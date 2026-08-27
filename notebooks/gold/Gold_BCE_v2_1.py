# Databricks notebook source

# =============================================================================
# CELDA 1 — IMPORTACIONES, CONFIGURACIÓN, VERSIONADO Y UTILIDADES
# =============================================================================

import json
import hashlib
import time
import platform
import importlib.metadata as importlib_metadata

from pathlib import Path
from datetime import datetime, timezone
from typing import Any, Dict, List

%pip install openpyxl
import pandas as pd
import openpyxl

from pyspark.sql import DataFrame
from pyspark.sql import functions as F
from pyspark.sql import types as T
from pyspark.sql.window import Window


PIPELINE_VERSION = "2.1"
GOLD_MANIFEST_SCHEMA_VERSION = "1.0"
SUPPORTED_SILVER_VERSIONS = {"2.1.0"}

CATALOG = "camaronera_2026"
SILVER_SCHEMA = "plata"
GOLD_SCHEMA = "oro"

SOURCE_CODE = "BCE"
GOLD_METADATA_VOLUME = "metadatos_bce_gold"
DELIVERABLE_VOLUME = "entregables"

RUN_STARTED_AT_UTC = datetime.now(timezone.utc)
RUN_STARTED_MONOTONIC = time.perf_counter()
RUN_ID = RUN_STARTED_AT_UTC.strftime("%Y%m%dT%H%M%S%fZ")

SILVER_TABLE = f"{CATALOG}.{SILVER_SCHEMA}.bce_exportaciones_camaron_subpartida"
LEGACY_SILVER_TABLE = f"{CATALOG}.{SILVER_SCHEMA}.exportaciones_camaron"

# Todas las tablas Gold son source-specific para impedir colisiones con CNA.
GOLD_TABLES = {
    "dim_fecha": f"{CATALOG}.{GOLD_SCHEMA}.bce_dim_fecha",
    "dim_producto": f"{CATALOG}.{GOLD_SCHEMA}.bce_dim_producto",
    "dim_subpartida": f"{CATALOG}.{GOLD_SCHEMA}.bce_dim_subpartida",
    "fact_exportaciones": f"{CATALOG}.{GOLD_SCHEMA}.bce_fact_exportaciones_subpartida",
    "resumen_mensual": f"{CATALOG}.{GOLD_SCHEMA}.bce_resumen_mensual",
    "resumen_anual": f"{CATALOG}.{GOLD_SCHEMA}.bce_resumen_anual",
    "indicadores_kpi": f"{CATALOG}.{GOLD_SCHEMA}.bce_indicadores_kpi",
    "serving_tableau": f"{CATALOG}.{GOLD_SCHEMA}.bce_serving_tableau",
}

GOLD_TABLE_ROLES = {
    "dim_fecha": "dimension",
    "dim_producto": "dimension",
    "dim_subpartida": "dimension",
    "fact_exportaciones": "fact",
    "resumen_mensual": "analytic",
    "resumen_anual": "analytic",
    "indicadores_kpi": "analytic",
    "serving_tableau": "serving",
}

# ÚNICA hoja que se permite exportar al Excel final.
SERVING_EXPORTS = {
    "tabla_tableau_bce": GOLD_TABLES["serving_tableau"],
}

GOLD_METADATA_ROOT = Path(f"/Volumes/{CATALOG}/{GOLD_SCHEMA}/{GOLD_METADATA_VOLUME}")
DELIVERABLE_ROOT = Path(f"/Volumes/{CATALOG}/{GOLD_SCHEMA}/{DELIVERABLE_VOLUME}")
EXCEL_PATH = DELIVERABLE_ROOT / "Exportaciones_BCE_Gold_Tableau.xlsx"

# Tolerancias numéricas.
ABS_USD_CONVERSION_TOL = 1e-6
ABS_PRICE_TOL = 1e-9
ABS_PARTICIPATION_TOL = 1e-8

# Mantenimiento Delta solo con justificación física.
OPTIMIZE_MIN_BYTES = 256 * 1024 * 1024
OPTIMIZE_MIN_FILES = 8


# Etiquetas cortas de PRESENTACIÓN para Tableau.
# La clave es el código arancelario institucional, más estable que la descripción textual.
# Silver conserva la descripción oficial; Gold no la destruye.
SUBPARTIDA_CORTA_POR_CODIGO = {
    "0306160000": "Camarón de agua fría",
    "0306171100": "Camarón entero",
    "0306171200": "Camarón cola sin caparazón",
    "0306171300": "Camarón cola crudo",
    "0306171400": "Camarón cola cocido",
    "0306171900": "Otros camarones",
    "0306179100": "Camarón de río",
    "0306179900": "Otros camarones",
    "0306179910": "Camarón pomada amarillo",
    "0306179920": "Camarón pomada negro",
    "0306179990": "Otros camarones",
    "0306350020": "Camarón fresco/refrigerado",
    "0306361900": "Otros camarones",
    "0306390090": "Otros camarones",
}


def subpartida_corta_expr(code_col: str = "codigo_subpartida"):
    """Devuelve la etiqueta corta configurada para el código de subpartida."""
    items = []
    for codigo, etiqueta in SUBPARTIDA_CORTA_POR_CODIGO.items():
        items.extend([F.lit(codigo), F.lit(etiqueta)])
    return F.create_map(*items)[F.col(code_col)]


def titulo(texto: str) -> None:
    print("\n" + "=" * 100)
    print(texto)
    print("=" * 100)


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def atomic_json_dump(payload: Dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    tmp.replace(path)


def sql_escape(value: Any) -> str:
    return str(value).replace("'", "''")


def table_properties(table_ref: str) -> Dict[str, str]:
    rows = spark.sql(f"SHOW TBLPROPERTIES {table_ref}").collect()
    return {str(r[0]): str(r[1]) for r in rows}


def software_versions() -> Dict[str, str]:
    import os
    import platform
    versions = {
        "python": platform.python_version(),
        "spark": getattr(spark, "version", "unknown"),
        "databricks_runtime": os.getenv("DATABRICKS_RUNTIME_VERSION", "unknown"),
    }
    for pkg in ["pyspark", "pandas", "openpyxl", "delta-spark", "playwright"]:
        try:
            versions[pkg] = importlib_metadata.version(pkg)
        except Exception:
            versions[pkg] = "not_installed"
    try:
        versions["spark_databricks_version"] = spark.conf.get(
            "spark.databricks.clusterUsageTags.sparkVersion"
        )
    except Exception:
        versions["spark_databricks_version"] = "unavailable"
    return versions

spark.sql(f"CREATE SCHEMA IF NOT EXISTS `{CATALOG}`.`{GOLD_SCHEMA}`")
spark.sql(f"CREATE VOLUME IF NOT EXISTS `{CATALOG}`.`{GOLD_SCHEMA}`.`{GOLD_METADATA_VOLUME}`")
spark.sql(f"CREATE VOLUME IF NOT EXISTS `{CATALOG}`.`{GOLD_SCHEMA}`.`{DELIVERABLE_VOLUME}`")

CONFIG_PAYLOAD = {
    "pipeline_version": PIPELINE_VERSION,
    "silver_table": SILVER_TABLE,
    "gold_tables": GOLD_TABLES,
    "serving_exports": SERVING_EXPORTS,
    "fact_grain": "fecha_periodo x codigo_producto_principal x codigo_subpartida",
    "fob_source_unit": "thousand_USD",
    "fob_canonical_unit": "USD",
    "price_formula": "fob_usd / (toneladas_metricas_peso_neto * 1000)",
    "tableau_contract": "one_flat_sheet_no_relationships",
    "subpartida_presentation_mapping": SUBPARTIDA_CORTA_POR_CODIGO,
}
CONFIG_HASH = sha256_text(json.dumps(CONFIG_PAYLOAD, sort_keys=True, ensure_ascii=False))

titulo("CONFIGURACIÓN GOLD BCE v2")
print(f"Run ID             : {RUN_ID}")
print(f"Pipeline version   : {PIPELINE_VERSION}")
print(f"Catálogo           : {CATALOG}")
print(f"Silver origen      : {SILVER_TABLE}")
print(f"Gold schema        : {GOLD_SCHEMA}")
print(f"Config SHA-256     : {CONFIG_HASH}")
print(f"Entregable final   : {EXCEL_PATH}")


# COMMAND ----------

# =============================================================================
# CELDA 2 — CONTRATO SILVER→GOLD Y VALIDACIÓN DE LINEAGE
# =============================================================================

EXPECTED_SILVER_COLUMNS = [
    "periodo",
    "fecha_periodo",
    "anio",
    "mes_numero",
    "codigo_nivel3",
    "nivel3",
    "codigo_producto_principal",
    "producto_principal",
    "codigo_subpartida",
    "subpartida",
    "toneladas_metricas_peso_neto",
    "fob_miles_usd",
    "fob_usd",
]


titulo("CONTRATO SILVER → GOLD BCE")

if not spark.catalog.tableExists(SILVER_TABLE):
    raise RuntimeError(
        f"No existe la Silver BCE v2 requerida: {SILVER_TABLE}. "
        "Ejecute primero Silver_BCE_v2_1_0."
    )

if spark.catalog.tableExists(LEGACY_SILVER_TABLE):
    print(
        f"WARN | Existe la tabla heredada {LEGACY_SILVER_TABLE}, pero este Gold NO la utilizará."
    )

DF_SILVER = spark.table(SILVER_TABLE)

if DF_SILVER.columns != EXPECTED_SILVER_COLUMNS:
    raise RuntimeError(
        "Contrato de columnas Silver incompatible.\n"
        f"Esperado: {EXPECTED_SILVER_COLUMNS}\n"
        f"Obtenido: {DF_SILVER.columns}"
    )

SILVER_PROPS = table_properties(SILVER_TABLE)

required_props = {
    "pipeline.layer": "silver",
    "pipeline.source": SOURCE_CODE,
}

for key, expected_value in required_props.items():
    observed = SILVER_PROPS.get(key)
    if observed != expected_value:
        raise RuntimeError(
            f"Propiedad Silver inválida {key}: esperado={expected_value!r}, obtenido={observed!r}"
        )

SILVER_VERSION = SILVER_PROPS.get("pipeline.version", "")
if SILVER_VERSION not in SUPPORTED_SILVER_VERSIONS:
    raise RuntimeError(
        f"Versión Silver no soportada: {SILVER_VERSION!r}. "
        f"Soportadas={sorted(SUPPORTED_SILVER_VERSIONS)}"
    )

SILVER_RUN_ID = SILVER_PROPS.get("pipeline.silver_run_id", "").strip()
BRONZE_SHA256 = SILVER_PROPS.get("lineage.bronze_sha256", "").strip()
BRONZE_BATCH_ID = SILVER_PROPS.get("lineage.bronze_batch_id", "").strip()
REPORT_TITLE = SILVER_PROPS.get("source.report", "").strip()
PERIOD_MIN_PROP = SILVER_PROPS.get("source.period_min", "").strip()
PERIOD_MAX_PROP = SILVER_PROPS.get("source.period_max", "").strip()
SOURCE_FOB_UNIT = SILVER_PROPS.get("source.fob_unit", "").strip()
CANONICAL_FOB_UNIT = SILVER_PROPS.get("canonical.fob_unit", "").strip()

if not SILVER_RUN_ID:
    raise RuntimeError("Silver no declara pipeline.silver_run_id.")
if not BRONZE_SHA256:
    raise RuntimeError("Silver no declara lineage.bronze_sha256.")
if SOURCE_FOB_UNIT != "thousand_USD":
    raise RuntimeError(f"Unidad FOB de origen inesperada: {SOURCE_FOB_UNIT!r}")
if CANONICAL_FOB_UNIT != "USD":
    raise RuntimeError(f"Unidad FOB canónica inesperada: {CANONICAL_FOB_UNIT!r}")

SILVER_ROW_COUNT = DF_SILVER.count()
if SILVER_ROW_COUNT <= 0:
    raise RuntimeError("La tabla Silver BCE está vacía.")

SILVER_LINEAGE = {
    "silver_table": SILVER_TABLE,
    "silver_version": SILVER_VERSION,
    "silver_run_id": SILVER_RUN_ID,
    "bronze_batch_id": BRONZE_BATCH_ID,
    "bronze_sha256": BRONZE_SHA256,
    "report_title": REPORT_TITLE,
    "period_min_property": PERIOD_MIN_PROP,
    "period_max_property": PERIOD_MAX_PROP,
    "source_fob_unit": SOURCE_FOB_UNIT,
    "canonical_fob_unit": CANONICAL_FOB_UNIT,
}

print(f"PASS | Silver rows        : {SILVER_ROW_COUNT}")
print(f"PASS | Silver version     : {SILVER_VERSION}")
print(f"PASS | Silver run ID      : {SILVER_RUN_ID}")
print(f"PASS | Bronze SHA-256     : {BRONZE_SHA256}")
print(f"PASS | FOB origen         : {SOURCE_FOB_UNIT}")
print(f"PASS | FOB canónico       : {CANONICAL_FOB_UNIT}")
print("PASS | El Gold no depende de la tabla Silver heredada.")


# COMMAND ----------

# =============================================================================
# CELDA 3 — PERFILADO DE SILVER SIN CONTEOS RÍGIDOS
# =============================================================================

titulo("PERFILADO SILVER BCE")

profile_row = (
    DF_SILVER
    .agg(
        F.count(F.lit(1)).alias("filas"),
        F.countDistinct("fecha_periodo").alias("periodos"),
        F.min("fecha_periodo").alias("periodo_min"),
        F.max("fecha_periodo").alias("periodo_max"),
        F.countDistinct("codigo_producto_principal").alias("productos"),
        F.countDistinct("codigo_subpartida").alias("subpartidas"),
        F.sum("toneladas_metricas_peso_neto").alias("total_tm"),
        F.sum("fob_miles_usd").alias("total_fob_miles_usd"),
        F.sum("fob_usd").alias("total_fob_usd"),
    )
    .first()
    .asDict()
)

rows_per_period = (
    DF_SILVER
    .groupBy("fecha_periodo")
    .agg(F.count(F.lit(1)).alias("filas"))
)
period_density = rows_per_period.agg(
    F.min("filas").alias("min_filas_periodo"),
    F.max("filas").alias("max_filas_periodo"),
    F.avg("filas").alias("promedio_filas_periodo"),
).first().asDict()

SILVER_PROFILE = {**profile_row, **period_density}

actual_period_min_iso = SILVER_PROFILE["periodo_min"].isoformat()
actual_period_max_iso = SILVER_PROFILE["periodo_max"].isoformat()

if PERIOD_MIN_PROP and PERIOD_MIN_PROP != actual_period_min_iso:
    raise RuntimeError(
        f"Lineage temporal inconsistente: source.period_min={PERIOD_MIN_PROP}, "
        f"datos={actual_period_min_iso}"
    )
if PERIOD_MAX_PROP and PERIOD_MAX_PROP != actual_period_max_iso:
    raise RuntimeError(
        f"Lineage temporal inconsistente: source.period_max={PERIOD_MAX_PROP}, "
        f"datos={actual_period_max_iso}"
    )

for key, value in SILVER_PROFILE.items():
    print(f"{key:28s}: {value}")

print("\nNota: no se exige un conteo fijo como 534 filas ni 14 subpartidas por mes.")
print("Gold trabaja con los períodos y registros realmente materializados por Silver v2.")


# COMMAND ----------

# =============================================================================
# CELDA 4 — BASE ANALÍTICA, TIEMPO Y MÉTRICAS DERIVADAS A NIVEL SUBPARTIDA
# =============================================================================

MESES_ES = F.create_map(
    *sum(
        ([F.lit(i), F.lit(nombre)] for i, nombre in enumerate([
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
        ], start=1)),
        [],
    )
)

base = (
    DF_SILVER
    .withColumn("fecha_key", F.date_format("fecha_periodo", "yyyyMMdd").cast("int"))
    .withColumn("mes_nombre", MESES_ES[F.col("mes_numero")])
    .withColumn("anio_mes", F.date_format("fecha_periodo", "yyyy-MM"))
    .withColumn("trimestre", F.quarter("fecha_periodo"))
    .withColumn(
        "semestre",
        F.when(F.col("mes_numero") <= 6, F.lit(1)).otherwise(F.lit(2)),
    )
    .withColumn(
        "precio_subpartida_usd_kg",
        F.when(
            F.col("toneladas_metricas_peso_neto") > 0,
            F.col("fob_usd") / (F.col("toneladas_metricas_peso_neto") * F.lit(1000.0)),
        ).otherwise(F.lit(None).cast("double")),
    )
)

# YoY por la misma subpartida y el mismo mes calendario.
window_yoy = Window.partitionBy("codigo_subpartida", "mes_numero").orderBy("anio")

DF_ENRICHED = (
    base
    .withColumn("anio_prev", F.lag("anio").over(window_yoy))
    .withColumn(
        "tm_prev_anio_mismo_mes_subpartida",
        F.lag("toneladas_metricas_peso_neto").over(window_yoy),
    )
    .withColumn(
        "fob_usd_prev_anio_mismo_mes_subpartida",
        F.lag("fob_usd").over(window_yoy),
    )
    .withColumn(
        "variacion_yoy_tm_subpartida",
        F.when(
            (F.col("anio_prev") == F.col("anio") - 1)
            & (F.col("tm_prev_anio_mismo_mes_subpartida") > 0),
            F.col("toneladas_metricas_peso_neto")
            / F.col("tm_prev_anio_mismo_mes_subpartida")
            - F.lit(1.0),
        ),
    )
    .withColumn(
        "variacion_yoy_fob_subpartida",
        F.when(
            (F.col("anio_prev") == F.col("anio") - 1)
            & (F.col("fob_usd_prev_anio_mismo_mes_subpartida") > 0),
            F.col("fob_usd")
            / F.col("fob_usd_prev_anio_mismo_mes_subpartida")
            - F.lit(1.0),
        ),
    )
)

titulo("BASE ANALÍTICA BCE")
print(f"Filas enriquecidas : {DF_ENRICHED.count()}")
print("Grano               : fecha × producto principal × subpartida")
DF_ENRICHED.select(
    "fecha_periodo",
    "codigo_subpartida",
    "toneladas_metricas_peso_neto",
    "fob_usd",
    "precio_subpartida_usd_kg",
    "variacion_yoy_fob_subpartida",
).orderBy("fecha_periodo", "codigo_subpartida").show(10, truncate=False)


# COMMAND ----------

# =============================================================================
# CELDA 5 — DIMENSIONES GOLD CON CLAVES DETERMINÍSTICAS
# =============================================================================

# Validar cobertura del catálogo de etiquetas cortas antes de construir Gold.
codigos_subpartida_actuales = [
    r["codigo_subpartida"]
    for r in DF_ENRICHED.select("codigo_subpartida").distinct().collect()
]

codigos_sin_etiqueta = sorted(
    codigo for codigo in codigos_subpartida_actuales
    if codigo not in SUBPARTIDA_CORTA_POR_CODIGO
)

codigos_configurados_no_presentes = sorted(
    codigo for codigo in SUBPARTIDA_CORTA_POR_CODIGO
    if codigo not in set(codigos_subpartida_actuales)
)

MAPPING_AUDIT = {
    "codigos_actuales": len(codigos_subpartida_actuales),
    "codigos_mapeados": len(codigos_subpartida_actuales) - len(codigos_sin_etiqueta),
    "codigos_sin_etiqueta": codigos_sin_etiqueta,
    "codigos_configurados_no_presentes": codigos_configurados_no_presentes,
}

if codigos_sin_etiqueta:
    raise RuntimeError(
        "Existen códigos de subpartida sin etiqueta corta de presentación: "
        f"{codigos_sin_etiqueta}. Actualice SUBPARTIDA_CORTA_POR_CODIGO antes de publicar Tableau."
    )

# Antes de deduplicar dimensiones se comprueba que una misma clave institucional
# no tenga descripciones/hierarquías contradictorias.
producto_conflicts = (
    DF_ENRICHED
    .groupBy("codigo_producto_principal")
    .agg(
        F.countDistinct("producto_principal").alias("n_producto"),
        F.countDistinct("codigo_nivel3").alias("n_nivel3_codigo"),
        F.countDistinct("nivel3").alias("n_nivel3"),
    )
    .filter(
        (F.col("n_producto") > 1)
        | (F.col("n_nivel3_codigo") > 1)
        | (F.col("n_nivel3") > 1)
    )
    .count()
)

subpartida_conflicts = (
    DF_ENRICHED
    .groupBy("codigo_subpartida")
    .agg(
        F.countDistinct("subpartida").alias("n_descripciones"),
        F.countDistinct("codigo_producto_principal").alias("n_productos"),
    )
    .filter(
        (F.col("n_descripciones") > 1)
        | (F.col("n_productos") > 1)
    )
    .count()
)

if producto_conflicts > 0 or subpartida_conflicts > 0:
    raise RuntimeError(
        "Conflicto semántico al construir dimensiones: "
        f"producto_conflicts={producto_conflicts}, "
        f"subpartida_conflicts={subpartida_conflicts}"
    )

DF_DIM_FECHA = (
    DF_ENRICHED
    .select(
        "fecha_key",
        F.col("fecha_periodo").alias("fecha"),
        "anio",
        "mes_numero",
        "mes_nombre",
        "anio_mes",
        "trimestre",
        "semestre",
    )
    .dropDuplicates(["fecha_key"])
    .orderBy("fecha_key")
)

# El código institucional se utiliza como llave estable: no depende del orden físico de Spark.
DF_DIM_PRODUCTO = (
    DF_ENRICHED
    .select(
        F.col("codigo_producto_principal").alias("producto_key"),
        "codigo_nivel3",
        "nivel3",
        F.col("codigo_producto_principal"),
        "producto_principal",
    )
    .dropDuplicates(["producto_key"])
    .orderBy("producto_key")
)

DF_DIM_SUBPARTIDA = (
    DF_ENRICHED
    .withColumn("subpartida_corta", subpartida_corta_expr("codigo_subpartida"))
    .select(
        F.col("codigo_subpartida").alias("subpartida_key"),
        "codigo_subpartida",
        F.col("subpartida").alias("subpartida_oficial"),
        "subpartida_corta",
        "codigo_producto_principal",
    )
    .dropDuplicates(["subpartida_key"])
    .orderBy("subpartida_key")
)

DIMENSION_CHECKS = {
    "dim_fecha_rows": DF_DIM_FECHA.count(),
    "dim_producto_rows": DF_DIM_PRODUCTO.count(),
    "dim_subpartida_rows": DF_DIM_SUBPARTIDA.count(),
    "producto_semantic_conflicts": producto_conflicts,
    "subpartida_semantic_conflicts": subpartida_conflicts,
    "dim_fecha_duplicate_keys": DF_DIM_FECHA.groupBy("fecha_key").count().filter(F.col("count") > 1).count(),
    "dim_producto_duplicate_keys": DF_DIM_PRODUCTO.groupBy("producto_key").count().filter(F.col("count") > 1).count(),
    "dim_subpartida_duplicate_keys": DF_DIM_SUBPARTIDA.groupBy("subpartida_key").count().filter(F.col("count") > 1).count(),
}

if any(DIMENSION_CHECKS[k] > 0 for k in [
    "dim_fecha_duplicate_keys",
    "dim_producto_duplicate_keys",
    "dim_subpartida_duplicate_keys",
]):
    raise RuntimeError(f"Se detectaron claves duplicadas en dimensiones: {DIMENSION_CHECKS}")

titulo("DIMENSIONES GOLD BCE")
for key, value in DIMENSION_CHECKS.items():
    print(f"{key:32s}: {value}")

print("\nMapeo de presentación de subpartidas:")
print(f"  Códigos actuales       : {MAPPING_AUDIT['codigos_actuales']}")
print(f"  Códigos mapeados       : {MAPPING_AUDIT['codigos_mapeados']}")
print(f"  Códigos sin etiqueta   : {MAPPING_AUDIT['codigos_sin_etiqueta']}")


# COMMAND ----------

# =============================================================================
# CELDA 6 — TABLA DE HECHOS Y CONSERVACIÓN DE MEDIDAS SILVER→GOLD
# =============================================================================

DF_FACT_EXPORTACIONES = (
    DF_ENRICHED
    .select(
        "fecha_key",
        F.col("codigo_producto_principal").alias("producto_key"),
        F.col("codigo_subpartida").alias("subpartida_key"),
        "toneladas_metricas_peso_neto",
        "fob_miles_usd",
        "fob_usd",
        "precio_subpartida_usd_kg",
    )
    .orderBy("fecha_key", "subpartida_key")
)

FACT_GRAIN = ["fecha_key", "producto_key", "subpartida_key"]

fact_duplicate_groups = (
    DF_FACT_EXPORTACIONES
    .groupBy(*FACT_GRAIN)
    .count()
    .filter(F.col("count") > 1)
    .count()
)

silver_totals = DF_SILVER.agg(
    F.sum("toneladas_metricas_peso_neto").alias("tm"),
    F.sum("fob_miles_usd").alias("fob_miles"),
    F.sum("fob_usd").alias("fob_usd"),
).first().asDict()

fact_totals = DF_FACT_EXPORTACIONES.agg(
    F.sum("toneladas_metricas_peso_neto").alias("tm"),
    F.sum("fob_miles_usd").alias("fob_miles"),
    F.sum("fob_usd").alias("fob_usd"),
).first().asDict()

FACT_CHECKS = {
    "silver_rows": SILVER_ROW_COUNT,
    "fact_rows": DF_FACT_EXPORTACIONES.count(),
    "duplicate_grain_groups": fact_duplicate_groups,
    "tm_difference": abs(float(silver_totals["tm"] or 0) - float(fact_totals["tm"] or 0)),
    "fob_miles_difference": abs(float(silver_totals["fob_miles"] or 0) - float(fact_totals["fob_miles"] or 0)),
    "fob_usd_difference": abs(float(silver_totals["fob_usd"] or 0) - float(fact_totals["fob_usd"] or 0)),
}

if FACT_CHECKS["silver_rows"] != FACT_CHECKS["fact_rows"]:
    raise RuntimeError(f"Pérdida o duplicación de filas Silver→Fact: {FACT_CHECKS}")
if FACT_CHECKS["duplicate_grain_groups"] != 0:
    raise RuntimeError(f"La fact viola su grano: {FACT_CHECKS}")
if (
    FACT_CHECKS["tm_difference"] > 1e-6
    or FACT_CHECKS["fob_miles_difference"] > 1e-6
    or FACT_CHECKS["fob_usd_difference"] > 1e-3
):
    raise RuntimeError(f"Las medidas no se conservaron Silver→Fact: {FACT_CHECKS}")

titulo("FACT EXPORTACIONES BCE")
for key, value in FACT_CHECKS.items():
    print(f"{key:28s}: {value}")
print(f"Grano fact: {FACT_GRAIN}")


# COMMAND ----------

# =============================================================================
# CELDA 7 — RESÚMENES ANALÍTICOS MENSUAL Y ANUAL
# =============================================================================

DF_RESUMEN_MENSUAL = (
    DF_ENRICHED
    .groupBy("fecha_key", "fecha_periodo", "anio", "mes_numero", "mes_nombre", "anio_mes", "trimestre", "semestre")
    .agg(
        F.sum("toneladas_metricas_peso_neto").alias("toneladas_metricas_totales"),
        F.sum("fob_miles_usd").alias("fob_total_miles_usd"),
        F.sum("fob_usd").alias("fob_total_usd"),
        F.countDistinct("codigo_subpartida").alias("subpartidas_reportadas"),
    )
    .withColumn(
        "precio_promedio_ponderado_usd_kg",
        F.when(
            F.col("toneladas_metricas_totales") > 0,
            F.col("fob_total_usd") / (F.col("toneladas_metricas_totales") * F.lit(1000.0)),
        ),
    )
    .orderBy("fecha_key")
)

window_anio = Window.orderBy("anio")

DF_RESUMEN_ANUAL_BASE = (
    DF_RESUMEN_MENSUAL
    .groupBy("anio")
    .agg(
        F.sum("toneladas_metricas_totales").alias("toneladas_metricas_totales"),
        F.sum("fob_total_miles_usd").alias("fob_total_miles_usd"),
        F.sum("fob_total_usd").alias("fob_total_usd"),
        F.countDistinct("mes_numero").alias("meses_reportados"),
    )
    .withColumn("anio_completo", F.col("meses_reportados") == F.lit(12))
    .withColumn(
        "precio_promedio_ponderado_usd_kg",
        F.when(
            F.col("toneladas_metricas_totales") > 0,
            F.col("fob_total_usd") / (F.col("toneladas_metricas_totales") * F.lit(1000.0)),
        ),
    )
)

DF_RESUMEN_ANUAL = (
    DF_RESUMEN_ANUAL_BASE
    .withColumn("anio_prev", F.lag("anio").over(window_anio))
    .withColumn("anio_completo_prev", F.lag("anio_completo").over(window_anio))
    .withColumn("tm_prev", F.lag("toneladas_metricas_totales").over(window_anio))
    .withColumn("fob_prev", F.lag("fob_total_usd").over(window_anio))
    .withColumn(
        "variacion_yoy_tm_anual",
        F.when(
            (F.col("anio_prev") == F.col("anio") - 1)
            & F.col("anio_completo")
            & F.col("anio_completo_prev")
            & (F.col("tm_prev") > 0),
            F.col("toneladas_metricas_totales") / F.col("tm_prev") - F.lit(1.0),
        ),
    )
    .withColumn(
        "variacion_yoy_fob_anual",
        F.when(
            (F.col("anio_prev") == F.col("anio") - 1)
            & F.col("anio_completo")
            & F.col("anio_completo_prev")
            & (F.col("fob_prev") > 0),
            F.col("fob_total_usd") / F.col("fob_prev") - F.lit(1.0),
        ),
    )
    .drop("anio_prev", "anio_completo_prev", "tm_prev", "fob_prev")
    .orderBy("anio")
)

titulo("RESÚMENES ANALÍTICOS BCE")
print(f"Resumen mensual : {DF_RESUMEN_MENSUAL.count()} filas")
print(f"Resumen anual   : {DF_RESUMEN_ANUAL.count()} filas")
print("El YoY anual solo se calcula cuando ambos años contienen 12 meses.")
DF_RESUMEN_ANUAL.show(truncate=False)


# COMMAND ----------

# =============================================================================
# CELDA 8 — SERVING PLANO ÚNICO PARA TABLEAU
# =============================================================================

# Ventanas mensuales para participación y ranking dentro de cada período.
window_period = Window.partitionBy("fecha_periodo")
window_rank_fob = Window.partitionBy("fecha_periodo").orderBy(
    F.col("fob_usd").desc_nulls_last(),
    F.col("codigo_subpartida").asc(),
)
window_rank_tm = Window.partitionBy("fecha_periodo").orderBy(
    F.col("toneladas_metricas_peso_neto").desc_nulls_last(),
    F.col("codigo_subpartida").asc(),
)

annual_coverage = (
    DF_RESUMEN_ANUAL
    .select(
        "anio",
        "meses_reportados",
        "anio_completo",
    )
    .withColumn(
        "cobertura_anio",
        F.when(F.col("anio_completo"), F.lit("COMPLETO")).otherwise(F.lit("PARCIAL")),
    )
)

max_period = DF_ENRICHED.agg(F.max("fecha_periodo").alias("max_period")).first()["max_period"]

serving = (
    DF_ENRICHED
    .withColumn("subpartida_corta", subpartida_corta_expr("codigo_subpartida"))
    .withColumn("total_tm_mes_tmp", F.sum("toneladas_metricas_peso_neto").over(window_period))
    .withColumn("total_fob_mes_tmp", F.sum("fob_usd").over(window_period))
    .withColumn(
        "participacion_tm_mes",
        F.when(
            F.col("total_tm_mes_tmp") > 0,
            F.col("toneladas_metricas_peso_neto") / F.col("total_tm_mes_tmp"),
        ),
    )
    .withColumn(
        "participacion_fob_mes",
        F.when(
            F.col("total_fob_mes_tmp") > 0,
            F.col("fob_usd") / F.col("total_fob_mes_tmp"),
        ),
    )
    .withColumn("ranking_tm_mes", F.row_number().over(window_rank_tm))
    .withColumn("ranking_fob_mes", F.row_number().over(window_rank_fob))
    .drop("total_tm_mes_tmp", "total_fob_mes_tmp", "anio_prev")
    .join(annual_coverage, on="anio", how="left")
    .withColumn("es_ultimo_periodo_disponible", F.col("fecha_periodo") == F.lit(max_period))
)

# Una sola tabla desnormalizada. No incluye agregados mensuales repetidos que puedan
# inflar SUM() en Tableau. La columna `subpartida` usa la ETIQUETA CORTA de presentación;
# la descripción oficial permanece preservada en Silver y en bce_dim_subpartida.
# Los totales nacionales se obtienen sumando TM/FOB al nivel de fecha; el precio nacional
# ponderado debe calcularse como SUM(FOB USD)/(SUM(TM)*1000).
DF_SERVING_TABLEAU = (
    serving
    .select(
        "fecha_periodo",
        "periodo",
        "anio",
        "mes_numero",
        "mes_nombre",
        "anio_mes",
        "trimestre",
        "semestre",
        "meses_reportados",
        "cobertura_anio",
        "es_ultimo_periodo_disponible",
        "codigo_nivel3",
        "nivel3",
        "codigo_producto_principal",
        "producto_principal",
        "codigo_subpartida",
        F.col("subpartida_corta").alias("subpartida"),
        "toneladas_metricas_peso_neto",
        "fob_miles_usd",
        "fob_usd",
        "precio_subpartida_usd_kg",
        "participacion_tm_mes",
        "participacion_fob_mes",
        "ranking_tm_mes",
        "ranking_fob_mes",
        "tm_prev_anio_mismo_mes_subpartida",
        "fob_usd_prev_anio_mismo_mes_subpartida",
        "variacion_yoy_tm_subpartida",
        "variacion_yoy_fob_subpartida",
    )
    .orderBy("fecha_periodo", "ranking_fob_mes", "codigo_subpartida")
)

SERVING_EXPORTS = {
    "tabla_tableau_bce": GOLD_TABLES["serving_tableau"],
}

titulo("SERVING PLANO PARA TABLEAU")
print(f"Filas     : {DF_SERVING_TABLEAU.count()}")
print(f"Columnas  : {len(DF_SERVING_TABLEAU.columns)}")
print("Hoja XLSX: tabla_tableau_bce")
print("Relaciones Tableau requeridas: NO")
print("Subpartida en XLSX          : ETIQUETA CORTA DE PRESENTACIÓN")
DF_SERVING_TABLEAU.show(10, truncate=False)


# COMMAND ----------

# =============================================================================
# CELDA 9 — INDICADORES KPI INTERNOS CON PERÍODOS COMPARABLES
# =============================================================================

latest_period = DF_RESUMEN_MENSUAL.agg(F.max("fecha_periodo")).first()[0]
previous_year_same_month = latest_period.replace(year=latest_period.year - 1)

latest_month_row = (
    DF_RESUMEN_MENSUAL
    .filter(F.col("fecha_periodo") == F.lit(latest_period))
    .first()
)
previous_year_row = (
    DF_RESUMEN_MENSUAL
    .filter(F.col("fecha_periodo") == F.lit(previous_year_same_month))
    .first()
)

top_subpartida_row = (
    DF_SERVING_TABLEAU
    .filter(F.col("fecha_periodo") == F.lit(latest_period))
    .orderBy("ranking_fob_mes")
    .select("codigo_subpartida", "subpartida", "fob_usd", "participacion_fob_mes")
    .first()
)

totals_all = DF_FACT_EXPORTACIONES.agg(
    F.sum("toneladas_metricas_peso_neto").alias("tm_total"),
    F.sum("fob_usd").alias("fob_usd_total"),
).first()

latest_tm = float(latest_month_row["toneladas_metricas_totales"] or 0.0)
latest_fob = float(latest_month_row["fob_total_usd"] or 0.0)

prev_tm = None if previous_year_row is None else previous_year_row["toneladas_metricas_totales"]
prev_fob = None if previous_year_row is None else previous_year_row["fob_total_usd"]

yoy_tm_latest = (
    float(latest_tm / float(prev_tm) - 1.0)
    if prev_tm is not None and float(prev_tm) > 0
    else None
)
yoy_fob_latest = (
    float(latest_fob / float(prev_fob) - 1.0)
    if prev_fob is not None and float(prev_fob) > 0
    else None
)

KPI_SCHEMA = T.StructType([
    T.StructField("periodo_min", T.DateType(), False),
    T.StructField("periodo_max", T.DateType(), False),
    T.StructField("periodos_disponibles", T.IntegerType(), False),
    T.StructField("subpartidas_disponibles", T.IntegerType(), False),
    T.StructField("toneladas_metricas_total_historico", T.DoubleType(), True),
    T.StructField("fob_usd_total_historico", T.DoubleType(), True),
    T.StructField("tm_ultimo_periodo", T.DoubleType(), True),
    T.StructField("fob_usd_ultimo_periodo", T.DoubleType(), True),
    T.StructField("variacion_yoy_tm_ultimo_mes_comparable", T.DoubleType(), True),
    T.StructField("variacion_yoy_fob_ultimo_mes_comparable", T.DoubleType(), True),
    T.StructField("codigo_subpartida_top_fob_ultimo_periodo", T.StringType(), True),
    T.StructField("subpartida_top_fob_ultimo_periodo", T.StringType(), True),
])

DF_INDICADORES_KPI = spark.createDataFrame([
    (
        SILVER_PROFILE["periodo_min"],
        SILVER_PROFILE["periodo_max"],
        int(SILVER_PROFILE["periodos"]),
        int(SILVER_PROFILE["subpartidas"]),
        float(totals_all["tm_total"] or 0.0),
        float(totals_all["fob_usd_total"] or 0.0),
        latest_tm,
        latest_fob,
        yoy_tm_latest,
        yoy_fob_latest,
        None if top_subpartida_row is None else str(top_subpartida_row["codigo_subpartida"]),
        None if top_subpartida_row is None else str(top_subpartida_row["subpartida"]),
    )
], schema=KPI_SCHEMA)

titulo("KPI INTERNOS BCE")
print(f"Último período             : {latest_period}")
print(f"Comparación interanual con : {previous_year_same_month}")
DF_INDICADORES_KPI.show(truncate=False)
print("Nota: la tabla KPI es interna Gold y NO se exporta como hoja adicional a Tableau.")


# COMMAND ----------

# =============================================================================
# CELDA 10 — REGLAS DE CALIDAD GOLD BLOQUEANTES Y ADVERTENCIAS
# =============================================================================

QUALITY_RESULTS: List[Dict[str, Any]] = []


def quality(rule: str, severity: str, violations: int, checked: int, detail: str) -> None:
    status = "PASS" if violations == 0 else ("FAIL" if severity == "CRITICAL" else "WARN")
    QUALITY_RESULTS.append({
        "rule": rule,
        "severity": severity,
        "status": status,
        "violations": int(violations),
        "checked": int(checked),
        "detail": detail,
    })
    print(
        f"{status:4s} | {rule:42s} | "
        f"violaciones={violations:4d} | revisados={checked:4d}"
    )


titulo("REGLAS DE CALIDAD GOLD BCE")

# Q1. Medidas no negativas.
q1_checked = DF_FACT_EXPORTACIONES.count()
q1_viol = DF_FACT_EXPORTACIONES.filter(
    (F.col("toneladas_metricas_peso_neto") < 0)
    | (F.col("fob_miles_usd") < 0)
    | (F.col("fob_usd") < 0)
).count()
quality("Q1_MEDIDAS_NO_NEGATIVAS", "CRITICAL", q1_viol, q1_checked, "TM y FOB deben ser >= 0.")

# Q2. Conversión explícita miles USD -> USD.
q2_viol = DF_FACT_EXPORTACIONES.filter(
    F.abs(F.col("fob_usd") - F.col("fob_miles_usd") * F.lit(1000.0))
    > F.lit(ABS_USD_CONVERSION_TOL)
).count()
quality("Q2_CONVERSION_FOB_USD", "CRITICAL", q2_viol, q1_checked, "fob_usd = fob_miles_usd * 1000.")

# Q3. Precio subpartida = USD / kg.
eligible_price = DF_FACT_EXPORTACIONES.filter(
    (F.col("toneladas_metricas_peso_neto") > 0)
    & F.col("precio_subpartida_usd_kg").isNotNull()
)
q3_viol = eligible_price.filter(
    F.abs(
        F.col("precio_subpartida_usd_kg")
        - F.col("fob_usd") / (F.col("toneladas_metricas_peso_neto") * F.lit(1000.0))
    ) > F.lit(ABS_PRICE_TOL)
).count()
quality("Q3_PRECIO_USD_KG", "CRITICAL", q3_viol, eligible_price.count(), "Precio unitario correcto por subpartida.")

# Q4. Grano único de la fact.
q4_viol = DF_FACT_EXPORTACIONES.groupBy(*FACT_GRAIN).count().filter(F.col("count") > 1).count()
quality("Q4_GRANO_FACT_UNICO", "CRITICAL", q4_viol, DF_FACT_EXPORTACIONES.count(), str(FACT_GRAIN))

# Q5-Q7. Integridad referencial.
q5_viol = DF_FACT_EXPORTACIONES.join(DF_DIM_FECHA.select("fecha_key"), "fecha_key", "left_anti").count()
quality("Q5_FK_FECHA", "CRITICAL", q5_viol, q1_checked, "Todas las facts deben tener fecha.")

q6_viol = DF_FACT_EXPORTACIONES.join(DF_DIM_PRODUCTO.select("producto_key"), "producto_key", "left_anti").count()
quality("Q6_FK_PRODUCTO", "CRITICAL", q6_viol, q1_checked, "Todas las facts deben tener producto.")

q7_viol = DF_FACT_EXPORTACIONES.join(DF_DIM_SUBPARTIDA.select("subpartida_key"), "subpartida_key", "left_anti").count()
quality("Q7_FK_SUBPARTIDA", "CRITICAL", q7_viol, q1_checked, "Todas las facts deben tener subpartida.")

# Q8. Participaciones en rango.
q8_checked = DF_SERVING_TABLEAU.filter(F.col("participacion_fob_mes").isNotNull()).count()
q8_viol = DF_SERVING_TABLEAU.filter(
    F.col("participacion_fob_mes").isNotNull()
    & (~F.col("participacion_fob_mes").between(0.0, 1.0))
).count()
quality("Q8_PARTICIPACION_RANGO", "CRITICAL", q8_viol, q8_checked, "Participación FOB debe estar en [0,1].")

# Q9. Participaciones mensuales deben sumar 1 cuando existe FOB positivo.
part_sums = (
    DF_SERVING_TABLEAU
    .groupBy("fecha_periodo")
    .agg(
        F.sum("participacion_fob_mes").alias("participacion_total"),
        F.sum("fob_usd").alias("fob_total"),
    )
    .filter(F.col("fob_total") > 0)
)
q9_viol = part_sums.filter(
    F.abs(F.col("participacion_total") - F.lit(1.0)) > F.lit(ABS_PARTICIPATION_TOL)
).count()
quality("Q9_PARTICIPACION_MENSUAL_TOTAL", "CRITICAL", q9_viol, part_sums.count(), "SUM(participación FOB)=1 por mes.")

# Q10. Serving conserva una fila por observación de la fact.
q10_viol = abs(DF_SERVING_TABLEAU.count() - DF_FACT_EXPORTACIONES.count())
quality("Q10_SERVING_CONSERVA_FILAS", "CRITICAL", q10_viol, DF_FACT_EXPORTACIONES.count(), "Serving y fact deben tener igual número de filas.")

# Q11. El contrato actual no admite dimensión país.
country_columns = [c for c in DF_SERVING_TABLEAU.columns if c.lower() in {"pais", "pais_destino", "id_pais", "pais_key"}]
q11_viol = len(country_columns)
quality("Q11_SIN_PAIS_INVENTADO", "CRITICAL", q11_viol, len(DF_SERVING_TABLEAU.columns), f"columnas país detectadas={country_columns}")

# Q12. Producto principal contractual del estudio.
q12_viol = DF_SERVING_TABLEAU.filter(F.col("codigo_producto_principal") != F.lit("140201")).count()
quality("Q12_PRODUCTO_CAMARON_140201", "CRITICAL", q12_viol, DF_SERVING_TABLEAU.count(), "Código PP esperado para camarones.")

# Q13. Ceros se documentan, no se eliminan silenciosamente.
zero_rows = DF_FACT_EXPORTACIONES.filter(
    (F.col("toneladas_metricas_peso_neto") == 0) | (F.col("fob_usd") == 0)
).count()
quality("Q13_VALORES_CERO", "WARNING", zero_rows, q1_checked, "Los ceros se conservan y se reportan como advertencia.")

# Q14. El YoY anual no debe existir para años parciales.
q14_viol = DF_RESUMEN_ANUAL.filter(
    (~F.col("anio_completo"))
    & (
        F.col("variacion_yoy_tm_anual").isNotNull()
        | F.col("variacion_yoy_fob_anual").isNotNull()
    )
).count()
quality("Q14_YOY_ANUAL_SOLO_COMPLETOS", "CRITICAL", q14_viol, DF_RESUMEN_ANUAL.count(), "No comparar año parcial con año completo.")


# Q15. Todas las subpartidas presentes deben tener etiqueta corta de presentación.
q15_checked = len(codigos_subpartida_actuales)
q15_viol = len(codigos_sin_etiqueta)
quality(
    "Q15_SUBPARTIDAS_CORTAS_MAPEADAS",
    "CRITICAL",
    q15_viol,
    q15_checked,
    "Cada código BCE presente debe tener una etiqueta corta explícita para Tableau.",
)

critical_failures = [
    x for x in QUALITY_RESULTS
    if x["severity"] == "CRITICAL" and x["status"] == "FAIL"
]
if critical_failures:
    raise RuntimeError(f"Gold BCE no supera las reglas críticas: {critical_failures}")

print("\n✓ Reglas críticas Gold BCE superadas.")
print("Las advertencias se registran; no se interpretan como 100% de calidad del dominio.")


# COMMAND ----------

# =============================================================================
# CELDA 11 — PERSISTENCIA DELTA Y LINEAGE SILVER→GOLD
# =============================================================================

GOLD_DATAFRAMES = {
    "dim_fecha": DF_DIM_FECHA,
    "dim_producto": DF_DIM_PRODUCTO,
    "dim_subpartida": DF_DIM_SUBPARTIDA,
    "fact_exportaciones": DF_FACT_EXPORTACIONES,
    "resumen_mensual": DF_RESUMEN_MENSUAL,
    "resumen_anual": DF_RESUMEN_ANUAL,
    "indicadores_kpi": DF_INDICADORES_KPI,
    "serving_tableau": DF_SERVING_TABLEAU,
}

PERSISTENCE_RESULTS: List[Dict[str, Any]] = []

titulo("PERSISTENCIA GOLD BCE EN UNITY CATALOG")

for logical_name, df in GOLD_DATAFRAMES.items():
    table_ref = GOLD_TABLES[logical_name]

    (
        df.write
        .format("delta")
        .mode("overwrite")
        .option("overwriteSchema", "true")
        .saveAsTable(table_ref)
    )

    props = {
        "pipeline.layer": "gold",
        "pipeline.source": SOURCE_CODE,
        "pipeline.version": PIPELINE_VERSION,
        "pipeline.gold_run_id": RUN_ID,
        "lineage.silver_table": SILVER_TABLE,
        "lineage.silver_run_id": SILVER_RUN_ID,
        "lineage.silver_version": SILVER_VERSION,
        "lineage.bronze_sha256": BRONZE_SHA256,
        "source.period_min": str(SILVER_PROFILE["periodo_min"]),
        "source.period_max": str(SILVER_PROFILE["periodo_max"]),
        "source.fob_unit": SOURCE_FOB_UNIT,
        "canonical.fob_unit": CANONICAL_FOB_UNIT,
        "gold.table_role": GOLD_TABLE_ROLES[logical_name],
        "gold.config_sha256": CONFIG_HASH,
        "gold.fact_grain": CONFIG_PAYLOAD["fact_grain"],
    }
    props_sql = ", ".join(
        f"'{sql_escape(k)}'='{sql_escape(v)}'"
        for k, v in props.items()
    )
    spark.sql(f"ALTER TABLE {table_ref} SET TBLPROPERTIES ({props_sql})")

    rows = spark.table(table_ref).count()
    PERSISTENCE_RESULTS.append({
        "logical_name": logical_name,
        "table": table_ref,
        "role": GOLD_TABLE_ROLES[logical_name],
        "rows": rows,
        "status": "WRITTEN",
    })
    print(f"PASS | {logical_name:24s} | {rows:5d} filas | {table_ref}")

print(f"\nTablas Gold persistidas: {len(PERSISTENCE_RESULTS)}")


# COMMAND ----------

# =============================================================================
# CELDA 12 — VALIDACIÓN POST-ESCRITURA Y QA TÉCNICO
# =============================================================================

POSTWRITE_RESULTS: List[Dict[str, Any]] = []
QA_RESULTS: List[Dict[str, Any]] = []

titulo("VALIDACIÓN POST-ESCRITURA GOLD BCE")

for logical_name, source_df in GOLD_DATAFRAMES.items():
    table_ref = GOLD_TABLES[logical_name]
    persisted_df = spark.table(table_ref)

    source_count = source_df.count()
    persisted_count = persisted_df.count()

    same_columns = source_df.columns == persisted_df.columns
    same_schema = (
        [f.dataType.simpleString() for f in source_df.schema.fields]
        == [f.dataType.simpleString() for f in persisted_df.schema.fields]
    )

    source_minus = source_df.exceptAll(persisted_df).limit(1).count()
    persisted_minus = persisted_df.exceptAll(source_df).limit(1).count()
    same_data = source_minus == 0 and persisted_minus == 0

    detail = spark.sql(f"DESCRIBE DETAIL {table_ref}").first().asDict()
    is_delta = str(detail.get("format", "")).lower() == "delta"

    props = table_properties(table_ref)
    lineage_ok = (
        props.get("pipeline.layer") == "gold"
        and props.get("pipeline.source") == SOURCE_CODE
        and props.get("pipeline.gold_run_id") == RUN_ID
        and props.get("lineage.silver_run_id") == SILVER_RUN_ID
        and props.get("lineage.bronze_sha256") == BRONZE_SHA256
    )

    passed = (
        source_count == persisted_count
        and same_columns
        and same_schema
        and same_data
        and is_delta
        and lineage_ok
    )

    POSTWRITE_RESULTS.append({
        "logical_name": logical_name,
        "source_rows": source_count,
        "persisted_rows": persisted_count,
        "same_columns": same_columns,
        "same_schema": same_schema,
        "same_data": same_data,
        "is_delta": is_delta,
        "lineage_ok": lineage_ok,
        "status": "PASS" if passed else "FAIL",
    })

    print(
        f"{'PASS' if passed else 'FAIL'} | {logical_name:24s} | "
        f"filas={persisted_count:5d} | data_equal={same_data} | lineage={lineage_ok}"
    )

postwrite_failures = [x for x in POSTWRITE_RESULTS if x["status"] == "FAIL"]
if postwrite_failures:
    raise RuntimeError(f"Validación post-escritura fallida: {postwrite_failures}")


def qa(name: str, passed: bool, detail: str) -> None:
    QA_RESULTS.append({
        "test": name,
        "status": "PASS" if passed else "FAIL",
        "detail": detail,
    })


for logical_name, table_ref in GOLD_TABLES.items():
    qa(f"access_{logical_name}", spark.catalog.tableExists(table_ref), table_ref)

qa(
    "namespace_bce",
    all(".bce_" in table_ref for table_ref in GOLD_TABLES.values()),
    "Todas las tablas del notebook usan prefijo bce_.",
)

qa(
    "silver_source_v2",
    SILVER_TABLE.endswith(".bce_exportaciones_camaron_subpartida"),
    SILVER_TABLE,
)

qa(
    "tableau_una_hoja_plana",
    list(SERVING_EXPORTS.keys()) == ["tabla_tableau_bce"],
    str(list(SERVING_EXPORTS.keys())),
)

qa(
    "serving_sin_pais",
    not any(c.lower() in {"pais", "pais_destino", "id_pais", "pais_key"} for c in DF_SERVING_TABLEAU.columns),
    "El reporte BCE v2 no contiene país de destino.",
)

qa(
    "kpi_una_fila",
    spark.table(GOLD_TABLES["indicadores_kpi"]).count() == 1,
    "bce_indicadores_kpi debe contener una sola fila.",
)

qa(
    "fact_conserva_filas_silver",
    spark.table(GOLD_TABLES["fact_exportaciones"]).count() == SILVER_ROW_COUNT,
    f"silver={SILVER_ROW_COUNT}",
)

qa_failures = [x for x in QA_RESULTS if x["status"] == "FAIL"]

titulo("QA TÉCNICO GOLD BCE")
for x in QA_RESULTS:
    print(f"{x['status']:4s} | {x['test']:34s} | {x['detail']}")

if qa_failures:
    raise RuntimeError(f"QA Gold BCE detectó fallos: {qa_failures}")

print("\n✓ QA técnico Gold BCE superado.")


# COMMAND ----------

# =============================================================================
# CELDA 13 — MANTENIMIENTO DELTA CON CRITERIO OBJETIVO
# =============================================================================

MAINTENANCE_RESULTS: List[Dict[str, Any]] = []

titulo("MANTENIMIENTO DELTA CONDICIONADO")

for logical_name, table_ref in GOLD_TABLES.items():
    detail = spark.sql(f"DESCRIBE DETAIL {table_ref}").first().asDict()
    size_bytes = int(detail.get("sizeInBytes") or 0)
    num_files = int(detail.get("numFiles") or 0)

    should_optimize = (
        size_bytes >= OPTIMIZE_MIN_BYTES
        or num_files >= OPTIMIZE_MIN_FILES
    )

    if not should_optimize:
        status = "SKIPPED_SMALL_TABLE"
        print(
            f"SKIP | {logical_name:24s} | "
            f"{size_bytes:,} bytes | {num_files} archivos"
        )
    else:
        try:
            spark.sql(f"OPTIMIZE {table_ref}")
            status = "OPTIMIZED"
            print(
                f"PASS | {logical_name:24s} | "
                f"{size_bytes:,} bytes | {num_files} archivos"
            )
        except Exception as exc:
            status = f"SKIPPED_{type(exc).__name__}"
            print(f"WARN | {logical_name:24s} | {status}")

    MAINTENANCE_RESULTS.append({
        "logical_name": logical_name,
        "table": table_ref,
        "size_bytes_before": size_bytes,
        "num_files_before": num_files,
        "status": status,
    })

print("\n✓ Mantenimiento evaluado con umbrales objetivos.")


# COMMAND ----------

# =============================================================================
# CELDA 14 — MANIFIESTO GOLD, MÉTRICAS Y REPORTE FINAL PREVIO AL XLSX
# =============================================================================

RUN_FINISHED_AT_UTC = datetime.now(timezone.utc)
PIPELINE_DURATION_SECONDS = round(time.perf_counter() - RUN_STARTED_MONOTONIC, 6)

TABLE_METRICS = []
for logical_name, table_ref in GOLD_TABLES.items():
    df = spark.table(table_ref)
    detail = spark.sql(f"DESCRIBE DETAIL {table_ref}").first().asDict()
    TABLE_METRICS.append({
        "logical_name": logical_name,
        "table": table_ref,
        "role": GOLD_TABLE_ROLES[logical_name],
        "rows": df.count(),
        "columns": len(df.columns),
        "size_bytes": int(detail.get("sizeInBytes") or 0),
        "num_files": int(detail.get("numFiles") or 0),
    })

manifest = {
    "manifest_schema_version": GOLD_MANIFEST_SCHEMA_VERSION,
    "pipeline_version": PIPELINE_VERSION,
    "run_id": RUN_ID,
    "source": SOURCE_CODE,
    "status": "SUCCESS_READY_FOR_TABLEAU_EXPORT",
    "started_at_utc": RUN_STARTED_AT_UTC.isoformat(),
    "finished_at_utc": RUN_FINISHED_AT_UTC.isoformat(),
    "duration_seconds": PIPELINE_DURATION_SECONDS,
    "config_sha256": CONFIG_HASH,
    "silver_lineage": SILVER_LINEAGE,
    "silver_profile": SILVER_PROFILE,
    "dimension_checks": DIMENSION_CHECKS,
    "subpartida_presentation_mapping": {
        "strategy": "codigo_subpartida_to_short_label",
        "mapping": SUBPARTIDA_CORTA_POR_CODIGO,
        "audit": MAPPING_AUDIT,
    },
    "fact_checks": FACT_CHECKS,
    "quality": QUALITY_RESULTS,
    "persistence": PERSISTENCE_RESULTS,
    "postwrite_validation": POSTWRITE_RESULTS,
    "technical_qa": QA_RESULTS,
    "delta_maintenance": MAINTENANCE_RESULTS,
    "gold_tables": TABLE_METRICS,
    "serving_tables_for_tableau": SERVING_EXPORTS,
    "tableau_contract": {
        "sheet_count": 1,
        "sheets": ["tabla_tableau_bce"],
        "subpartida_label": "short_presentation_label",
        "relationships_required": False,
    },
    "software_versions": software_versions(),
}

GOLD_MANIFEST_PATH = GOLD_METADATA_ROOT / f"gold_manifest_{RUN_ID}_BCE.json"
atomic_json_dump(manifest, GOLD_MANIFEST_PATH)

reloaded = json.loads(GOLD_MANIFEST_PATH.read_text(encoding="utf-8"))
if (
    reloaded.get("run_id") != RUN_ID
    or reloaded.get("status") != "SUCCESS_READY_FOR_TABLEAU_EXPORT"
):
    raise RuntimeError("El manifiesto Gold BCE no pudo verificarse después de su escritura.")

manifest_bytes = GOLD_MANIFEST_PATH.read_bytes()
manifest_sha = sha256_bytes(manifest_bytes)

titulo("REPORTE FINAL — GOLD BCE v2")
print("Estado                    : SUCCESS_READY_FOR_TABLEAU_EXPORT")
print(f"Run ID                    : {RUN_ID}")
print(f"Pipeline version          : {PIPELINE_VERSION}")
print(f"Silver run                : {SILVER_RUN_ID}")
print(f"Bronze SHA-256            : {BRONZE_SHA256}")
print(f"Período mínimo            : {SILVER_PROFILE['periodo_min']}")
print(f"Período máximo            : {SILVER_PROFILE['periodo_max']}")
print(f"Duración medida           : {PIPELINE_DURATION_SECONDS:.3f} segundos")
print(f"Tablas Gold               : {len(GOLD_TABLES)}")
print(f"Hojas Tableau exportables : {len(SERVING_EXPORTS)}")
print(f"Reglas calidad            : {len(QUALITY_RESULTS)}")
print(f"Warnings calidad          : {sum(1 for x in QUALITY_RESULTS if x['status'] == 'WARN')}")
print(f"Tests QA                  : {len(QA_RESULTS)}")
print(f"Manifiesto                : {GOLD_MANIFEST_PATH}")
print(f"SHA-256 manifiesto        : {manifest_sha}")

print("\nÚNICA hoja que será exportada a Tableau en la ÚLTIMA CELDA:")
print(f"  - tabla_tableau_bce <- {SERVING_EXPORTS['tabla_tableau_bce']}")
print("\n✓ MODELO GOLD BCE LISTO PARA GENERAR EL XLSX DE SERVING.")


# COMMAND ----------

# =============================================================================
# CELDA 15 — ÚNICA EXPORTACIÓN XLSX PARA TABLEAU
# =============================================================================
# Esta celda SOLO materializa el archivo Excel de serving.
# El libro final contiene EXACTAMENTE UNA hoja plana/autocontenida:
#   - tabla_tableau_bce
# No se exportan dimensiones, facts, resúmenes ni KPI internos.

from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font


titulo("EXPORTACIÓN FINAL — UNA SOLA TABLA PLANA PARA TABLEAU")

EXPECTED_SHEETS = ["tabla_tableau_bce"]

if list(SERVING_EXPORTS.keys()) != EXPECTED_SHEETS:
    raise RuntimeError(
        f"Contrato Tableau alterado. Esperado={EXPECTED_SHEETS}; "
        f"obtenido={list(SERVING_EXPORTS.keys())}"
    )

sheet_name = "tabla_tableau_bce"
table_ref = SERVING_EXPORTS[sheet_name]

df = spark.table(table_ref).orderBy(
    "fecha_periodo",
    "ranking_fob_mes",
    "codigo_subpartida",
)

spark_rows = df.count()
if spark_rows <= 0:
    raise RuntimeError(f"La tabla plana para Tableau está vacía: {table_ref}")

pdf = df.toPandas()

if len(pdf) != spark_rows:
    raise RuntimeError(
        f"Conversión Spark→Pandas alteró el conteo: spark={spark_rows}, pandas={len(pdf)}"
    )

print(f"PASS | {sheet_name:24s} | {len(pdf):5d} filas | {len(pdf.columns):2d} columnas")

buffer = BytesIO()

with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
    pdf.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for column_cells in ws.columns:
        max_len = 0
        for cell in list(column_cells)[:300]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_cells[0].column_letter].width = min(
            max(max_len + 2, 10), 45
        )

payload = buffer.getvalue()
if not payload:
    raise RuntimeError("El XLSX generado está vacío.")

# Validación estructural en memoria antes de escribir en el Volume.
wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)

if wb.sheetnames != EXPECTED_SHEETS:
    raise RuntimeError(
        f"El XLSX final debe contener exclusivamente {EXPECTED_SHEETS}; "
        f"obtenido={wb.sheetnames}"
    )

ws = wb[sheet_name]
actual_rows = max(int(ws.max_row) - 1, 0)
actual_cols = int(ws.max_column)

if actual_rows != len(pdf):
    raise RuntimeError(
        f"Conteo XLSX incorrecto: esperado={len(pdf)}, obtenido={actual_rows}"
    )
if actual_cols != len(pdf.columns):
    raise RuntimeError(
        f"Columnas XLSX incorrectas: esperado={len(pdf.columns)}, obtenido={actual_cols}"
    )

wb.close()

DELIVERABLE_ROOT.mkdir(parents=True, exist_ok=True)
EXCEL_PATH.write_bytes(payload)

if not EXCEL_PATH.exists() or EXCEL_PATH.stat().st_size != len(payload):
    raise RuntimeError("El XLSX no pudo verificarse después de persistirlo.")

excel_sha256 = sha256_bytes(payload)

print("\nArchivo Tableau creado y validado.")
print(f"Ruta       : {EXCEL_PATH}")
print(f"Tamaño     : {len(payload):,} bytes")
print(f"SHA-256    : {excel_sha256}")
print(f"Hojas      : {EXPECTED_SHEETS}")

print("\n✓ El archivo contiene EXACTAMENTE 1 hoja.")
print("✓ tabla_tableau_bce es plana y autocontenida.")
print("✓ La columna subpartida usa nombres cortos para el dashboard.")
print("✓ La descripción oficial permanece preservada internamente en Silver/Gold.")
print("✓ No se exportan dimensiones, facts, resúmenes ni KPI auxiliares.")
print("✓ No existe dimensión país porque el reporte BCE v2 no contiene País Destino.")
print("✓ Tableau puede consumir esta hoja como una única fuente sin relaciones.")
print("✓ Para precio nacional ponderado en Tableau use:")
print("  SUM([fob_usd]) / (SUM([toneladas_metricas_peso_neto]) * 1000)")


# COMMAND ----------
