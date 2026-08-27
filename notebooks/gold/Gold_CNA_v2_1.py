# Databricks notebook source

# =============================================================================
# CELDA 1 — IMPORTACIONES, CONFIGURACIÓN, VERSIONADO Y UTILIDADES
# =============================================================================

import json
import hashlib
import time
import platform
import importlib.metadata as importlib_metadata

from pathlib import Path
from datetime import datetime, timezone
from typing import Any, Dict, List, Tuple
%pip install openpyxl

import pandas as pd
import openpyxl

from pyspark.sql import DataFrame
from pyspark.sql import functions as F
from pyspark.sql import types as T
from pyspark.sql.window import Window


PIPELINE_VERSION = "2.1"
GOLD_MANIFEST_SCHEMA_VERSION = "1.0"
SUPPORTED_SILVER_VERSIONS = {"2.1.0"}

CATALOG = "camaronera_2026"
SILVER_SCHEMA = "plata"
GOLD_SCHEMA = "oro"

SOURCE_CODE = "CNA"
GOLD_METADATA_VOLUME = "metadatos_cna_gold"
DELIVERABLE_VOLUME = "entregables"

RUN_STARTED_AT_UTC = datetime.now(timezone.utc)
RUN_STARTED_MONOTONIC = time.perf_counter()
RUN_ID = RUN_STARTED_AT_UTC.strftime("%Y%m%dT%H%M%S%fZ")

SILVER_TABLES = {
    "exportaciones_mensuales": f"{CATALOG}.{SILVER_SCHEMA}.exportaciones_mensuales",
    "exportaciones_mensuales_dolares_normalizada": f"{CATALOG}.{SILVER_SCHEMA}.exportaciones_mensuales_dolares_normalizada",
    "exportaciones_mensuales_libras_normalizada": f"{CATALOG}.{SILVER_SCHEMA}.exportaciones_mensuales_libras_normalizada",
    "mercado_pais_mayo_normalizado": f"{CATALOG}.{SILVER_SCHEMA}.mercado_pais_mayo_normalizado",
    "mercado_pais_acumulado_normalizado": f"{CATALOG}.{SILVER_SCHEMA}.mercado_pais_acumulado_normalizado",
    "exportaciones_mayo_historico": f"{CATALOG}.{SILVER_SCHEMA}.exportaciones_mayo_historico",
    "exportaciones_acumuladas_historico": f"{CATALOG}.{SILVER_SCHEMA}.exportaciones_acumuladas_historico",
}

# Tablas Gold internas. Se usa prefijo CNA para impedir colisiones con BCE.
GOLD_TABLES = {
    "dim_fecha": f"{CATALOG}.{GOLD_SCHEMA}.cna_dim_fecha",
    "dim_destino": f"{CATALOG}.{GOLD_SCHEMA}.cna_dim_destino",
    "dim_corte": f"{CATALOG}.{GOLD_SCHEMA}.cna_dim_corte",
    "fact_nacional": f"{CATALOG}.{GOLD_SCHEMA}.cna_fact_exportaciones_nacional",
    "fact_mercados": f"{CATALOG}.{GOLD_SCHEMA}.cna_fact_mercados",
    "fact_comparativo": f"{CATALOG}.{GOLD_SCHEMA}.cna_fact_comparativo_historico",
    "reconciliacion_nacional": f"{CATALOG}.{GOLD_SCHEMA}.cna_reconciliacion_nacional",
    "audit_mercados_agregados": f"{CATALOG}.{GOLD_SCHEMA}.cna_audit_mercados_agregados",
    "serving_nacional_mensual": f"{CATALOG}.{GOLD_SCHEMA}.cna_serving_nacional_mensual",
    "serving_nacional_anual": f"{CATALOG}.{GOLD_SCHEMA}.cna_serving_nacional_anual",
    "serving_mercados": f"{CATALOG}.{GOLD_SCHEMA}.cna_serving_mercados",
    "serving_regiones": f"{CATALOG}.{GOLD_SCHEMA}.cna_serving_regiones_agregadas",
    "serving_comparativo": f"{CATALOG}.{GOLD_SCHEMA}.cna_serving_comparativo_historico",
    "serving_kpi": f"{CATALOG}.{GOLD_SCHEMA}.cna_serving_kpi",
}

GOLD_TABLE_ROLES = {
    "dim_fecha": "dimension",
    "dim_destino": "dimension",
    "dim_corte": "dimension",
    "fact_nacional": "fact",
    "fact_mercados": "fact",
    "fact_comparativo": "fact",
    "reconciliacion_nacional": "audit",
    "audit_mercados_agregados": "audit",
    "serving_nacional_mensual": "serving",
    "serving_nacional_anual": "serving",
    "serving_mercados": "serving",
    "serving_regiones": "serving",
    "serving_comparativo": "serving",
    "serving_kpi": "serving",
}

# ÚNICAS hojas que salen al XLSX final de Tableau.
# Cada una es una tabla plana/autocontenida; no se exportan dimensiones ni facts.
SERVING_EXPORTS = {
    "tabla_tableau_nacional": GOLD_TABLES["serving_nacional_mensual"],
    "tabla_tableau_destinos": GOLD_TABLES["serving_mercados"],
    "tabla_tableau_regiones": GOLD_TABLES["serving_regiones"],
}

GOLD_METADATA_ROOT = Path(f"/Volumes/{CATALOG}/{GOLD_SCHEMA}/{GOLD_METADATA_VOLUME}")
DELIVERABLE_ROOT = Path(f"/Volumes/{CATALOG}/{GOLD_SCHEMA}/{DELIVERABLE_VOLUME}")
EXCEL_PATH = DELIVERABLE_ROOT / "Exportaciones_CNA_Gold_Tableau.xlsx"

ABS_RECON_TOL = 1.0
REL_RECON_TOL = 1e-9
ABS_VARIATION_TOL = 1e-6
REL_VARIATION_TOL = 1e-5

# Mantenimiento Delta: solo se ejecuta si la tabla lo justifica físicamente.
OPTIMIZE_MIN_BYTES = 256 * 1024 * 1024
OPTIMIZE_MIN_FILES = 8


def titulo(texto: str) -> None:
    print("\n" + "=" * 100)
    print(texto)
    print("=" * 100)


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def atomic_json_dump(payload: Dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    tmp.replace(path)


def sql_escape(value: Any) -> str:
    return str(value).replace("'", "''")


def table_properties(table_ref: str) -> Dict[str, str]:
    rows = spark.sql(f"SHOW TBLPROPERTIES {table_ref}").collect()
    return {str(r[0]): str(r[1]) for r in rows}


def software_versions() -> Dict[str, str]:
    import os
    import platform
    versions = {
        "python": platform.python_version(),
        "spark": getattr(spark, "version", "unknown"),
        "databricks_runtime": os.getenv("DATABRICKS_RUNTIME_VERSION", "unknown"),
    }
    for pkg in ["pyspark", "pandas", "openpyxl", "delta-spark", "playwright"]:
        try:
            versions[pkg] = importlib_metadata.version(pkg)
        except Exception:
            versions[pkg] = "not_installed"
    try:
        versions["spark_databricks_version"] = spark.conf.get(
            "spark.databricks.clusterUsageTags.sparkVersion"
        )
    except Exception:
        versions["spark_databricks_version"] = "unavailable"
    return versions

spark.sql(f"CREATE SCHEMA IF NOT EXISTS `{CATALOG}`.`{GOLD_SCHEMA}`")
spark.sql(f"CREATE VOLUME IF NOT EXISTS `{CATALOG}`.`{GOLD_SCHEMA}`.`{GOLD_METADATA_VOLUME}`")
spark.sql(f"CREATE VOLUME IF NOT EXISTS `{CATALOG}`.`{GOLD_SCHEMA}`.`{DELIVERABLE_VOLUME}`")

CONFIG_PAYLOAD = {
    "pipeline_version": PIPELINE_VERSION,
    "silver_tables": SILVER_TABLES,
    "gold_tables": GOLD_TABLES,
    "serving_exports": SERVING_EXPORTS,
    "reconciliation_rule": (
        "En periodos coincidentes se prioriza exportaciones_mensuales; "
        "las matrices normalizadas se utilizan para extensión histórica."
    ),
    "annual_price_formula": "SUM(dolares) / SUM(libras)",
    "market_semantics": (
        "source field 'pais' is retained only at the Silver contract boundary; "
        "Gold models the concept as 'destino'; regional aggregates are separated by "
        "nivel_destino and excluded from detailed rankings/participations; no geopolitical "
        "PAIS/TERRITORIO classification is used as a calculation rule"
    ),
}
CONFIG_HASH = sha256_text(json.dumps(CONFIG_PAYLOAD, sort_keys=True, ensure_ascii=False))

titulo("CONFIGURACIÓN GOLD CNA v2")
print(f"Run ID             : {RUN_ID}")
print(f"Pipeline version   : {PIPELINE_VERSION}")
print(f"Catálogo           : {CATALOG}")
print(f"Silver schema      : {SILVER_SCHEMA}")
print(f"Gold schema        : {GOLD_SCHEMA}")
print(f"Config SHA-256     : {CONFIG_HASH}")
print(f"Entregable final   : {EXCEL_PATH}")


# COMMAND ----------

# =============================================================================
# CELDA 2 — CONTRATO SILVER→GOLD Y VALIDACIÓN DE LINEAGE
# =============================================================================

EXPECTED_SILVER_COLUMNS = {
    "exportaciones_mensuales": [
        "fecha_periodo", "anio", "mes_numero", "libras", "dolares", "precio_promedio"
    ],
    "exportaciones_mensuales_dolares_normalizada": [
        "fecha_periodo", "anio", "mes_numero", "valor"
    ],
    "exportaciones_mensuales_libras_normalizada": [
        "fecha_periodo", "anio", "mes_numero", "valor"
    ],
    "mercado_pais_mayo_normalizado": [
        "pais", "indicador", "valor"
    ],
    "mercado_pais_acumulado_normalizado": [
        "pais", "indicador", "valor"
    ],
    "exportaciones_mayo_historico": [
        "anio", "libras", "dolares", "libras_variacion", "dolares_variacion"
    ],
    "exportaciones_acumuladas_historico": [
        "periodo", "libras", "dolares", "libras_variacion", "dolares_variacion"
    ],
}

REQUIRED_SILVER_PROPERTIES = {
    "pipeline.layer": "silver",
    "pipeline.source": SOURCE_CODE,
}

SILVER_CONTRACT_RESULTS: List[Dict[str, Any]] = []
SILVER_LINEAGE_BY_TABLE: Dict[str, Dict[str, str]] = {}
SILVER_DFS: Dict[str, DataFrame] = {}

titulo("VALIDACIÓN DEL CONTRATO SILVER→GOLD")

for logical_name, table_ref in SILVER_TABLES.items():
    if not spark.catalog.tableExists(table_ref):
        raise RuntimeError(f"Falta la tabla Silver requerida: {table_ref}")

    df = spark.table(table_ref)
    actual_columns = df.columns
    expected_columns = EXPECTED_SILVER_COLUMNS[logical_name]

    if actual_columns != expected_columns:
        raise RuntimeError(
            f"Contrato Silver incompatible en {table_ref}.\n"
            f"Esperado: {expected_columns}\n"
            f"Obtenido: {actual_columns}"
        )

    props = table_properties(table_ref)

    for key, expected_value in REQUIRED_SILVER_PROPERTIES.items():
        actual_value = props.get(key)
        if actual_value != expected_value:
            raise RuntimeError(
                f"Lineage inválido en {table_ref}: {key}={actual_value!r}; "
                f"se esperaba {expected_value!r}."
            )

    silver_version = props.get("pipeline.version")
    if silver_version not in SUPPORTED_SILVER_VERSIONS:
        raise RuntimeError(
            f"Versión Silver no soportada en {table_ref}: {silver_version!r}. "
            f"Versiones admitidas: {sorted(SUPPORTED_SILVER_VERSIONS)}"
        )

    silver_run_id = props.get("pipeline.silver_run_id")
    bronze_sha = props.get("lineage.bronze_sha256")
    report_period = props.get("lineage.report_period")

    if not silver_run_id or not bronze_sha or not report_period:
        raise RuntimeError(
            f"Lineage incompleto en {table_ref}. "
            "Se requieren pipeline.silver_run_id, lineage.bronze_sha256 y lineage.report_period."
        )

    row_count = df.count()
    if row_count <= 0:
        raise RuntimeError(f"La tabla Silver está vacía: {table_ref}")

    SILVER_DFS[logical_name] = df
    SILVER_LINEAGE_BY_TABLE[logical_name] = props
    SILVER_CONTRACT_RESULTS.append({
        "table": table_ref,
        "rows": row_count,
        "silver_run_id": silver_run_id,
        "bronze_sha256": bronze_sha,
        "report_period": report_period,
        "status": "PASS",
    })

    print(
        f"PASS | {logical_name:48s} | filas={row_count:4d} | "
        f"silver_run={silver_run_id}"
    )

silver_run_ids = {x["silver_run_id"] for x in SILVER_CONTRACT_RESULTS}
bronze_hashes = {x["bronze_sha256"] for x in SILVER_CONTRACT_RESULTS}
report_periods = {x["report_period"] for x in SILVER_CONTRACT_RESULTS}

if len(silver_run_ids) != 1:
    raise RuntimeError(f"Las siete tablas Silver no pertenecen a la misma ejecución: {silver_run_ids}")
if len(bronze_hashes) != 1:
    raise RuntimeError(f"Las siete tablas Silver no comparten el mismo SHA Bronze: {bronze_hashes}")
if len(report_periods) != 1:
    raise RuntimeError(f"Las siete tablas Silver no comparten el mismo período documental: {report_periods}")

SILVER_RUN_ID = next(iter(silver_run_ids))
BRONZE_SHA256 = next(iter(bronze_hashes))
REPORT_PERIOD = next(iter(report_periods))

print("\nContrato Silver v2 compatible.")
print(f"Silver run ID      : {SILVER_RUN_ID}")
print(f"Bronze SHA-256     : {BRONZE_SHA256}")
print(f"Período documental : {REPORT_PERIOD}")


# COMMAND ----------

# =============================================================================
# CELDA 3 — PERFILADO DE ENTRADAS SILVER SIN SUPONER CONTEOS FIJOS
# =============================================================================

PROFILE_RESULTS: List[Dict[str, Any]] = []

titulo("PERFILADO SILVER CNA")

for name, df in SILVER_DFS.items():
    row_count = df.count()
    exact_dups = (
        df.groupBy(*df.columns).count().filter(F.col("count") > 1).count()
        if df.columns else 0
    )
    null_exprs = [
        F.sum(F.when(F.col(c).isNull(), 1).otherwise(0)).alias(c)
        for c in df.columns
    ]
    nulls = df.agg(*null_exprs).first().asDict() if row_count else {c: 0 for c in df.columns}
    total_nulls = sum(int(v or 0) for v in nulls.values())

    PROFILE_RESULTS.append({
        "table": name,
        "rows": row_count,
        "columns": len(df.columns),
        "exact_duplicate_groups": exact_dups,
        "total_null_cells": total_nulls,
        "schema": {f.name: f.dataType.simpleString() for f in df.schema.fields},
    })

    if exact_dups > 0:
        raise RuntimeError(f"Silver contiene duplicados exactos inesperados en {name}: {exact_dups}")

    print(
        f"PASS | {name:48s} | filas={row_count:4d} | "
        f"columnas={len(df.columns):2d} | nulos={total_nulls:3d}"
    )

# Validar las claves naturales básicas que Gold necesita.
for name in [
    "exportaciones_mensuales_dolares_normalizada",
    "exportaciones_mensuales_libras_normalizada",
]:
    df = SILVER_DFS[name]
    dup = (
        df.groupBy("fecha_periodo")
          .count()
          .filter(F.col("count") > 1)
          .count()
    )
    if dup:
        raise RuntimeError(f"{name} no tiene grano único por fecha_periodo: {dup} grupos duplicados.")

dup_direct = (
    SILVER_DFS["exportaciones_mensuales"]
    .groupBy("fecha_periodo")
    .count()
    .filter(F.col("count") > 1)
    .count()
)
if dup_direct:
    raise RuntimeError(f"exportaciones_mensuales no es única por fecha_periodo: {dup_direct} grupos.")

print("\n✓ Perfilado Silver completado sin imponer conteos históricos rígidos.")


# COMMAND ----------

# =============================================================================
# CELDA 4 — SERIE NACIONAL CANÓNICA Y RECONCILIACIÓN ENTRE ESTRUCTURAS CNA
# =============================================================================

df_dolares_matriz = (
    SILVER_DFS["exportaciones_mensuales_dolares_normalizada"]
    .select(
        "fecha_periodo",
        F.col("valor").alias("dolares_matriz"),
    )
)

df_libras_matriz = (
    SILVER_DFS["exportaciones_mensuales_libras_normalizada"]
    .select(
        "fecha_periodo",
        F.col("valor").alias("libras_matriz"),
    )
)

df_matriz = (
    df_dolares_matriz
    .join(df_libras_matriz, on="fecha_periodo", how="full_outer")
)

df_directo = (
    SILVER_DFS["exportaciones_mensuales"]
    .select(
        "fecha_periodo",
        F.col("libras").alias("libras_directo"),
        F.col("dolares").alias("dolares_directo"),
        F.col("precio_promedio").alias("precio_reportado_directo"),
    )
)

df_reconciliacion = (
    df_matriz
    .join(df_directo, on="fecha_periodo", how="full_outer")
    .withColumn("anio", F.year("fecha_periodo"))
    .withColumn("mes_numero", F.month("fecha_periodo"))
    .withColumn(
        "dif_abs_libras",
        F.when(
            F.col("libras_directo").isNotNull() & F.col("libras_matriz").isNotNull(),
            F.abs(F.col("libras_directo") - F.col("libras_matriz")),
        ),
    )
    .withColumn(
        "dif_abs_dolares",
        F.when(
            F.col("dolares_directo").isNotNull() & F.col("dolares_matriz").isNotNull(),
            F.abs(F.col("dolares_directo") - F.col("dolares_matriz")),
        ),
    )
    .withColumn(
        "coincide_libras",
        F.when(
            F.col("libras_directo").isNull() | F.col("libras_matriz").isNull(),
            F.lit(None).cast("boolean"),
        ).otherwise(
            F.col("dif_abs_libras")
            <= F.greatest(
                F.lit(ABS_RECON_TOL),
                F.abs(F.col("libras_directo")) * F.lit(REL_RECON_TOL),
            )
        ),
    )
    .withColumn(
        "coincide_dolares",
        F.when(
            F.col("dolares_directo").isNull() | F.col("dolares_matriz").isNull(),
            F.lit(None).cast("boolean"),
        ).otherwise(
            F.col("dif_abs_dolares")
            <= F.greatest(
                F.lit(ABS_RECON_TOL),
                F.abs(F.col("dolares_directo")) * F.lit(REL_RECON_TOL),
            )
        ),
    )
    .withColumn(
        "estado_reconciliacion",
        F.when(
            F.col("libras_directo").isNotNull()
            & F.col("libras_matriz").isNotNull()
            & F.col("dolares_directo").isNotNull()
            & F.col("dolares_matriz").isNotNull()
            & F.coalesce(F.col("coincide_libras"), F.lit(False))
            & F.coalesce(F.col("coincide_dolares"), F.lit(False)),
            F.lit("COINCIDE"),
        )
        .when(
            F.col("libras_directo").isNotNull()
            | F.col("dolares_directo").isNotNull(),
            F.when(
                F.col("libras_matriz").isNotNull() | F.col("dolares_matriz").isNotNull(),
                F.lit("DIFIERE_PRIORIZA_DIRECTO"),
            ).otherwise(F.lit("SOLO_DIRECTO")),
        )
        .otherwise(F.lit("SOLO_MATRIZ")),
    )
    # Regla canónica explícita:
    # 1) tabla mensual directa cuando existe;
    # 2) matriz histórica como extensión de cobertura.
    .withColumn("libras", F.coalesce("libras_directo", "libras_matriz"))
    .withColumn("dolares", F.coalesce("dolares_directo", "dolares_matriz"))
    .withColumn(
        "fuente_libras",
        F.when(F.col("libras_directo").isNotNull(), F.lit("exportaciones_mensuales"))
         .otherwise(F.lit("matriz_historica")),
    )
    .withColumn(
        "fuente_dolares",
        F.when(F.col("dolares_directo").isNotNull(), F.lit("exportaciones_mensuales"))
         .otherwise(F.lit("matriz_historica")),
    )
    .withColumn(
        "precio_promedio_usd_lb",
        F.when(
            F.col("libras").isNotNull() & (F.col("libras") > 0) & F.col("dolares").isNotNull(),
            F.col("dolares") / F.col("libras"),
        ),
    )
)

# Solo períodos con al menos una magnitud analítica.
DF_NACIONAL_BASE = (
    df_reconciliacion
    .filter(F.col("libras").isNotNull() | F.col("dolares").isNotNull())
    .select(
        "fecha_periodo", "anio", "mes_numero",
        "libras", "dolares", "precio_promedio_usd_lb",
        "fuente_libras", "fuente_dolares",
        "estado_reconciliacion",
    )
)

DF_RECONCILIACION_NACIONAL = (
    df_reconciliacion
    .select(
        "fecha_periodo", "anio", "mes_numero",
        "libras_directo", "libras_matriz", "dif_abs_libras", "coincide_libras",
        "dolares_directo", "dolares_matriz", "dif_abs_dolares", "coincide_dolares",
        "precio_reportado_directo",
        "estado_reconciliacion",
    )
)

if DF_NACIONAL_BASE.count() == 0:
    raise RuntimeError("La reconciliación nacional no produjo observaciones.")

dup_nacional = (
    DF_NACIONAL_BASE.groupBy("fecha_periodo")
    .count().filter(F.col("count") > 1).count()
)
if dup_nacional:
    raise RuntimeError(f"La serie nacional canónica tiene {dup_nacional} fechas duplicadas.")

negative_nacional = DF_NACIONAL_BASE.filter(
    (F.col("libras") < 0) | (F.col("dolares") < 0)
).count()
if negative_nacional:
    raise RuntimeError(f"La serie nacional contiene {negative_nacional} observaciones negativas.")

RECON_COUNTS = {
    r["estado_reconciliacion"]: r["count"]
    for r in DF_RECONCILIACION_NACIONAL.groupBy("estado_reconciliacion").count().collect()
}

titulo("RECONCILIACIÓN SERIE NACIONAL")
for k, v in sorted(RECON_COUNTS.items()):
    print(f"{k:28s}: {v}")

print(f"\nObservaciones nacionales canónicas: {DF_NACIONAL_BASE.count()}")
print("Regla: se prioriza la tabla mensual directa cuando existe; la matriz extiende la historia.")


# COMMAND ----------

# =============================================================================
# CELDA 5 — MERCADOS: DESTINOS ANALÍTICOS A GRANO DETALLE × CORTE × AÑO
# =============================================================================
# DECISIÓN SEMÁNTICA:
# - Silver conserva el contrato fuente de CNA, cuyo campo se denomina `pais`.
# - Gold NO interpreta automáticamente ese campo como "Estado soberano".
# - Gold crea `destino_fuente` (etiqueta proveniente de Silver), `destino`
#   (etiqueta analítica homologada) y `nivel_destino`.
# - `nivel_destino` controla la regla matemática:
#       DETALLADO          -> participa en ranking, denominador y participación.
#       AGREGADO_REGIONAL  -> se conserva para auditoría, nunca se mezcla con detalle.
# - No se utiliza una clasificación geopolítica PAIS/TERRITORIO para decidir cálculos.
#   Esto evita introducir supuestos externos innecesarios sobre Aruba, Hong Kong,
#   Puerto Rico, Taiwán, Martinica, Reunión u otras economías/territorios.

PATRON_NO_DESTINO_ANALITICO = r"^(TOTAL\b.*|SUBTOTAL\b.*|OTROS\b.*|RESTO\b.*|MUNDO|GLOBAL)$"

AGREGADOS_REGIONALES = ["AFRICA", "AMERICA", "EUROPA", "OCEANIA"]

# Homologaciones explícitas observadas en los archivos CNA procesados.
# Se aplican únicamente a la etiqueta analítica Gold. La etiqueta proveniente
# de Silver queda disponible en `destino_fuente` y el original físico permanece
# íntegro en Bronze.
NORMALIZACION_ETIQUETAS_DESTINO = {
    "MARTINICA (COLONIA FRANCIA)": "MARTINICA",
    "REUNION (COLONIA FRANCIA)": "REUNION",
}

_DESTINO_MAP_EXPR = F.create_map(*[
    item
    for origen, destino in NORMALIZACION_ETIQUETAS_DESTINO.items()
    for item in (F.lit(origen), F.lit(destino))
])


def preparar_mercado(df: DataFrame, corte: str) -> Tuple[DataFrame, DataFrame]:
    # `pais` es nombre de campo del contrato Silver; a partir de aquí se trata
    # explícitamente como etiqueta de destino publicada por la fuente.
    base_pre = (
        df
        .withColumn("destino_fuente", F.upper(F.trim(F.col("pais"))))
        .filter(F.col("destino_fuente").isNotNull() & (F.col("destino_fuente") != ""))
        .filter(~F.col("destino_fuente").rlike(PATRON_NO_DESTINO_ANALITICO))
    )

    base = (
        base_pre
        .withColumn(
            "destino",
            F.coalesce(_DESTINO_MAP_EXPR[F.col("destino_fuente")], F.col("destino_fuente"))
        )
        .withColumn(
            "nivel_destino",
            F.when(F.col("destino").isin(AGREGADOS_REGIONALES), F.lit("AGREGADO_REGIONAL"))
             .otherwise(F.lit("DETALLADO"))
        )
    )

    # Evita que dos etiquetas fuente distintas se fusionen silenciosamente en
    # un mismo destino/indicador después de la homologación. Si en una edición
    # futura aparece un nuevo alias, el pipeline se detiene y exige revisión.
    alias_collisions = (
        base
        .groupBy("destino", "indicador")
        .agg(F.countDistinct("destino_fuente").alias("n_etiquetas_fuente"))
        .filter(F.col("n_etiquetas_fuente") > 1)
        .count()
    )
    if alias_collisions:
        raise RuntimeError(
            f"Se detectaron {alias_collisions} colisiones de etiquetas fuente tras homologación; "
            "revise NORMALIZACION_ETIQUETAS_DESTINO antes de continuar."
        )

    if corte == "MAYO":
        year_expr = F.regexp_extract("indicador", r"^(?:dolares|libras)_(\d{4})_05$", 1)
        unit_expr = F.regexp_extract("indicador", r"^(dolares|libras)_\d{4}_05$", 1)
    elif corte == "ACUM_ENE_MAY":
        year_expr = F.regexp_extract("indicador", r"^(?:dolares|libras)_ene_may_(\d{4})$", 1)
        unit_expr = F.regexp_extract("indicador", r"^(dolares|libras)_ene_may_\d{4}$", 1)
    else:
        raise ValueError(f"Corte no soportado: {corte}")

    # Catálogo determinístico de etiquetas de procedencia por destino analítico.
    # concat_ws(sort_array(collect_set())) preserva toda variante fuente si en
    # futuras ediciones varias etiquetas convergen al mismo destino.
    source_labels = (
        base
        .select("destino", "nivel_destino", "destino_fuente")
        .distinct()
        .groupBy("destino", "nivel_destino")
        .agg(
            F.concat_ws(
                " | ",
                F.sort_array(F.collect_set("destino_fuente"))
            ).alias("destino_fuente")
        )
    )

    values_long = (
        base
        .withColumn("anio_txt", year_expr)
        .withColumn("unidad", unit_expr)
        .filter((F.col("anio_txt") != "") & (F.col("unidad") != ""))
        .withColumn("anio", F.col("anio_txt").cast("int"))
        .select(
            "destino", "nivel_destino",
            F.lit(corte).alias("corte"),
            "anio", "unidad",
            F.col("valor").cast("double").alias("valor"),
        )
    )

    values_wide = (
        values_long
        .groupBy("destino", "nivel_destino", "corte", "anio")
        .pivot("unidad", ["dolares", "libras"])
        .agg(F.first("valor"))
        .join(source_labels, on=["destino", "nivel_destino"], how="left")
        .select(
            "destino_fuente", "destino", "nivel_destino",
            "corte", "anio", "libras", "dolares"
        )
    )

    # Los indicadores publicados sin año explícito corresponden a destinos
    # detallados. Los agregados regionales se excluyen de esta rama para impedir
    # que una participación de región se compare con una participación de destino.
    source_indicators = (
        base
        .filter(F.col("indicador").isin(
            "variacion_dolares", "variacion_libras", "participacion_libras"
        ))
        .filter(F.col("nivel_destino") == "DETALLADO")
        .groupBy("destino")
        .pivot("indicador", ["variacion_dolares", "variacion_libras", "participacion_libras"])
        .agg(F.first("valor"))
        .withColumn("corte", F.lit(corte))
        .select(
            "destino", "corte",
            F.col("variacion_dolares").alias("variacion_dolares_fuente"),
            F.col("variacion_libras").alias("variacion_libras_fuente"),
            F.col("participacion_libras").alias("participacion_libras_fuente"),
        )
    )

    return values_wide, source_indicators


# Auditoría explícita de homologaciones. No modifica Bronze ni Silver.
DF_DESTINO_NORMALIZACIONES_AUDIT = (
    SILVER_DFS["mercado_pais_mayo_normalizado"]
    .select(F.upper(F.trim(F.col("pais"))).alias("destino_fuente"))
    .unionByName(
        SILVER_DFS["mercado_pais_acumulado_normalizado"]
        .select(F.upper(F.trim(F.col("pais"))).alias("destino_fuente"))
    )
    .filter(F.col("destino_fuente").isin(list(NORMALIZACION_ETIQUETAS_DESTINO.keys())))
    .distinct()
    .withColumn("destino", _DESTINO_MAP_EXPR[F.col("destino_fuente")])
    .orderBy("destino_fuente")
)

mercado_mayo, indicadores_mayo = preparar_mercado(
    SILVER_DFS["mercado_pais_mayo_normalizado"], "MAYO"
)
mercado_acum, indicadores_acum = preparar_mercado(
    SILVER_DFS["mercado_pais_acumulado_normalizado"], "ACUM_ENE_MAY"
)

df_mercado_todos = mercado_mayo.unionByName(mercado_acum)

# Auditoría persistible: agregados regionales encontrados en la fuente.
DF_MERCADO_AGREGADOS_EXCLUIDOS = (
    df_mercado_todos
    .filter(F.col("nivel_destino") == "AGREGADO_REGIONAL")
    .orderBy("corte", "anio", "destino")
)

# Modelo analítico: únicamente destinos individuales/detallados.
df_mercado = (
    df_mercado_todos
    .filter(F.col("nivel_destino") == "DETALLADO")
)

df_indicadores_fuente = indicadores_mayo.unionByName(indicadores_acum)

w_destination = Window.partitionBy("destino", "corte").orderBy("anio")
w_market_year = Window.partitionBy("corte", "anio")
w_rank_dol = Window.partitionBy("corte", "anio").orderBy(
    F.col("dolares").desc_nulls_last(), F.col("destino")
)
w_rank_lb = Window.partitionBy("corte", "anio").orderBy(
    F.col("libras").desc_nulls_last(), F.col("destino")
)

df_mercado_calc = (
    df_mercado
    .withColumn(
        "precio_promedio_usd_lb",
        F.when(
            F.col("libras").isNotNull()
            & (F.col("libras") > 0)
            & F.col("dolares").isNotNull(),
            F.col("dolares") / F.col("libras"),
        ),
    )
    .withColumn("total_libras_corte_anio", F.sum("libras").over(w_market_year))
    .withColumn(
        "participacion_libras",
        F.when(
            F.col("libras").isNotNull()
            & (F.col("total_libras_corte_anio") > 0),
            F.col("libras") / F.col("total_libras_corte_anio"),
        ),
    )
    .withColumn("ranking_dolares", F.dense_rank().over(w_rank_dol))
    .withColumn("ranking_libras", F.dense_rank().over(w_rank_lb))
    .withColumn("anio_anterior", F.lag("anio").over(w_destination))
    .withColumn("dolares_anterior", F.lag("dolares").over(w_destination))
    .withColumn("libras_anterior", F.lag("libras").over(w_destination))
    .withColumn(
        "variacion_yoy_dolares",
        F.when(
            (F.col("anio") - F.col("anio_anterior") == 1)
            & F.col("dolares_anterior").isNotNull()
            & (F.col("dolares_anterior") != 0)
            & F.col("dolares").isNotNull(),
            (F.col("dolares") - F.col("dolares_anterior")) / F.col("dolares_anterior"),
        ),
    )
    .withColumn(
        "variacion_yoy_libras",
        F.when(
            (F.col("anio") - F.col("anio_anterior") == 1)
            & F.col("libras_anterior").isNotNull()
            & (F.col("libras_anterior") != 0)
            & F.col("libras").isNotNull(),
            (F.col("libras") - F.col("libras_anterior")) / F.col("libras_anterior"),
        ),
    )
    .drop(
        "total_libras_corte_anio",
        "anio_anterior", "dolares_anterior", "libras_anterior",
    )
)

latest_year_by_cut = (
    df_mercado_calc.groupBy("corte").agg(F.max("anio").alias("anio_ultimo"))
)

DF_MERCADO_BASE = (
    df_mercado_calc
    .join(df_indicadores_fuente, on=["destino", "corte"], how="left")
    .join(latest_year_by_cut, on="corte", how="left")
    # Los indicadores publicados sin año propio corresponden al corte más reciente;
    # no se replican artificialmente sobre años anteriores.
    .withColumn(
        "variacion_dolares_fuente",
        F.when(F.col("anio") == F.col("anio_ultimo"), F.col("variacion_dolares_fuente")),
    )
    .withColumn(
        "variacion_libras_fuente",
        F.when(F.col("anio") == F.col("anio_ultimo"), F.col("variacion_libras_fuente")),
    )
    .withColumn(
        "participacion_libras_fuente",
        F.when(F.col("anio") == F.col("anio_ultimo"), F.col("participacion_libras_fuente")),
    )
    .withColumn(
        "dif_variacion_dolares",
        F.when(
            (F.col("anio") == F.col("anio_ultimo"))
            & F.col("variacion_dolares_fuente").isNotNull()
            & F.col("variacion_yoy_dolares").isNotNull(),
            F.abs(F.col("variacion_dolares_fuente") - F.col("variacion_yoy_dolares")),
        ),
    )
    .withColumn(
        "dif_variacion_libras",
        F.when(
            (F.col("anio") == F.col("anio_ultimo"))
            & F.col("variacion_libras_fuente").isNotNull()
            & F.col("variacion_yoy_libras").isNotNull(),
            F.abs(F.col("variacion_libras_fuente") - F.col("variacion_yoy_libras")),
        ),
    )
    .drop("anio_ultimo")
)

dup_market = (
    DF_MERCADO_BASE.groupBy("destino", "corte", "anio")
    .count()
    .filter(F.col("count") > 1)
    .count()
)
if dup_market:
    raise RuntimeError(
        f"Mercados no cumple el grano destino×corte×año: {dup_market} duplicados."
    )

negative_market = DF_MERCADO_BASE.filter(
    (F.col("libras") < 0) | (F.col("dolares") < 0)
).count()
if negative_market:
    raise RuntimeError(f"Mercados contiene {negative_market} registros con medidas negativas.")

# Auditoría de etiquetas y nivel de detalle.
unexpected_levels = df_mercado_todos.filter(
    ~F.col("nivel_destino").isin("DETALLADO", "AGREGADO_REGIONAL")
).count()
if unexpected_levels:
    raise RuntimeError(f"nivel_destino contiene {unexpected_levels} valores fuera de dominio.")

unexpected_aggregates = DF_MERCADO_AGREGADOS_EXCLUIDOS.filter(
    ~F.col("destino").isin(AGREGADOS_REGIONALES)
).count()
if unexpected_aggregates:
    raise RuntimeError(
        f"Se detectaron {unexpected_aggregates} agregados no contemplados por la taxonomía controlada."
    )

titulo("MERCADOS CNA — DESTINOS DETALLADOS")
print(f"Observaciones analíticas destino×corte×año : {DF_MERCADO_BASE.count()}")
print(f"Destinos detallados                        : {DF_MERCADO_BASE.select('destino').distinct().count()}")
print(f"Filas regionales excluidas                 : {DF_MERCADO_AGREGADOS_EXCLUIDOS.count()}")
print(f"Regiones excluidas                         : {AGREGADOS_REGIONALES}")
print("Regla de cálculo                           : solo nivel_destino=DETALLADO")
print("Clasificación geopolítica para cálculos    : NO")
print("Homologaciones Gold explícitas             :")
for origen, destino in NORMALIZACION_ETIQUETAS_DESTINO.items():
    print(f"  - {origen} -> {destino}")


# COMMAND ----------

# =============================================================================
# CELDA 6 — COMPARATIVO HISTÓRICO CON CORTES HOMOGÉNEOS
# =============================================================================

df_hist_mayo = (
    SILVER_DFS["exportaciones_mayo_historico"]
    .select(
        F.col("anio").cast("int").alias("anio"),
        F.lit("MAYO").alias("corte"),
        F.col("libras").cast("double").alias("libras"),
        F.col("dolares").cast("double").alias("dolares"),
        F.col("libras_variacion").cast("double").alias("variacion_libras_fuente"),
        F.col("dolares_variacion").cast("double").alias("variacion_dolares_fuente"),
    )
)

df_hist_acum = (
    SILVER_DFS["exportaciones_acumuladas_historico"]
    .withColumn("anio", F.regexp_extract("periodo", r"(\d{4})", 1).cast("int"))
    .select(
        "anio",
        F.lit("ACUM_ENE_MAY").alias("corte"),
        F.col("libras").cast("double").alias("libras"),
        F.col("dolares").cast("double").alias("dolares"),
        F.col("libras_variacion").cast("double").alias("variacion_libras_fuente"),
        F.col("dolares_variacion").cast("double").alias("variacion_dolares_fuente"),
    )
)

df_hist = df_hist_mayo.unionByName(df_hist_acum)

if df_hist.filter(F.col("anio").isNull()).count():
    raise RuntimeError("No se pudo derivar el año de todos los registros históricos.")

w_hist = Window.partitionBy("corte").orderBy("anio")

DF_COMPARATIVO_BASE = (
    df_hist
    .withColumn("anio_anterior", F.lag("anio").over(w_hist))
    .withColumn("libras_anterior", F.lag("libras").over(w_hist))
    .withColumn("dolares_anterior", F.lag("dolares").over(w_hist))
    .withColumn(
        "variacion_yoy_libras",
        F.when(
            (F.col("anio") - F.col("anio_anterior") == 1)
            & F.col("libras_anterior").isNotNull()
            & (F.col("libras_anterior") != 0),
            (F.col("libras") - F.col("libras_anterior")) / F.col("libras_anterior"),
        ),
    )
    .withColumn(
        "variacion_yoy_dolares",
        F.when(
            (F.col("anio") - F.col("anio_anterior") == 1)
            & F.col("dolares_anterior").isNotNull()
            & (F.col("dolares_anterior") != 0),
            (F.col("dolares") - F.col("dolares_anterior")) / F.col("dolares_anterior"),
        ),
    )
    .withColumn(
        "precio_promedio_usd_lb",
        F.when(
            F.col("libras").isNotNull() & (F.col("libras") > 0) & F.col("dolares").isNotNull(),
            F.col("dolares") / F.col("libras"),
        ),
    )
    .withColumn(
        "dif_variacion_libras",
        F.when(
            F.col("variacion_libras_fuente").isNotNull()
            & F.col("variacion_yoy_libras").isNotNull(),
            F.abs(F.col("variacion_libras_fuente") - F.col("variacion_yoy_libras")),
        ),
    )
    .withColumn(
        "dif_variacion_dolares",
        F.when(
            F.col("variacion_dolares_fuente").isNotNull()
            & F.col("variacion_yoy_dolares").isNotNull(),
            F.abs(F.col("variacion_dolares_fuente") - F.col("variacion_yoy_dolares")),
        ),
    )
    .drop("anio_anterior", "libras_anterior", "dolares_anterior")
)

dup_hist = (
    DF_COMPARATIVO_BASE.groupBy("anio", "corte")
    .count().filter(F.col("count") > 1).count()
)
if dup_hist:
    raise RuntimeError(f"Comparativo histórico no cumple grano año×corte: {dup_hist} duplicados.")

if DF_COMPARATIVO_BASE.filter((F.col("libras") < 0) | (F.col("dolares") < 0)).count():
    raise RuntimeError("Comparativo histórico contiene magnitudes negativas.")

titulo("COMPARATIVO HISTÓRICO")
print(f"Registros : {DF_COMPARATIVO_BASE.count()}")
DF_COMPARATIVO_BASE.orderBy("corte", "anio").show(20, truncate=False)


# COMMAND ----------

# =============================================================================
# CELDA 7 — DIMENSIONES CON CLAVES ESTABLES
# =============================================================================

# Dimensión fecha: únicamente fechas mensuales reales.
date_candidates = (
    DF_NACIONAL_BASE.select("fecha_periodo")
    .unionByName(
        DF_MERCADO_BASE.select(
            F.make_date(F.col("anio"), F.lit(5), F.lit(1)).alias("fecha_periodo")
        )
    )
    .unionByName(
        DF_COMPARATIVO_BASE.select(
            F.make_date(F.col("anio"), F.lit(5), F.lit(1)).alias("fecha_periodo")
        )
    )
    .filter(F.col("fecha_periodo").isNotNull())
    .distinct()
)

DF_DIM_FECHA = (
    date_candidates
    .withColumn(
        "fecha_key",
        (
            F.year("fecha_periodo") * F.lit(10000)
            + F.month("fecha_periodo") * F.lit(100)
            + F.dayofmonth("fecha_periodo")
        ).cast("int"),
    )
    .withColumn("anio", F.year("fecha_periodo"))
    .withColumn("mes_numero", F.month("fecha_periodo"))
    .withColumn(
        "mes_nombre",
        F.element_at(
            F.array(*[
                F.lit(x) for x in [
                    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
                ]
            ]),
            F.col("mes_numero"),
        ),
    )
    .withColumn("trimestre", F.quarter("fecha_periodo"))
    .withColumn("semestre", F.when(F.col("mes_numero") <= 6, 1).otherwise(2))
    .withColumn("anio_mes", F.date_format("fecha_periodo", "yyyy-MM"))
    .select(
        "fecha_key", "fecha_periodo", "anio", "mes_numero",
        "mes_nombre", "trimestre", "semestre", "anio_mes",
    )
)

if DF_DIM_FECHA.groupBy("fecha_key").count().filter(F.col("count") > 1).count():
    raise RuntimeError("fecha_key no es única.")
if DF_DIM_FECHA.groupBy("fecha_periodo").count().filter(F.col("count") > 1).count():
    raise RuntimeError("fecha_periodo no es única.")

# Dimensión de destino: evita afirmar que toda unidad reportada por CNA sea un
# Estado soberano. `destino_key` depende solo de la etiqueta analítica homologada.
DF_DIM_DESTINO = (
    DF_MERCADO_BASE
    .groupBy("destino")
    .agg(
        F.concat_ws(
            " | ",
            F.sort_array(F.collect_set("destino_fuente"))
        ).alias("etiquetas_fuente")
    )
    .withColumn("destino_key", F.sha2(F.col("destino"), 256))
    .withColumn("nivel_destino", F.lit("DETALLADO"))
    .select("destino_key", "destino", "etiquetas_fuente", "nivel_destino")
)

if DF_DIM_DESTINO.groupBy("destino_key").count().filter(F.col("count") > 1).count():
    raise RuntimeError("Colisión detectada en destino_key SHA-256.")
if DF_DIM_DESTINO.groupBy("destino").count().filter(F.col("count") > 1).count():
    raise RuntimeError("La dimensión destino contiene etiquetas analíticas duplicadas.")

# Dimensión corte: semántica separada del calendario.
CORTE_SCHEMA = T.StructType([
    T.StructField("corte_key", T.IntegerType(), False),
    T.StructField("corte", T.StringType(), False),
    T.StructField("descripcion", T.StringType(), False),
    T.StructField("mes_fin", T.IntegerType(), False),
])

DF_DIM_CORTE = spark.createDataFrame([
    (1, "MAYO", "Valor correspondiente únicamente al mes de mayo", 5),
    (2, "ACUM_ENE_MAY", "Valor acumulado desde enero hasta mayo", 5),
], schema=CORTE_SCHEMA)

titulo("DIMENSIONES GOLD CNA")
print(f"cna_dim_fecha   : {DF_DIM_FECHA.count()} filas")
print(f"cna_dim_destino : {DF_DIM_DESTINO.count()} filas (destinos detallados)")
print(f"cna_dim_corte   : {DF_DIM_CORTE.count()} filas")


# COMMAND ----------

# =============================================================================
# CELDA 8 — CONSTRUCCIÓN DE TABLAS DE HECHOS
# =============================================================================

# Hecho nacional mensual.
DF_FACT_NACIONAL = (
    DF_NACIONAL_BASE.alias("n")
    .join(
        DF_DIM_FECHA.select("fecha_key", "fecha_periodo").alias("d"),
        on="fecha_periodo",
        how="left",
    )
    .select(
        "fecha_key", "fecha_periodo", "anio", "mes_numero",
        "libras", "dolares", "precio_promedio_usd_lb",
        "fuente_libras", "fuente_dolares", "estado_reconciliacion",
    )
)

# Hecho mercados: grano destino detallado × corte × año.
DF_FACT_MERCADOS = (
    DF_MERCADO_BASE.alias("m")
    .join(
        DF_DIM_DESTINO.select("destino_key", "destino").alias("d"),
        on="destino",
        how="left",
    )
    .join(DF_DIM_CORTE.select("corte_key", "corte").alias("c"), on="corte", how="left")
    .withColumn("fecha_periodo", F.make_date(F.col("anio"), F.lit(5), F.lit(1)))
    .join(DF_DIM_FECHA.select("fecha_key", "fecha_periodo"), on="fecha_periodo", how="left")
    .select(
        "destino_key", "corte_key", "fecha_key",
        "destino_fuente", "destino", "nivel_destino", "corte", "anio",
        "libras", "dolares", "precio_promedio_usd_lb",
        "participacion_libras", "ranking_dolares", "ranking_libras",
        "variacion_yoy_dolares", "variacion_yoy_libras",
        "variacion_dolares_fuente", "variacion_libras_fuente",
        "participacion_libras_fuente",
        "dif_variacion_dolares", "dif_variacion_libras",
    )
)

# Hecho comparativo histórico.
DF_FACT_COMPARATIVO = (
    DF_COMPARATIVO_BASE.alias("h")
    .join(DF_DIM_CORTE.select("corte_key", "corte"), on="corte", how="left")
    .withColumn("fecha_periodo", F.make_date(F.col("anio"), F.lit(5), F.lit(1)))
    .join(DF_DIM_FECHA.select("fecha_key", "fecha_periodo"), on="fecha_periodo", how="left")
    .select(
        "fecha_key", "corte_key", "anio", "corte",
        "libras", "dolares", "precio_promedio_usd_lb",
        "variacion_yoy_libras", "variacion_yoy_dolares",
        "variacion_libras_fuente", "variacion_dolares_fuente",
        "dif_variacion_libras", "dif_variacion_dolares",
    )
)

# Validaciones de integridad referencial y grano.
FACT_CHECKS = []


def add_check(name: str, condition: bool, detail: str) -> None:
    FACT_CHECKS.append({"check": name, "status": "PASS" if condition else "FAIL", "detail": detail})
    print(f"{'PASS' if condition else 'FAIL'} | {name:35s} | {detail}")


titulo("VALIDACIÓN DE HECHOS")

add_check(
    "fact_nacional_fecha_fk",
    DF_FACT_NACIONAL.filter(F.col("fecha_key").isNull()).count() == 0,
    "0 fechas huérfanas esperadas",
)
add_check(
    "fact_nacional_grano",
    DF_FACT_NACIONAL.groupBy("fecha_key").count().filter(F.col("count") > 1).count() == 0,
    "1 fila por fecha_key",
)
add_check(
    "fact_mercados_fk",
    DF_FACT_MERCADOS.filter(
        F.col("destino_key").isNull() | F.col("corte_key").isNull() | F.col("fecha_key").isNull()
    ).count() == 0,
    "0 claves huérfanas esperadas",
)
add_check(
    "fact_mercados_grano",
    DF_FACT_MERCADOS.groupBy("destino_key", "corte_key", "anio")
    .count().filter(F.col("count") > 1).count() == 0,
    "1 fila por destino detallado×corte×año",
)
add_check(
    "fact_comparativo_fk",
    DF_FACT_COMPARATIVO.filter(
        F.col("corte_key").isNull() | F.col("fecha_key").isNull()
    ).count() == 0,
    "0 claves huérfanas esperadas",
)
add_check(
    "fact_comparativo_grano",
    DF_FACT_COMPARATIVO.groupBy("corte_key", "anio")
    .count().filter(F.col("count") > 1).count() == 0,
    "1 fila por corte×año",
)

if any(x["status"] == "FAIL" for x in FACT_CHECKS):
    raise RuntimeError(f"Fallaron controles de hechos: {FACT_CHECKS}")

print("\n✓ Tablas de hechos construidas con grano e integridad referencial válidos.")


# COMMAND ----------

# =============================================================================
# CELDA 9 — SERVING NACIONAL: MENSUAL Y ANUAL
# =============================================================================

# Serving mensual completamente desnormalizado.
base_month = (
    DF_FACT_NACIONAL.alias("f")
    .join(
        DF_DIM_FECHA.select(
            "fecha_key", "mes_nombre", "trimestre", "semestre", "anio_mes"
        ).alias("d"),
        on="fecha_key",
        how="inner",
    )
)

prev_month = DF_FACT_NACIONAL.select(
    F.add_months("fecha_periodo", 1).alias("fecha_periodo"),
    F.col("libras").alias("libras_mes_anterior"),
    F.col("dolares").alias("dolares_mes_anterior"),
)

prev_year = DF_FACT_NACIONAL.select(
    F.add_months("fecha_periodo", 12).alias("fecha_periodo"),
    F.col("libras").alias("libras_mismo_mes_anterior"),
    F.col("dolares").alias("dolares_mismo_mes_anterior"),
)

DF_SERVING_NACIONAL_MENSUAL = (
    base_month
    .join(prev_month, on="fecha_periodo", how="left")
    .join(prev_year, on="fecha_periodo", how="left")
    .withColumn(
        "variacion_mom_libras",
        F.when(
            F.col("libras_mes_anterior").isNotNull()
            & (F.col("libras_mes_anterior") != 0)
            & F.col("libras").isNotNull(),
            (F.col("libras") - F.col("libras_mes_anterior")) / F.col("libras_mes_anterior"),
        ),
    )
    .withColumn(
        "variacion_mom_dolares",
        F.when(
            F.col("dolares_mes_anterior").isNotNull()
            & (F.col("dolares_mes_anterior") != 0)
            & F.col("dolares").isNotNull(),
            (F.col("dolares") - F.col("dolares_mes_anterior")) / F.col("dolares_mes_anterior"),
        ),
    )
    .withColumn(
        "variacion_yoy_libras",
        F.when(
            F.col("libras_mismo_mes_anterior").isNotNull()
            & (F.col("libras_mismo_mes_anterior") != 0)
            & F.col("libras").isNotNull(),
            (F.col("libras") - F.col("libras_mismo_mes_anterior")) / F.col("libras_mismo_mes_anterior"),
        ),
    )
    .withColumn(
        "variacion_yoy_dolares",
        F.when(
            F.col("dolares_mismo_mes_anterior").isNotNull()
            & (F.col("dolares_mismo_mes_anterior") != 0)
            & F.col("dolares").isNotNull(),
            (F.col("dolares") - F.col("dolares_mismo_mes_anterior")) / F.col("dolares_mismo_mes_anterior"),
        ),
    )
    .select(
        "fecha_periodo", "anio", "mes_numero", "mes_nombre", "anio_mes",
        "trimestre", "semestre",
        "libras", "dolares", "precio_promedio_usd_lb",
        "variacion_mom_libras", "variacion_mom_dolares",
        "variacion_yoy_libras", "variacion_yoy_dolares",
        "fuente_libras", "fuente_dolares", "estado_reconciliacion",
    )
    .orderBy("fecha_periodo")
)

# Serving anual con precio ponderado correcto.
df_anual_base = (
    DF_FACT_NACIONAL
    .groupBy("anio")
    .agg(
        F.sum("libras").alias("libras_totales"),
        F.sum("dolares").alias("dolares_totales"),
        F.countDistinct("mes_numero").alias("meses_reportados"),
    )
    .withColumn("anio_completo", F.col("meses_reportados") == 12)
    .withColumn(
        "precio_promedio_ponderado_usd_lb",
        F.when(
            F.col("libras_totales").isNotNull()
            & (F.col("libras_totales") > 0)
            & F.col("dolares_totales").isNotNull(),
            F.col("dolares_totales") / F.col("libras_totales"),
        ),
    )
)

prev_anual = df_anual_base.select(
    (F.col("anio") + 1).alias("anio"),
    F.col("libras_totales").alias("libras_anio_anterior"),
    F.col("dolares_totales").alias("dolares_anio_anterior"),
    F.col("anio_completo").alias("anio_anterior_completo"),
)

DF_SERVING_NACIONAL_ANUAL = (
    df_anual_base
    .join(prev_anual, on="anio", how="left")
    .withColumn(
        "variacion_yoy_libras",
        F.when(
            F.col("anio_completo")
            & F.col("anio_anterior_completo")
            & F.col("libras_anio_anterior").isNotNull()
            & (F.col("libras_anio_anterior") != 0),
            (F.col("libras_totales") - F.col("libras_anio_anterior")) / F.col("libras_anio_anterior"),
        ),
    )
    .withColumn(
        "variacion_yoy_dolares",
        F.when(
            F.col("anio_completo")
            & F.col("anio_anterior_completo")
            & F.col("dolares_anio_anterior").isNotNull()
            & (F.col("dolares_anio_anterior") != 0),
            (F.col("dolares_totales") - F.col("dolares_anio_anterior")) / F.col("dolares_anio_anterior"),
        ),
    )
    .select(
        "anio", "meses_reportados", "anio_completo",
        "libras_totales", "dolares_totales",
        "precio_promedio_ponderado_usd_lb",
        "variacion_yoy_libras", "variacion_yoy_dolares",
    )
    .orderBy("anio")
)

titulo("SERVING NACIONAL")
print(f"Nacional mensual : {DF_SERVING_NACIONAL_MENSUAL.count()} filas")
print(f"Nacional anual   : {DF_SERVING_NACIONAL_ANUAL.count()} filas")


# COMMAND ----------

# =============================================================================
# CELDA 10 — SERVING DE DESTINOS INDEPENDIENTE PARA TABLEAU
# =============================================================================

DF_SERVING_MERCADOS = (
    DF_FACT_MERCADOS
    .select(
        "destino_fuente", "destino", "nivel_destino", "corte", "anio",
        "libras", "dolares", "precio_promedio_usd_lb",
        "participacion_libras",
        "ranking_dolares", "ranking_libras",
        "variacion_yoy_libras", "variacion_yoy_dolares",
        "variacion_libras_fuente", "variacion_dolares_fuente",
        "participacion_libras_fuente",
        "dif_variacion_libras", "dif_variacion_dolares",
    )
    .withColumn(
        "validacion_variacion_libras",
        F.when(
            F.col("dif_variacion_libras").isNull(),
            F.lit("NO_COMPARABLE"),
        ).when(
            F.col("dif_variacion_libras")
            <= F.greatest(
                F.lit(ABS_VARIATION_TOL),
                F.abs(F.col("variacion_yoy_libras")) * F.lit(REL_VARIATION_TOL),
            ),
            F.lit("COINCIDE"),
        ).otherwise(F.lit("REVISAR")),
    )
    .withColumn(
        "validacion_variacion_dolares",
        F.when(
            F.col("dif_variacion_dolares").isNull(),
            F.lit("NO_COMPARABLE"),
        ).when(
            F.col("dif_variacion_dolares")
            <= F.greatest(
                F.lit(ABS_VARIATION_TOL),
                F.abs(F.col("variacion_yoy_dolares")) * F.lit(REL_VARIATION_TOL),
            ),
            F.lit("COINCIDE"),
        ).otherwise(F.lit("REVISAR")),
    )
    .orderBy("corte", "anio", "ranking_dolares", "destino")
)

titulo("SERVING DESTINOS")
print(f"Filas : {DF_SERVING_MERCADOS.count()}")
print("Grano : destino detallado × corte × año")
print("Regla : nivel_destino=DETALLADO")


# COMMAND ----------

# =============================================================================
# CELDA 10B — SERVING DE AGREGADOS REGIONALES INDEPENDIENTE PARA TABLEAU
# =============================================================================
# Los agregados regionales publicados por CNA se conservan, pero NUNCA se
# mezclan con destinos detallados en rankings, denominadores o participaciones.
# Esta hoja existe para análisis regional separado y trazabilidad de la fuente.

w_region_lb = Window.partitionBy("corte", "anio").orderBy(
    F.col("libras").desc_nulls_last(), F.col("destino")
)
w_region_usd = Window.partitionBy("corte", "anio").orderBy(
    F.col("dolares").desc_nulls_last(), F.col("destino")
)
w_region_yoy = Window.partitionBy("destino", "corte").orderBy("anio")

DF_SERVING_REGIONES = (
    DF_MERCADO_AGREGADOS_EXCLUIDOS
    .withColumn(
        "precio_promedio_usd_lb",
        F.when(
            F.col("libras").isNotNull() & (F.col("libras") > 0) & F.col("dolares").isNotNull(),
            F.col("dolares") / F.col("libras"),
        ),
    )
    .withColumn("ranking_libras_regional", F.dense_rank().over(w_region_lb))
    .withColumn("ranking_dolares_regional", F.dense_rank().over(w_region_usd))
    .withColumn("anio_anterior", F.lag("anio").over(w_region_yoy))
    .withColumn("libras_anterior", F.lag("libras").over(w_region_yoy))
    .withColumn("dolares_anterior", F.lag("dolares").over(w_region_yoy))
    .withColumn(
        "variacion_yoy_libras",
        F.when(
            (F.col("anio") - F.col("anio_anterior") == 1)
            & F.col("libras_anterior").isNotNull() & (F.col("libras_anterior") != 0)
            & F.col("libras").isNotNull(),
            (F.col("libras") - F.col("libras_anterior")) / F.col("libras_anterior"),
        ),
    )
    .withColumn(
        "variacion_yoy_dolares",
        F.when(
            (F.col("anio") - F.col("anio_anterior") == 1)
            & F.col("dolares_anterior").isNotNull() & (F.col("dolares_anterior") != 0)
            & F.col("dolares").isNotNull(),
            (F.col("dolares") - F.col("dolares_anterior")) / F.col("dolares_anterior"),
        ),
    )
    .drop("anio_anterior", "libras_anterior", "dolares_anterior")
    .withColumn("uso_analitico", F.lit("AGREGADO_REGIONAL_SEPARADO"))
    .select(
        "destino_fuente", "destino", "nivel_destino", "corte", "anio",
        "libras", "dolares", "precio_promedio_usd_lb",
        "ranking_libras_regional", "ranking_dolares_regional",
        "variacion_yoy_libras", "variacion_yoy_dolares", "uso_analitico",
    )
    .orderBy("corte", "anio", "ranking_dolares_regional", "destino")
)

# Contratos mínimos: grano único, solo regiones controladas y sin mezcla con detalle.
if DF_SERVING_REGIONES.groupBy("destino", "corte", "anio").count().filter(F.col("count") > 1).count():
    raise RuntimeError("El serving regional contiene duplicados destino×corte×año.")
if DF_SERVING_REGIONES.filter(F.col("nivel_destino") != "AGREGADO_REGIONAL").count():
    raise RuntimeError("El serving regional contiene filas que no son AGREGADO_REGIONAL.")
if DF_SERVING_REGIONES.filter(~F.col("destino").isin(AGREGADOS_REGIONALES)).count():
    raise RuntimeError("El serving regional contiene etiquetas fuera del catálogo controlado.")

titulo("SERVING REGIONES AGREGADAS")
print(f"Filas : {DF_SERVING_REGIONES.count()}")
print("Grano : agregado regional × corte × año")
print(f"Catálogo controlado: {AGREGADOS_REGIONALES}")
print("Regla : se exporta separado; NO se suma ni se mezcla con destinos detallados.")


# COMMAND ----------

# =============================================================================
# CELDA 11 — SERVING COMPARATIVO HISTÓRICO Y KPI GLOBAL
# =============================================================================

DF_SERVING_COMPARATIVO = (
    DF_FACT_COMPARATIVO
    .select(
        "anio", "corte",
        "libras", "dolares", "precio_promedio_usd_lb",
        "variacion_yoy_libras", "variacion_yoy_dolares",
        "variacion_libras_fuente", "variacion_dolares_fuente",
        "dif_variacion_libras", "dif_variacion_dolares",
    )
    .withColumn(
        "validacion_variacion_libras",
        F.when(F.col("dif_variacion_libras").isNull(), F.lit("NO_COMPARABLE"))
        .when(
            F.col("dif_variacion_libras")
            <= F.greatest(
                F.lit(ABS_VARIATION_TOL),
                F.abs(F.col("variacion_yoy_libras")) * F.lit(REL_VARIATION_TOL),
            ),
            F.lit("COINCIDE"),
        )
        .otherwise(F.lit("REVISAR")),
    )
    .withColumn(
        "validacion_variacion_dolares",
        F.when(F.col("dif_variacion_dolares").isNull(), F.lit("NO_COMPARABLE"))
        .when(
            F.col("dif_variacion_dolares")
            <= F.greatest(
                F.lit(ABS_VARIATION_TOL),
                F.abs(F.col("variacion_yoy_dolares")) * F.lit(REL_VARIATION_TOL),
            ),
            F.lit("COINCIDE"),
        )
        .otherwise(F.lit("REVISAR")),
    )
    .orderBy("corte", "anio")
)

# Último período mensual disponible.
latest_month = (
    DF_SERVING_NACIONAL_MENSUAL
    .orderBy(F.col("fecha_periodo").desc())
    .limit(1)
    .first()
)
if latest_month is None:
    raise RuntimeError("No existe un último período mensual para KPI.")

# Último corte acumulado comparable.
latest_acum = (
    DF_SERVING_COMPARATIVO
    .filter(F.col("corte") == "ACUM_ENE_MAY")
    .orderBy(F.col("anio").desc())
    .limit(1)
    .first()
)

# Destino líder del último acumulado.
latest_market_year_row = (
    DF_SERVING_MERCADOS
    .filter(F.col("corte") == "ACUM_ENE_MAY")
    .agg(F.max("anio").alias("anio"))
    .first()
)
latest_market_year = latest_market_year_row["anio"] if latest_market_year_row else None

top_destination = None
if latest_market_year is not None:
    top_destination = (
        DF_SERVING_MERCADOS
        .filter(
            (F.col("corte") == "ACUM_ENE_MAY")
            & (F.col("anio") == F.lit(latest_market_year))
        )
        .orderBy(F.col("dolares").desc_nulls_last(), F.col("destino"))
        .limit(1)
        .first()
    )

years = DF_SERVING_NACIONAL_ANUAL.agg(
    F.min("anio").alias("anio_inicio"),
    F.max("anio").alias("anio_fin"),
).first()

KPI_SCHEMA = T.StructType([
    T.StructField("anio_inicio_historico", T.IntegerType(), True),
    T.StructField("anio_mas_reciente", T.IntegerType(), True),
    T.StructField("ultimo_periodo_mensual", T.DateType(), True),
    T.StructField("libras_ultimo_mes", T.DoubleType(), True),
    T.StructField("dolares_ultimo_mes", T.DoubleType(), True),
    T.StructField("precio_ultimo_mes_usd_lb", T.DoubleType(), True),
    T.StructField("anio_ultimo_acumulado_ene_may", T.IntegerType(), True),
    T.StructField("libras_ultimo_acumulado_ene_may", T.DoubleType(), True),
    T.StructField("dolares_ultimo_acumulado_ene_may", T.DoubleType(), True),
    T.StructField("crecimiento_yoy_libras_acumulado_ene_may", T.DoubleType(), True),
    T.StructField("crecimiento_yoy_dolares_acumulado_ene_may", T.DoubleType(), True),
    T.StructField("destino_top_acumulado_dolares", T.StringType(), True),
    T.StructField("dolares_destino_top_acumulado", T.DoubleType(), True),
])

kpi_row = [(
    int(years["anio_inicio"]) if years["anio_inicio"] is not None else None,
    int(years["anio_fin"]) if years["anio_fin"] is not None else None,
    latest_month["fecha_periodo"],
    float(latest_month["libras"]) if latest_month["libras"] is not None else None,
    float(latest_month["dolares"]) if latest_month["dolares"] is not None else None,
    float(latest_month["precio_promedio_usd_lb"]) if latest_month["precio_promedio_usd_lb"] is not None else None,
    int(latest_acum["anio"]) if latest_acum is not None and latest_acum["anio"] is not None else None,
    float(latest_acum["libras"]) if latest_acum is not None and latest_acum["libras"] is not None else None,
    float(latest_acum["dolares"]) if latest_acum is not None and latest_acum["dolares"] is not None else None,
    float(latest_acum["variacion_yoy_libras"]) if latest_acum is not None and latest_acum["variacion_yoy_libras"] is not None else None,
    float(latest_acum["variacion_yoy_dolares"]) if latest_acum is not None and latest_acum["variacion_yoy_dolares"] is not None else None,
    str(top_destination["destino"]) if top_destination is not None else None,
    float(top_destination["dolares"]) if top_destination is not None and top_destination["dolares"] is not None else None,
)]

DF_SERVING_KPI = spark.createDataFrame(kpi_row, schema=KPI_SCHEMA)

titulo("SERVING COMPARATIVO Y KPI")
print(f"Comparativo histórico : {DF_SERVING_COMPARATIVO.count()} filas")
print(f"KPI resumen            : {DF_SERVING_KPI.count()} fila")


# COMMAND ----------

# =============================================================================
# CELDA 12 — REGLAS DE CALIDAD GOLD BLOQUEANTES Y ADVERTENCIAS
# =============================================================================

QUALITY_RESULTS: List[Dict[str, Any]] = []


def quality(rule: str, severity: str, violations: int, checked: int, detail: str) -> None:
    status = "PASS" if violations == 0 else ("FAIL" if severity == "CRITICAL" else "WARN")
    QUALITY_RESULTS.append({
        "rule": rule,
        "severity": severity,
        "status": status,
        "violations": int(violations),
        "checked": int(checked),
        "detail": detail,
    })
    print(
        f"{status:4s} | {rule:40s} | "
        f"violaciones={violations:4d} | revisados={checked:4d}"
    )


titulo("REGLAS DE CALIDAD GOLD CNA")

# Q1. Medidas nacionales no negativas.
q1_checked = DF_FACT_NACIONAL.count()
q1_viol = DF_FACT_NACIONAL.filter(
    (F.col("libras") < 0) | (F.col("dolares") < 0)
).count()
quality("Q1_NACIONAL_NO_NEGATIVO", "CRITICAL", q1_viol, q1_checked, "Libras y dólares deben ser >= 0.")

# Q2. Precio nacional = dólares/libras.
eligible_price = DF_FACT_NACIONAL.filter(
    F.col("libras").isNotNull() & (F.col("libras") > 0)
    & F.col("dolares").isNotNull()
    & F.col("precio_promedio_usd_lb").isNotNull()
)
q2_viol = eligible_price.filter(
    F.abs(
        F.col("precio_promedio_usd_lb") - F.col("dolares") / F.col("libras")
    ) > F.lit(1e-9)
).count()
quality("Q2_PRECIO_NACIONAL", "CRITICAL", q2_viol, eligible_price.count(), "Precio = dólares/libras.")

# Q3. Precio anual ponderado.
eligible_annual = DF_SERVING_NACIONAL_ANUAL.filter(
    F.col("libras_totales").isNotNull() & (F.col("libras_totales") > 0)
    & F.col("dolares_totales").isNotNull()
)
q3_viol = eligible_annual.filter(
    F.abs(
        F.col("precio_promedio_ponderado_usd_lb")
        - F.col("dolares_totales") / F.col("libras_totales")
    ) > F.lit(1e-9)
).count()
quality("Q3_PRECIO_ANUAL_PONDERADO", "CRITICAL", q3_viol, eligible_annual.count(), "SUM(dólares)/SUM(libras).")

# Q4. Participación de mercado en rango.
q4_checked = DF_FACT_MERCADOS.filter(F.col("participacion_libras").isNotNull()).count()
q4_viol = DF_FACT_MERCADOS.filter(
    F.col("participacion_libras").isNotNull()
    & (~F.col("participacion_libras").between(0.0, 1.0))
).count()
quality("Q4_PARTICIPACION_RANGO", "CRITICAL", q4_viol, q4_checked, "Participación en [0,1].")

# Q5. Suma de participaciones ~ 1 cuando existe volumen.
part_sums = (
    DF_FACT_MERCADOS
    .groupBy("corte", "anio")
    .agg(
        F.sum("participacion_libras").alias("participacion_total"),
        F.sum("libras").alias("libras_total"),
    )
    .filter(F.col("libras_total") > 0)
)
q5_viol = part_sums.filter(
    F.abs(F.col("participacion_total") - F.lit(1.0)) > F.lit(1e-8)
).count()
quality("Q5_PARTICIPACION_TOTAL", "CRITICAL", q5_viol, part_sums.count(), "Participaciones calculadas deben sumar 1.")


# Q5b. Reconciliación externa de destinos detallados contra total nacional del mismo corte.
# Esta prueba evita el falso positivo de una participación que suma 1 sobre un denominador semánticamente incorrecto.
market_totals = (
    DF_FACT_MERCADOS.groupBy("corte", "anio")
    .agg(F.sum("libras").alias("mercado_libras"), F.sum("dolares").alias("mercado_dolares"))
)
national_cut_totals = DF_COMPARATIVO_BASE.select(
    "corte", "anio",
    F.col("libras").alias("nacional_libras"),
    F.col("dolares").alias("nacional_dolares"),
)
market_reconciliation = market_totals.join(national_cut_totals, ["corte", "anio"], "inner")
market_reconciliation = (
    market_reconciliation
    .withColumn("dif_libras", F.abs(F.col("mercado_libras") - F.col("nacional_libras")))
    .withColumn("dif_dolares", F.abs(F.col("mercado_dolares") - F.col("nacional_dolares")))
)
q5b_viol = market_reconciliation.filter(
    (F.col("dif_libras") > F.lit(ABS_RECON_TOL)) |
    (F.col("dif_dolares") > F.lit(ABS_RECON_TOL))
).count()
quality(
    "Q5B_MERCADOS_RECONCILIA_NACIONAL", "CRITICAL", q5b_viol, market_reconciliation.count(),
    "La suma de destinos con nivel_destino=DETALLADO debe reconciliar con el total nacional para el mismo año y corte."
)

# Q5c. Solo el nivel DETALLADO puede entrar al hecho analítico de destinos.
q5c_viol = DF_FACT_MERCADOS.filter(
    F.col("nivel_destino") != "DETALLADO"
).count()
quality(
    "Q5C_SOLO_DETALLE_EN_FACT_MERCADOS",
    "CRITICAL",
    q5c_viol,
    DF_FACT_MERCADOS.count(),
    "El hecho analítico no admite agregados regionales.",
)

# Q5d. La etiqueta analítica no debe contener sufijos históricos homologados.
# `destino_fuente` sí puede conservarlos para trazabilidad.
q5d_viol = DF_FACT_MERCADOS.filter(
    F.col("destino").rlike(r"\(COLONIA FRANCIA\)")
).count()
quality(
    "Q5D_ETIQUETAS_DESTINO_HOMOLOGADAS",
    "CRITICAL",
    q5d_viol,
    DF_FACT_MERCADOS.count(),
    "La etiqueta analítica `destino` no debe contener el sufijo '(COLONIA FRANCIA)'.",
)

# Q5e. Dominio controlado del nivel de destino en todo el universo procesado.
q5e_viol = df_mercado_todos.filter(
    ~F.col("nivel_destino").isin("DETALLADO", "AGREGADO_REGIONAL")
).count()
quality(
    "Q5E_NIVEL_DESTINO_VALIDO",
    "CRITICAL",
    q5e_viol,
    df_mercado_todos.count(),
    "nivel_destino solo admite DETALLADO o AGREGADO_REGIONAL.",
)

# Q5f. Los agregados excluidos deben pertenecer al catálogo regional explícito.
q5f_viol = DF_MERCADO_AGREGADOS_EXCLUIDOS.filter(
    ~F.col("destino").isin(AGREGADOS_REGIONALES)
).count()
quality(
    "Q5F_AGREGADOS_REGIONALES_CONTROLADOS",
    "CRITICAL",
    q5f_viol,
    DF_MERCADO_AGREGADOS_EXCLUIDOS.count(),
    f"Agregados admitidos={AGREGADOS_REGIONALES}.",
)

# Q5g. Toda fila analítica conserva una etiqueta de procedencia de Silver.
q5g_viol = DF_FACT_MERCADOS.filter(
    F.col("destino_fuente").isNull() | (F.trim(F.col("destino_fuente")) == "")
).count()
quality(
    "Q5G_DESTINO_FUENTE_TRAZABLE",
    "CRITICAL",
    q5g_viol,
    DF_FACT_MERCADOS.count(),
    "Cada destino analítico debe conservar su etiqueta de procedencia desde Silver.",
)

# Q5h. Toda divergencia entre etiqueta fuente y analítica debe provenir del
# diccionario explícito de homologación; no se permiten transformaciones ocultas.
allowed_source_changes = list(NORMALIZACION_ETIQUETAS_DESTINO.keys())
q5h_viol = DF_FACT_MERCADOS.filter(
    (F.col("destino_fuente") != F.col("destino"))
    & (~F.col("destino_fuente").isin(allowed_source_changes))
).count()
quality(
    "Q5H_HOMOLOGACION_EXPLICITA",
    "CRITICAL",
    q5h_viol,
    DF_FACT_MERCADOS.count(),
    "Toda homologación de etiqueta debe estar declarada en NORMALIZACION_ETIQUETAS_DESTINO.",
)

# Q6. No comparar año parcial vs completo en resumen anual.
q6_viol = DF_SERVING_NACIONAL_ANUAL.filter(
    (~F.col("anio_completo")) &
    (F.col("variacion_yoy_libras").isNotNull() | F.col("variacion_yoy_dolares").isNotNull())
).count()
quality("Q6_YOY_ANUAL_SOLO_COMPLETOS", "CRITICAL", q6_viol, DF_SERVING_NACIONAL_ANUAL.count(), "YoY anual solo para dos años completos.")

# Q7. Reconciliaciones CNA divergentes: advertencia, no corrección oculta.
q7_viol = DF_RECONCILIACION_NACIONAL.filter(
    F.col("estado_reconciliacion") == "DIFIERE_PRIORIZA_DIRECTO"
).count()
quality(
    "Q7_RECONCILIACION_ESTRUCTURAS",
    "WARNING",
    q7_viol,
    DF_RECONCILIACION_NACIONAL.count(),
    "Las divergencias se conservan en cna_reconciliacion_nacional y la regla canónica queda documentada.",
)

# Q8. Diferencias entre variaciones de mercado reportadas y calculadas.
q8_viol = DF_SERVING_MERCADOS.filter(
    (F.col("validacion_variacion_libras") == "REVISAR")
    | (F.col("validacion_variacion_dolares") == "REVISAR")
).count()
quality(
    "Q8_VARIACIONES_MERCADO",
    "WARNING",
    q8_viol,
    DF_SERVING_MERCADOS.count(),
    "Comparación de variaciones publicadas por CNA contra cálculo Gold.",
)

critical_failures = [
    x for x in QUALITY_RESULTS
    if x["severity"] == "CRITICAL" and x["status"] == "FAIL"
]
if critical_failures:
    raise RuntimeError(f"Gold CNA no supera las reglas críticas: {critical_failures}")

print("\n✓ Reglas críticas Gold superadas.")
print("Las advertencias se conservan como evidencia; no se interpretan como 100% de calidad del dominio.")


# COMMAND ----------

# =============================================================================
# CELDA 13 — PERSISTENCIA DELTA Y LINEAGE SILVER→GOLD
# =============================================================================

GOLD_DATAFRAMES = {
    "dim_fecha": DF_DIM_FECHA,
    "dim_destino": DF_DIM_DESTINO,
    "dim_corte": DF_DIM_CORTE,
    "fact_nacional": DF_FACT_NACIONAL,
    "fact_mercados": DF_FACT_MERCADOS,
    "fact_comparativo": DF_FACT_COMPARATIVO,
    "reconciliacion_nacional": DF_RECONCILIACION_NACIONAL,
    "audit_mercados_agregados": DF_MERCADO_AGREGADOS_EXCLUIDOS,
    "serving_nacional_mensual": DF_SERVING_NACIONAL_MENSUAL,
    "serving_nacional_anual": DF_SERVING_NACIONAL_ANUAL,
    "serving_mercados": DF_SERVING_MERCADOS,
    "serving_regiones": DF_SERVING_REGIONES,
    "serving_comparativo": DF_SERVING_COMPARATIVO,
    "serving_kpi": DF_SERVING_KPI,
}

PERSISTENCE_RESULTS: List[Dict[str, Any]] = []

titulo("PERSISTENCIA GOLD CNA EN UNITY CATALOG")

for logical_name, df in GOLD_DATAFRAMES.items():
    table_ref = GOLD_TABLES[logical_name]

    (
        df.write
        .format("delta")
        .mode("overwrite")
        .option("overwriteSchema", "true")
        .saveAsTable(table_ref)
    )

    props = {
        "pipeline.layer": "gold",
        "pipeline.source": SOURCE_CODE,
        "pipeline.version": PIPELINE_VERSION,
        "pipeline.gold_run_id": RUN_ID,
        "lineage.silver_run_id": SILVER_RUN_ID,
        "lineage.bronze_sha256": BRONZE_SHA256,
        "lineage.report_period": REPORT_PERIOD,
        "gold.table_role": GOLD_TABLE_ROLES[logical_name],
        "gold.config_sha256": CONFIG_HASH,
    }
    props_sql = ", ".join(
        f"'{sql_escape(k)}'='{sql_escape(v)}'"
        for k, v in props.items()
    )
    spark.sql(f"ALTER TABLE {table_ref} SET TBLPROPERTIES ({props_sql})")

    rows = spark.table(table_ref).count()
    PERSISTENCE_RESULTS.append({
        "logical_name": logical_name,
        "table": table_ref,
        "role": GOLD_TABLE_ROLES[logical_name],
        "rows": rows,
        "status": "WRITTEN",
    })
    print(f"PASS | {logical_name:30s} | {rows:5d} filas | {table_ref}")

print(f"\nTablas Gold persistidas: {len(PERSISTENCE_RESULTS)}")


# COMMAND ----------

# =============================================================================
# CELDA 14 — VALIDACIÓN POST-ESCRITURA Y QA TÉCNICO
# =============================================================================

POSTWRITE_RESULTS: List[Dict[str, Any]] = []
QA_RESULTS: List[Dict[str, Any]] = []

titulo("VALIDACIÓN POST-ESCRITURA GOLD")

for logical_name, source_df in GOLD_DATAFRAMES.items():
    table_ref = GOLD_TABLES[logical_name]
    persisted_df = spark.table(table_ref)

    source_count = source_df.count()
    persisted_count = persisted_df.count()

    same_columns = source_df.columns == persisted_df.columns
    same_schema = (
        [f.dataType.simpleString() for f in source_df.schema.fields]
        == [f.dataType.simpleString() for f in persisted_df.schema.fields]
    )

    source_minus = source_df.exceptAll(persisted_df).limit(1).count()
    persisted_minus = persisted_df.exceptAll(source_df).limit(1).count()
    same_data = source_minus == 0 and persisted_minus == 0

    detail = spark.sql(f"DESCRIBE DETAIL {table_ref}").first().asDict()
    is_delta = str(detail.get("format", "")).lower() == "delta"

    props = table_properties(table_ref)
    lineage_ok = (
        props.get("pipeline.layer") == "gold"
        and props.get("pipeline.source") == SOURCE_CODE
        and props.get("pipeline.gold_run_id") == RUN_ID
        and props.get("lineage.silver_run_id") == SILVER_RUN_ID
        and props.get("lineage.bronze_sha256") == BRONZE_SHA256
    )

    passed = (
        source_count == persisted_count
        and same_columns
        and same_schema
        and same_data
        and is_delta
        and lineage_ok
    )

    POSTWRITE_RESULTS.append({
        "logical_name": logical_name,
        "source_rows": source_count,
        "persisted_rows": persisted_count,
        "same_columns": same_columns,
        "same_schema": same_schema,
        "same_data": same_data,
        "is_delta": is_delta,
        "lineage_ok": lineage_ok,
        "status": "PASS" if passed else "FAIL",
    })

    print(
        f"{'PASS' if passed else 'FAIL'} | {logical_name:30s} | "
        f"filas={persisted_count:5d} | data_equal={same_data} | lineage={lineage_ok}"
    )

postwrite_failures = [x for x in POSTWRITE_RESULTS if x["status"] == "FAIL"]
if postwrite_failures:
    raise RuntimeError(f"Validación post-escritura fallida: {postwrite_failures}")


def qa(name: str, passed: bool, detail: str) -> None:
    QA_RESULTS.append({
        "test": name,
        "status": "PASS" if passed else "FAIL",
        "detail": detail,
    })


# Accesibilidad y prefijos source-specific.
for logical_name, table_ref in GOLD_TABLES.items():
    qa(
        f"access_{logical_name}",
        spark.catalog.tableExists(table_ref),
        table_ref,
    )

qa(
    "namespace_cna",
    all(".cna_" in table_ref for table_ref in GOLD_TABLES.values()),
    "Todas las tablas Gold del notebook usan prefijo cna_.",
)

qa(
    "tableau_tres_hojas_planas_separadas",
    list(SERVING_EXPORTS.keys()) == [
        "tabla_tableau_nacional",
        "tabla_tableau_destinos",
        "tabla_tableau_regiones",
    ],
    str(list(SERVING_EXPORTS.keys())),
)

qa(
    "semantica_destino_sin_pais_geopolitico",
    "destino" in spark.table(GOLD_TABLES["serving_mercados"]).columns
    and "pais" not in spark.table(GOLD_TABLES["serving_mercados"]).columns
    and "tipo_destino" not in spark.table(GOLD_TABLES["serving_mercados"]).columns,
    "Serving de mercados usa `destino`/`nivel_destino`; no clasifica geopolíticamente para cálculos.",
)

qa(
    "serving_regiones_solo_agregados",
    spark.table(GOLD_TABLES["serving_regiones"]).filter(
        F.col("nivel_destino") != "AGREGADO_REGIONAL"
    ).count() == 0,
    "La hoja regional contiene únicamente agregados regionales y está separada del detalle.",
)

qa(
    "kpi_una_fila",
    spark.table(GOLD_TABLES["serving_kpi"]).count() == 1,
    "cna_serving_kpi debe contener una sola fila.",
)

qa_failures = [x for x in QA_RESULTS if x["status"] == "FAIL"]

titulo("QA TÉCNICO GOLD")
for x in QA_RESULTS:
    print(f"{x['status']:4s} | {x['test']:35s} | {x['detail']}")

if qa_failures:
    raise RuntimeError(f"QA Gold detectó fallos: {qa_failures}")

print("\n✓ QA técnico Gold superado.")


# COMMAND ----------

# =============================================================================
# CELDA 15 — MANTENIMIENTO DELTA CON CRITERIO OBJETIVO
# =============================================================================

MAINTENANCE_RESULTS: List[Dict[str, Any]] = []

titulo("MANTENIMIENTO DELTA CONDICIONADO")

for logical_name, table_ref in GOLD_TABLES.items():
    detail = spark.sql(f"DESCRIBE DETAIL {table_ref}").first().asDict()
    size_bytes = int(detail.get("sizeInBytes") or 0)
    num_files = int(detail.get("numFiles") or 0)

    should_optimize = (
        size_bytes >= OPTIMIZE_MIN_BYTES
        or num_files >= OPTIMIZE_MIN_FILES
    )

    if not should_optimize:
        status = "SKIPPED_SMALL_TABLE"
        print(
            f"SKIP | {logical_name:30s} | "
            f"{size_bytes:,} bytes | {num_files} archivos"
        )
    else:
        try:
            spark.sql(f"OPTIMIZE {table_ref}")
            status = "OPTIMIZED"
            print(
                f"PASS | {logical_name:30s} | "
                f"{size_bytes:,} bytes | {num_files} archivos"
            )
        except Exception as exc:
            status = f"SKIPPED_{type(exc).__name__}"
            print(f"WARN | {logical_name:30s} | {status}")

    MAINTENANCE_RESULTS.append({
        "logical_name": logical_name,
        "table": table_ref,
        "size_bytes_before": size_bytes,
        "num_files_before": num_files,
        "status": status,
    })

print("\n✓ Mantenimiento evaluado con umbrales objetivos; no se optimizan tablas pequeñas por defecto.")


# COMMAND ----------

# =============================================================================
# CELDA 16 — MANIFIESTO GOLD, MÉTRICAS Y REPORTE FINAL ANTES DEL SERVING XLSX
# =============================================================================

RUN_FINISHED_AT_UTC = datetime.now(timezone.utc)
PIPELINE_DURATION_SECONDS = round(time.perf_counter() - RUN_STARTED_MONOTONIC, 6)

TABLE_METRICS = []
for logical_name, table_ref in GOLD_TABLES.items():
    df = spark.table(table_ref)
    detail = spark.sql(f"DESCRIBE DETAIL {table_ref}").first().asDict()
    TABLE_METRICS.append({
        "logical_name": logical_name,
        "table": table_ref,
        "role": GOLD_TABLE_ROLES[logical_name],
        "rows": df.count(),
        "columns": len(df.columns),
        "size_bytes": int(detail.get("sizeInBytes") or 0),
        "num_files": int(detail.get("numFiles") or 0),
    })

manifest = {
    "manifest_schema_version": GOLD_MANIFEST_SCHEMA_VERSION,
    "pipeline_version": PIPELINE_VERSION,
    "run_id": RUN_ID,
    "source": SOURCE_CODE,
    "status": "SUCCESS",
    "started_at_utc": RUN_STARTED_AT_UTC.isoformat(),
    "finished_at_utc": RUN_FINISHED_AT_UTC.isoformat(),
    "duration_seconds": PIPELINE_DURATION_SECONDS,
    "config_sha256": CONFIG_HASH,
    "silver_lineage": {
        "silver_run_id": SILVER_RUN_ID,
        "bronze_sha256": BRONZE_SHA256,
        "report_period": REPORT_PERIOD,
        "tables": SILVER_CONTRACT_RESULTS,
    },
    "reconciliation_rule": CONFIG_PAYLOAD["reconciliation_rule"],
    "reconciliation_counts": RECON_COUNTS,
    "profile_silver": PROFILE_RESULTS,
    "quality": QUALITY_RESULTS,
    "fact_checks": FACT_CHECKS,
    "persistence": PERSISTENCE_RESULTS,
    "postwrite_validation": POSTWRITE_RESULTS,
    "technical_qa": QA_RESULTS,
    "delta_maintenance": MAINTENANCE_RESULTS,
    "gold_tables": TABLE_METRICS,
    "serving_tables_for_tableau": SERVING_EXPORTS,
    "software_versions": software_versions(),
    "market_semantics": {
        "grain": "destino_detallado_x_corte_x_anio",
        "silver_source_field": "pais",
        "gold_source_label_field": "destino_fuente",
        "gold_analytical_field": "destino",
        "classification_field": "nivel_destino",
        "analytical_level": "DETALLADO",
        "excluded_level": "AGREGADO_REGIONAL",
        "excluded_regional_labels": AGREGADOS_REGIONALES,
        "excluded_regional_rows_from_detailed_model": DF_MERCADO_AGREGADOS_EXCLUIDOS.count(),
        "regional_aggregates_exported_separately": True,
        "regional_serving_table": GOLD_TABLES["serving_regiones"],
        "geopolitical_classification_used_for_calculation": False,
        "label_normalization": NORMALIZACION_ETIQUETAS_DESTINO,
        "normalization_audit": [r.asDict() for r in DF_DESTINO_NORMALIZACIONES_AUDIT.collect()],
        "source_labels_preserved_upstream": True,
    },
}

GOLD_MANIFEST_PATH = GOLD_METADATA_ROOT / f"gold_manifest_{RUN_ID}_CNA.json"
atomic_json_dump(manifest, GOLD_MANIFEST_PATH)

reloaded = json.loads(GOLD_MANIFEST_PATH.read_text(encoding="utf-8"))
if reloaded.get("run_id") != RUN_ID or reloaded.get("status") != "SUCCESS":
    raise RuntimeError("El manifiesto Gold no pudo verificarse después de su escritura.")

manifest_bytes = GOLD_MANIFEST_PATH.read_bytes()
manifest_sha = sha256_bytes(manifest_bytes)

titulo("REPORTE FINAL — GOLD CNA v2")
print("Estado                  : SUCCESS")
print(f"Run ID                  : {RUN_ID}")
print(f"Pipeline version        : {PIPELINE_VERSION}")
print(f"Silver run              : {SILVER_RUN_ID}")
print(f"Bronze SHA-256          : {BRONZE_SHA256}")
print(f"Período documental      : {REPORT_PERIOD}")
print(f"Duración medida         : {PIPELINE_DURATION_SECONDS:.3f} segundos")
print(f"Tablas Gold             : {len(GOLD_TABLES)}")
print(f"Hojas Tableau exportables: {len(SERVING_EXPORTS)}")
print(f"Reglas calidad          : {len(QUALITY_RESULTS)}")
print(f"Warnings calidad        : {sum(1 for x in QUALITY_RESULTS if x['status'] == 'WARN')}")
print(f"Tests QA                : {len(QA_RESULTS)}")
print(f"Manifiesto              : {GOLD_MANIFEST_PATH}")
print(f"SHA-256 manifiesto      : {manifest_sha}")

print("\nÚNICAS hojas planas que serán exportadas a Tableau en la ÚLTIMA CELDA:")
for sheet, table_ref in SERVING_EXPORTS.items():
    print(f"  - {sheet:24s} <- {table_ref}")

print("\n✓ MODELO GOLD CNA LISTO PARA GENERAR EL XLSX DE SERVING.")


# COMMAND ----------

# =============================================================================
# CELDA 17 — ÚNICA EXPORTACIÓN XLSX PARA TABLEAU
# =============================================================================
# El Excel final contiene EXACTAMENTE TRES hojas planas/autocontenidas:
#   1) tabla_tableau_nacional
#   2) tabla_tableau_destinos (solo nivel DETALLADO)
#   3) tabla_tableau_regiones (solo nivel AGREGADO_REGIONAL)
#
# Los agregados regionales se exportan por separado y NUNCA se mezclan con
# destinos detallados en denominadores, participaciones o rankings de mercado.
# No se exportan dimensiones, facts, reconciliaciones ni tablas auxiliares.
# El objetivo es reproducir el contrato de consumo que ya funcionaba en Tableau:
# cada hoja se utiliza como fuente plana y no necesita relaciones con otra hoja.

from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font


titulo("EXPORTACIÓN FINAL — TRES TABLAS PLANAS PARA TABLEAU")

EXPECTED_SHEETS = [
    "tabla_tableau_nacional",
    "tabla_tableau_destinos",
    "tabla_tableau_regiones",
]

if list(SERVING_EXPORTS.keys()) != EXPECTED_SHEETS:
    raise RuntimeError(
        f"Contrato Tableau alterado. Esperado={EXPECTED_SHEETS}; "
        f"obtenido={list(SERVING_EXPORTS.keys())}"
    )

pandas_sheets: Dict[str, pd.DataFrame] = {}

for sheet_name, table_ref in SERVING_EXPORTS.items():
    df = spark.table(table_ref)

    if df.count() <= 0:
        raise RuntimeError(f"La tabla plana para Tableau está vacía: {table_ref}")

    # Orden determinístico.
    if sheet_name == "tabla_tableau_nacional":
        df = df.orderBy("fecha_periodo")
    elif sheet_name == "tabla_tableau_destinos":
        df = df.orderBy("corte", "anio", "ranking_dolares", "destino")
    elif sheet_name == "tabla_tableau_regiones":
        df = df.orderBy("corte", "anio", "ranking_dolares_regional", "destino")

    pdf = df.toPandas()
    pandas_sheets[sheet_name] = pdf
    print(
        f"PASS | {sheet_name:24s} | "
        f"{len(pdf):5d} filas | {len(pdf.columns):2d} columnas"
    )

# Validar explícitamente que SOLO existen las dos tablas permitidas.
if set(pandas_sheets.keys()) != set(EXPECTED_SHEETS):
    raise RuntimeError(
        f"Se intentaron exportar hojas no autorizadas: {list(pandas_sheets.keys())}"
    )

buffer = BytesIO()

with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
    for sheet_name in EXPECTED_SHEETS:
        pdf = pandas_sheets[sheet_name]
        pdf.to_excel(writer, sheet_name=sheet_name, index=False)

        ws = writer.book[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for column_cells in ws.columns:
            max_len = 0
            for cell in list(column_cells)[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[column_cells[0].column_letter].width = min(
                max(max_len + 2, 10), 40
            )

payload = buffer.getvalue()
if not payload:
    raise RuntimeError("El XLSX generado está vacío.")

# Validación estructural antes de persistir.
wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)

if wb.sheetnames != EXPECTED_SHEETS:
    raise RuntimeError(
        f"El XLSX final debe contener exclusivamente {EXPECTED_SHEETS}; "
        f"obtenido={wb.sheetnames}"
    )

for sheet_name in EXPECTED_SHEETS:
    ws = wb[sheet_name]
    expected_rows = len(pandas_sheets[sheet_name])
    actual_rows = max(int(ws.max_row) - 1, 0)

    if actual_rows != expected_rows:
        raise RuntimeError(
            f"Conteo XLSX incorrecto en {sheet_name}: "
            f"esperado={expected_rows}, obtenido={actual_rows}"
        )

wb.close()

# Persistir un único archivo.
DELIVERABLE_ROOT.mkdir(parents=True, exist_ok=True)
EXCEL_PATH.write_bytes(payload)

if not EXCEL_PATH.exists() or EXCEL_PATH.stat().st_size != len(payload):
    raise RuntimeError("El XLSX no pudo verificarse después de persistirlo.")

excel_sha256 = sha256_bytes(payload)

print("\nArchivo Tableau creado y validado.")
print(f"Ruta       : {EXCEL_PATH}")
print(f"Tamaño     : {len(payload):,} bytes")
print(f"SHA-256    : {excel_sha256}")
print(f"Hojas      : {EXPECTED_SHEETS}")

print("\n✓ El archivo contiene EXACTAMENTE 3 hojas.")
print("✓ tabla_tableau_nacional es plana y autocontenida.")
print("✓ tabla_tableau_destinos es plana, autocontenida y contiene únicamente destinos detallados.")
print("✓ tabla_tableau_regiones es plana, autocontenida y contiene únicamente agregados regionales separados.")
print("✓ Los agregados regionales NO intervienen en los rankings ni participaciones de destinos detallados.")
print("✓ No se exportan dimensiones, facts ni tablas auxiliares.")
print("✓ Tableau no necesita relacionar estas hojas para usar cada fuente de forma independiente.")


# COMMAND ----------
