# Databricks notebook source

# =============================================================================
# CELDA 1 — IMPORTACIONES, CONFIGURACIÓN Y RELOJ DE EJECUCIÓN
# =============================================================================

import json
import re
import time
import hashlib
import unicodedata
import zipfile
import importlib.metadata as importlib_metadata

from pathlib import Path
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

%pip install openpyxl
from openpyxl import load_workbook

from pyspark.sql import DataFrame
from pyspark.sql import functions as F
from pyspark.sql import types as T

PIPELINE_VERSION = "2.1.0"
SILVER_MANIFEST_SCHEMA_VERSION = "1.0"
EXPECTED_BRONZE_STORAGE_LAYOUT_VERSION = "2.0"

CATALOG = "camaronera_2026"
BRONZE_SCHEMA = "bronce"
SILVER_SCHEMA = "plata"

BRONZE_VOLUME = "datosbce"
SILVER_METADATA_VOLUME = "metadatos_bce"

SOURCE_CODE = "BCE"
OUTPUT_TABLE = "bce_exportaciones_camaron_subpartida"

BRONZE_ROOT = Path(f"/Volumes/{CATALOG}/{BRONZE_SCHEMA}/{BRONZE_VOLUME}")
BRONZE_RAW_ROOT = BRONZE_ROOT / "raw"
BRONZE_METADATA_ROOT = BRONZE_ROOT / "_metadata"

SUCCESS_BRONZE_STATUSES = {
    "downloaded",
    "updated_version",
    "unchanged",
    "recovered_copy",
}

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}

# Tolerancias numéricas para verificaciones aritméticas.
ABS_TOL_USD_CONVERSION = 1e-6

# Mantenimiento Delta solo cuando exista una justificación objetiva.
ENABLE_DELTA_MAINTENANCE = True
OPTIMIZE_MIN_FILES = 10
OPTIMIZE_MIN_SIZE_MB = 128.0

RUN_ID = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
PIPELINE_STARTED_AT_UTC = datetime.now(timezone.utc)
PIPELINE_STARTED_MONOTONIC = time.perf_counter()


def titulo(texto: str) -> None:
    print("\n" + "=" * 100)
    print(texto)
    print("=" * 100)


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> Tuple[str, int]:
    digest = hashlib.sha256()
    size = 0
    with path.open("rb") as fh:
        while True:
            chunk = fh.read(chunk_size)
            if not chunk:
                break
            digest.update(chunk)
            size += len(chunk)
    return digest.hexdigest(), size


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def atomic_json_dump(payload: Dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + f".{RUN_ID}.tmp")
    with tmp.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2, default=str)
    tmp.replace(path)


def package_version(name: str) -> str:
    try:
        return importlib_metadata.version(name)
    except importlib_metadata.PackageNotFoundError:
        return "unknown"


def software_versions() -> Dict[str, str]:
    import os
    import platform
    versions = {
        "python": platform.python_version(),
        "spark": getattr(spark, "version", "unknown"),
        "databricks_runtime": os.getenv("DATABRICKS_RUNTIME_VERSION", "unknown"),
    }
    for pkg in ["pyspark", "pandas", "openpyxl", "delta-spark", "playwright"]:
        try:
            versions[pkg] = importlib_metadata.version(pkg)
        except Exception:
            versions[pkg] = "not_installed"
    try:
        versions["spark_databricks_version"] = spark.conf.get(
            "spark.databricks.clusterUsageTags.sparkVersion"
        )
    except Exception:
        versions["spark_databricks_version"] = "unavailable"
    return versions

def normalize_text_for_match(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text).strip().upper()
    return text


titulo("SILVER BCE — INICIO")
print(f"Run ID                 : {RUN_ID}")
print(f"Pipeline version       : {PIPELINE_VERSION}")
print(f"Bronze                 : {BRONZE_ROOT}")
print(f"Destino Silver         : {CATALOG}.{SILVER_SCHEMA}.{OUTPUT_TABLE}")


# COMMAND ----------

# =============================================================================
# CELDA 2 — SELECCIÓN BRONZE, LINEAGE Y VERIFICACIÓN SHA-256
# =============================================================================

def normalize_volume_path(value: str) -> Path:
    raw = str(value or "").strip()
    if raw.startswith("dbfs:/Volumes/"):
        raw = "/Volumes/" + raw[len("dbfs:/Volumes/"):]
    return Path(raw)


def load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def compatible_manifest_candidates() -> List[Path]:
    if not BRONZE_METADATA_ROOT.is_dir():
        return []
    return sorted(
        BRONZE_METADATA_ROOT.glob("ingestion_manifest_*_BCE.json"),
        key=lambda p: p.name,
        reverse=True,
    )


def select_from_manifest() -> Optional[Dict[str, Any]]:
    for manifest_path in compatible_manifest_candidates():
        try:
            manifest = load_json(manifest_path)

            if str(manifest.get("source") or "").upper() != SOURCE_CODE:
                continue

            layout = str(manifest.get("storage_layout_version") or "")
            if layout and layout != EXPECTED_BRONZE_STORAGE_LAYOUT_VERSION:
                continue

            records = manifest.get("records") or []
            if not records:
                continue

            # BCE Bronze tiene un snapshot por ejecución.
            record = records[0]
            status = str(record.get("status") or "")
            if status and status not in SUCCESS_BRONZE_STATUSES:
                continue

            stored_path = normalize_volume_path(record.get("stored_path"))
            if not stored_path.is_file():
                continue

            if stored_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue

            return {
                "selection_mode": "bronze_manifest",
                "manifest_path": str(manifest_path),
                "manifest": manifest,
                "record": record,
                "stored_path": stored_path,
            }

        except Exception as exc:
            print(f"WARN | Manifiesto ignorado {manifest_path.name}: {exc}")

    return None


def select_fallback_raw() -> Dict[str, Any]:
    if not BRONZE_RAW_ROOT.is_dir():
        raise RuntimeError(
            f"No existe el directorio Bronze RAW esperado: {BRONZE_RAW_ROOT}"
        )

    candidates = [
        p for p in BRONZE_RAW_ROOT.rglob("*")
        if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS
    ]

    if not candidates:
        raise RuntimeError(
            f"No se encontraron archivos XLSX/XLSM en {BRONZE_RAW_ROOT}"
        )

    selected = max(candidates, key=lambda p: p.stat().st_mtime_ns)

    return {
        "selection_mode": "raw_fallback",
        "manifest_path": None,
        "manifest": None,
        "record": {},
        "stored_path": selected,
    }


selection = select_from_manifest() or select_fallback_raw()

BRONZE_XLSX_PATH: Path = selection["stored_path"]
BRONZE_MANIFEST = selection["manifest"]
BRONZE_RECORD = selection["record"]

if not BRONZE_XLSX_PATH.is_file():
    raise RuntimeError(f"El archivo Bronze seleccionado no existe: {BRONZE_XLSX_PATH}")

if BRONZE_XLSX_PATH.stat().st_size <= 0:
    raise RuntimeError(f"El archivo Bronze está vacío: {BRONZE_XLSX_PATH}")

if not zipfile.is_zipfile(BRONZE_XLSX_PATH):
    raise RuntimeError(
        f"El archivo seleccionado no posee estructura ZIP/XLSX válida: {BRONZE_XLSX_PATH}"
    )

actual_sha256, actual_size = sha256_file(BRONZE_XLSX_PATH)
expected_sha256 = str(BRONZE_RECORD.get("sha256") or "").strip().lower()

if expected_sha256 and actual_sha256.lower() != expected_sha256:
    raise RuntimeError(
        "Fallo de lineage Bronze→Silver: "
        f"SHA-256 manifiesto={expected_sha256} != SHA-256 físico={actual_sha256}"
    )

BRONZE_LINEAGE: Dict[str, Any] = {
    "selection_mode": selection["selection_mode"],
    "bronze_manifest_path": selection["manifest_path"],
    "bronze_batch_id": (
        (BRONZE_MANIFEST or {}).get("batch_id")
        or BRONZE_RECORD.get("batch_id")
    ),
    "bronze_status": BRONZE_RECORD.get("status"),
    "original_filename": (
        BRONZE_RECORD.get("original_filename")
        or BRONZE_XLSX_PATH.name
    ),
    "stored_filename": BRONZE_XLSX_PATH.name,
    "stored_path": str(BRONZE_XLSX_PATH),
    "sha256": actual_sha256,
    "size_bytes": actual_size,
    "query_fingerprint": (
        BRONZE_RECORD.get("query_fingerprint")
        or (BRONZE_MANIFEST or {}).get("query_fingerprint")
    ),
    "acquisition_mode": (
        BRONZE_RECORD.get("acquisition_mode")
        or (BRONZE_MANIFEST or {}).get("acquisition_mode")
    ),
    "requested_period_start": BRONZE_RECORD.get("requested_period_start"),
    "requested_period_end": BRONZE_RECORD.get("requested_period_end"),
}

titulo("BRONZE BCE SELECCIONADO Y VERIFICADO")
print(f"Modo selección          : {BRONZE_LINEAGE['selection_mode']}")
print(f"Manifiesto              : {BRONZE_LINEAGE['bronze_manifest_path'] or 'NO DISPONIBLE — fallback RAW'}")
print(f"Bronze batch            : {BRONZE_LINEAGE['bronze_batch_id'] or 'N/D'}")
print(f"Archivo                 : {BRONZE_LINEAGE['original_filename']}")
print(f"Ruta física             : {BRONZE_LINEAGE['stored_path']}")
print(f"Tamaño                  : {BRONZE_LINEAGE['size_bytes']} bytes")
print(f"SHA-256                 : {BRONZE_LINEAGE['sha256']}")

if BRONZE_LINEAGE["selection_mode"] == "raw_fallback":
    print(
        "WARN | Se utilizó fallback RAW porque no se encontró un manifiesto Bronze "
        "compatible que referenciara un archivo existente."
    )


# COMMAND ----------

# =============================================================================
# CELDA 3 — CONTRATO DINÁMICO DEL XLSX Y VALIDACIÓN FÍSICA
# =============================================================================

HEADER_ALIASES: Dict[str, Tuple[str, ...]] = {
    "periodo": ("PERIODO",),
    "codigo_nivel3": ("CODIGO NIVEL 3",),
    "nivel3": ("NIVEL 3",),
    "codigo_pp": ("CODIGO PP",),
    "pp": ("PP",),
    "codigo_subpartida": ("CODIGO SUBPARTIDA",),
    "subpartida": ("SUBPARTIDA",),
    "tm_peso_neto": ("TM (PESO NETO)", "TM PESO NETO"),
    "fob_miles_usd": ("FOB",),
}

EXPECTED_REPORT_TOKENS = (
    "EXPORTACIONES",
    "PRODUCTO PRINCIPAL",
    "SUBPARTIDA",
    "TM",
    "MILES DE USD",
)


def alias_match(value: Any, aliases: Tuple[str, ...]) -> bool:
    normalized = normalize_text_for_match(value)
    return normalized in {normalize_text_for_match(a) for a in aliases}


def detect_contract(workbook) -> Dict[str, Any]:
    matches = []

    for ws in workbook.worksheets:
        max_scan_row = min(max(ws.max_row, 1), 40)

        for row_idx in range(1, max_scan_row + 1):
            header_map: Dict[str, int] = {}

            for canonical, aliases in HEADER_ALIASES.items():
                for col_idx in range(1, ws.max_column + 1):
                    if alias_match(ws.cell(row=row_idx, column=col_idx).value, aliases):
                        header_map[canonical] = col_idx
                        break

            if len(header_map) == len(HEADER_ALIASES):
                matches.append({
                    "sheet": ws.title,
                    "header_row": row_idx,
                    "header_map": header_map,
                })

    if not matches:
        raise RuntimeError(
            "No se encontró una hoja que cumpla el contrato BCE actual. "
            f"Se requieren las columnas: {list(HEADER_ALIASES.keys())}"
        )

    if len(matches) > 1:
        raise RuntimeError(
            "El contrato BCE fue detectado en más de una ubicación; "
            f"se requiere revisión manual: {matches}"
        )

    return matches[0]


WORKBOOK = load_workbook(
    filename=BRONZE_XLSX_PATH,
    data_only=True,
    read_only=False,
)

CONTRACT = detect_contract(WORKBOOK)
SOURCE_SHEET = WORKBOOK[CONTRACT["sheet"]]
HEADER_ROW = int(CONTRACT["header_row"])
HEADER_MAP = dict(CONTRACT["header_map"])

# Detectar última fila con contenido en alguna columna del contrato.
LAST_DATA_ROW = HEADER_ROW
for row_idx in range(HEADER_ROW + 1, SOURCE_SHEET.max_row + 1):
    if any(
        SOURCE_SHEET.cell(row=row_idx, column=col_idx).value is not None
        for col_idx in HEADER_MAP.values()
    ):
        LAST_DATA_ROW = row_idx

if LAST_DATA_ROW <= HEADER_ROW:
    raise RuntimeError("El XLSX contiene encabezados, pero no contiene registros de datos.")

# Metadata textual previa al encabezado.
metadata_lines: List[str] = []
report_generated_at = None

for row_idx in range(1, HEADER_ROW):
    row_values = [
        SOURCE_SHEET.cell(row=row_idx, column=col_idx).value
        for col_idx in range(1, SOURCE_SHEET.max_column + 1)
    ]

    for value in row_values:
        if isinstance(value, datetime) and report_generated_at is None:
            report_generated_at = value

    text_values = [
        re.sub(r"\s+", " ", str(v)).strip()
        for v in row_values
        if v is not None and str(v).strip()
    ]

    if text_values:
        metadata_lines.append(" | ".join(text_values))

METADATA_TEXT = "\n".join(metadata_lines)
NORMALIZED_METADATA_TEXT = normalize_text_for_match(METADATA_TEXT)

missing_report_tokens = [
    token
    for token in EXPECTED_REPORT_TOKENS
    if normalize_text_for_match(token) not in NORMALIZED_METADATA_TEXT
]

if missing_report_tokens:
    raise RuntimeError(
        "El contenido superior del reporte BCE no coincide con el contrato esperado. "
        f"Tokens faltantes: {missing_report_tokens}"
    )

REPORT_TITLE = next(
    (line for line in metadata_lines if "EXPORTACIONES" in normalize_text_for_match(line)),
    "",
)
REPORT_SCOPE = next(
    (line for line in metadata_lines if "PERIODO:" in normalize_text_for_match(line)),
    "",
)

structure_payload = {
    "sheet": SOURCE_SHEET.title,
    "header_row": HEADER_ROW,
    "last_data_row": LAST_DATA_ROW,
    "header_map": HEADER_MAP,
    "report_title": REPORT_TITLE,
    "report_scope": REPORT_SCOPE,
}
WORKBOOK_STRUCTURE_FINGERPRINT = sha256_text(
    json.dumps(structure_payload, ensure_ascii=False, sort_keys=True)
)

WORKBOOK_VALIDATION = {
    **structure_payload,
    "report_generated_at": report_generated_at.isoformat() if isinstance(report_generated_at, datetime) else None,
    "workbook_structure_fingerprint": WORKBOOK_STRUCTURE_FINGERPRINT,
    "status": "PASS",
}

titulo("VALIDACIÓN FÍSICA DEL XLSX BCE")
print(f"Hoja                    : {SOURCE_SHEET.title}")
print(f"Fila encabezado         : {HEADER_ROW}")
print(f"Última fila de datos    : {LAST_DATA_ROW}")
print(f"Filas fuente detectadas : {LAST_DATA_ROW - HEADER_ROW}")
print(f"Título reporte          : {REPORT_TITLE}")
print(f"Ámbito/unidades         : {REPORT_SCOPE}")
print(f"Generado en origen      : {WORKBOOK_VALIDATION['report_generated_at'] or 'N/D'}")
print(f"Fingerprint estructura  : {WORKBOOK_STRUCTURE_FINGERPRINT}")


# COMMAND ----------

# =============================================================================
# CELDA 4 — EXTRACCIÓN CONTROLADA Y MATERIALIZACIÓN INICIAL EN SPARK
# =============================================================================

RAW_COLUMNS = [
    "periodo",
    "codigo_nivel3",
    "nivel3",
    "codigo_pp",
    "pp",
    "codigo_subpartida",
    "subpartida",
    "tm_peso_neto_raw",
    "fob_miles_usd_raw",
]


def identifier_cell_to_string(cell, width: int) -> Optional[str]:
    value = cell.value

    if value is None:
        return None

    if isinstance(value, bool):
        return str(value)

    if isinstance(value, (int, float)):
        number = float(value)
        if number.is_integer():
            return str(int(number)).zfill(width)

    text = str(value).strip()

    # Excel puede devolver un identificador entero como "1402.0".
    if re.fullmatch(r"\d+\.0+", text):
        text = str(int(float(text)))

    if text.isdigit() and len(text) < width:
        text = text.zfill(width)

    return text


def plain_cell_to_string(cell) -> Optional[str]:
    value = cell.value
    if value is None:
        return None
    return str(value)


def measure_cell_to_string(cell) -> Optional[str]:
    value = cell.value
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    return str(value).strip()


raw_rows: List[Dict[str, Any]] = []

for row_idx in range(HEADER_ROW + 1, LAST_DATA_ROW + 1):
    row_payload = {
        "periodo": plain_cell_to_string(
            SOURCE_SHEET.cell(row=row_idx, column=HEADER_MAP["periodo"])
        ),
        "codigo_nivel3": identifier_cell_to_string(
            SOURCE_SHEET.cell(row=row_idx, column=HEADER_MAP["codigo_nivel3"]), 4
        ),
        "nivel3": plain_cell_to_string(
            SOURCE_SHEET.cell(row=row_idx, column=HEADER_MAP["nivel3"])
        ),
        "codigo_pp": identifier_cell_to_string(
            SOURCE_SHEET.cell(row=row_idx, column=HEADER_MAP["codigo_pp"]), 6
        ),
        "pp": plain_cell_to_string(
            SOURCE_SHEET.cell(row=row_idx, column=HEADER_MAP["pp"])
        ),
        "codigo_subpartida": identifier_cell_to_string(
            SOURCE_SHEET.cell(row=row_idx, column=HEADER_MAP["codigo_subpartida"]), 10
        ),
        "subpartida": plain_cell_to_string(
            SOURCE_SHEET.cell(row=row_idx, column=HEADER_MAP["subpartida"])
        ),
        "tm_peso_neto_raw": measure_cell_to_string(
            SOURCE_SHEET.cell(row=row_idx, column=HEADER_MAP["tm_peso_neto"])
        ),
        "fob_miles_usd_raw": measure_cell_to_string(
            SOURCE_SHEET.cell(row=row_idx, column=HEADER_MAP["fob_miles_usd"])
        ),
        "_source_excel_row": int(row_idx),
    }

    # Filas completamente vacías no forman parte del dataset.
    if all(
        row_payload.get(col) is None or str(row_payload.get(col)).strip() == ""
        for col in RAW_COLUMNS
    ):
        continue

    raw_rows.append(row_payload)

if not raw_rows:
    raise RuntimeError("La extracción no produjo registros.")

RAW_SCHEMA = T.StructType([
    T.StructField("periodo", T.StringType(), True),
    T.StructField("codigo_nivel3", T.StringType(), True),
    T.StructField("nivel3", T.StringType(), True),
    T.StructField("codigo_pp", T.StringType(), True),
    T.StructField("pp", T.StringType(), True),
    T.StructField("codigo_subpartida", T.StringType(), True),
    T.StructField("subpartida", T.StringType(), True),
    T.StructField("tm_peso_neto_raw", T.StringType(), True),
    T.StructField("fob_miles_usd_raw", T.StringType(), True),
    T.StructField("_source_excel_row", T.IntegerType(), False),
])

DF_BCE_RAW = spark.createDataFrame(raw_rows, schema=RAW_SCHEMA)

RAW_ROW_COUNT = DF_BCE_RAW.count()

titulo("EXTRACCIÓN BCE → SPARK RAW")
print(f"Registros extraídos     : {RAW_ROW_COUNT}")
print(f"Columnas RAW            : {len(DF_BCE_RAW.columns)}")
DF_BCE_RAW.show(10, truncate=False)


# COMMAND ----------

# =============================================================================
# CELDA 5 — LIMPIEZA TEXTUAL Y DEDUPLICACIÓN EXACTA EN SPARK
# =============================================================================

NULL_TOKENS = ["", "NULL", "NONE", "NAN", "N/A", "NA"]


def clean_text_expr(column_name: str):
    c = F.col(column_name)
    c = F.regexp_replace(c, "[\\u200b\\u200c\\u200d\\ufeff\\u00a0]", "")
    c = F.regexp_replace(c, r"\s+", " ")
    c = F.trim(c)
    return F.when(F.upper(c).isin(NULL_TOKENS), F.lit(None)).otherwise(c)


TEXT_COLUMNS = [
    "periodo",
    "codigo_nivel3",
    "nivel3",
    "codigo_pp",
    "pp",
    "codigo_subpartida",
    "subpartida",
    "tm_peso_neto_raw",
    "fob_miles_usd_raw",
]

df_clean = DF_BCE_RAW

for column_name in TEXT_COLUMNS:
    df_clean = df_clean.withColumn(column_name, clean_text_expr(column_name))

BUSINESS_RAW_COLUMNS = [
    "periodo",
    "codigo_nivel3",
    "nivel3",
    "codigo_pp",
    "pp",
    "codigo_subpartida",
    "subpartida",
    "tm_peso_neto_raw",
    "fob_miles_usd_raw",
]

rows_before = df_clean.count()
rows_after_dedup = df_clean.dropDuplicates(BUSINESS_RAW_COLUMNS).count()
exact_duplicates_removed = rows_before - rows_after_dedup

DF_BCE_CLEAN = df_clean.dropDuplicates(BUSINESS_RAW_COLUMNS)

CLEANING_AUDIT = {
    "rows_before": rows_before,
    "rows_after": DF_BCE_CLEAN.count(),
    "exact_duplicates_removed": exact_duplicates_removed,
    "text_columns_cleaned": TEXT_COLUMNS,
    "semantic_relabeling_applied": False,
}

titulo("LIMPIEZA SILVER BCE")
print(f"Filas antes             : {CLEANING_AUDIT['rows_before']}")
print(f"Filas después           : {CLEANING_AUDIT['rows_after']}")
print(f"Duplicados exactos      : {CLEANING_AUDIT['exact_duplicates_removed']}")
print("Relabeling semántico    : NO — las descripciones oficiales se preservan")


# COMMAND ----------

# =============================================================================
# CELDA 6 — TIPADO CONTROLADO, PERÍODO Y NORMALIZACIÓN DE UNIDADES
# =============================================================================

def invalid_numeric_cast_count(df: DataFrame, raw_col: str) -> int:
    return (
        df.where(
            F.col(raw_col).isNotNull()
            & F.col(raw_col).cast("double").isNull()
        )
        .count()
    )


invalid_tm = invalid_numeric_cast_count(DF_BCE_CLEAN, "tm_peso_neto_raw")
invalid_fob = invalid_numeric_cast_count(DF_BCE_CLEAN, "fob_miles_usd_raw")

period_year = F.regexp_extract(F.col("periodo"), r"^(\d{4})\s*/\s*(\d{2})\s*-\s*(.+)$", 1)
period_month = F.regexp_extract(F.col("periodo"), r"^(\d{4})\s*/\s*(\d{2})\s*-\s*(.+)$", 2)

df_typed = (
    DF_BCE_CLEAN
    .withColumn("anio", period_year.cast("int"))
    .withColumn("mes_numero", period_month.cast("int"))
    .withColumn(
        "fecha_periodo",
        F.to_date(
            F.concat_ws(
                "-",
                F.col("anio").cast("string"),
                F.lpad(F.col("mes_numero").cast("string"), 2, "0"),
                F.lit("01"),
            ),
            "yyyy-MM-dd",
        ),
    )
    .withColumn(
        "toneladas_metricas_peso_neto",
        F.col("tm_peso_neto_raw").cast("double"),
    )
    .withColumn(
        "fob_miles_usd",
        F.col("fob_miles_usd_raw").cast("double"),
    )
    .withColumn(
        "fob_usd",
        F.col("fob_miles_usd") * F.lit(1000.0),
    )
    .withColumnRenamed("codigo_pp", "codigo_producto_principal")
    .withColumnRenamed("pp", "producto_principal")
)

invalid_periods = (
    df_typed
    .where(
        F.col("fecha_periodo").isNull()
        | F.col("anio").isNull()
        | F.col("mes_numero").isNull()
        | (~F.col("mes_numero").between(1, 12))
    )
    .count()
)

TYPING_AUDIT = {
    "invalid_tm_casts": invalid_tm,
    "invalid_fob_casts": invalid_fob,
    "invalid_periods": invalid_periods,
    "fob_source_unit": "thousand_USD",
    "fob_canonical_unit": "USD",
    "fob_conversion_factor": 1000.0,
}

if any([
    TYPING_AUDIT["invalid_tm_casts"] > 0,
    TYPING_AUDIT["invalid_fob_casts"] > 0,
    TYPING_AUDIT["invalid_periods"] > 0,
]):
    raise RuntimeError(
        "La conversión controlada detectó valores inválidos: "
        f"{TYPING_AUDIT}"
    )

FINAL_COLUMNS = [
    "periodo",
    "fecha_periodo",
    "anio",
    "mes_numero",
    "codigo_nivel3",
    "nivel3",
    "codigo_producto_principal",
    "producto_principal",
    "codigo_subpartida",
    "subpartida",
    "toneladas_metricas_peso_neto",
    "fob_miles_usd",
    "fob_usd",
]

DF_BCE_TYPED = df_typed.select(*FINAL_COLUMNS)

if DF_BCE_TYPED.count() <= 0:
    raise RuntimeError("El conjunto BCE quedó vacío después del tipado controlado.")

titulo("TIPADO CONTROLADO BCE")
print(f"Conversiones TM inválidas  : {invalid_tm}")
print(f"Conversiones FOB inválidas : {invalid_fob}")
print(f"Períodos inválidos         : {invalid_periods}")
print("Cobertura temporal         : se conserva completa según el recurso BCE seleccionado")
DF_BCE_TYPED.printSchema()


# COMMAND ----------

# =============================================================================
# CELDA 7 — CONSISTENCIA SEMÁNTICA Y JERARQUÍA DE PRODUCTO
# =============================================================================

SEMANTIC_RESULTS: List[Dict[str, Any]] = []


def semantic_record(check: str, severity: str, violations: int, detail: str) -> None:
    SEMANTIC_RESULTS.append({
        "check": check,
        "severity": severity,
        "status": "PASS" if violations == 0 else ("FAIL" if severity == "CRITICAL" else "WARN"),
        "violations": int(violations),
        "detail": detail,
    })


wrong_level3_code = DF_BCE_TYPED.where(F.col("codigo_nivel3") != F.lit("1402")).count()
wrong_pp_code = DF_BCE_TYPED.where(F.col("codigo_producto_principal") != F.lit("140201")).count()

wrong_level3_label = (
    DF_BCE_TYPED
    .where(~F.upper(F.col("nivel3")).contains("CAMARON"))
    .count()
)
wrong_pp_label = (
    DF_BCE_TYPED
    .where(~F.upper(F.col("producto_principal")).contains("CAMARON"))
    .count()
)

description_variations = (
    DF_BCE_TYPED
    .groupBy("codigo_subpartida")
    .agg(F.countDistinct("subpartida").alias("descripciones"))
    .where(F.col("descripciones") > 1)
    .count()
)

# Anomalías textuales observables en el origen se registran, no se corrigen
# silenciosamente. El código arancelario es la clave de agrupación.
source_text_anomalies = (
    DF_BCE_TYPED
    .where(
        F.upper(F.col("subpartida")).contains("FRESCVOS")
        | (
            F.upper(F.col("subpartida")).contains("DEMAS")
            & (~F.upper(F.col("subpartida")).contains("DEMÁS"))
        )
    )
    .select("codigo_subpartida", "subpartida")
    .distinct()
)

source_text_anomaly_count = source_text_anomalies.count()

semantic_record(
    "codigo_nivel3_1402",
    "CRITICAL",
    wrong_level3_code,
    "Todos los registros del reporte deben pertenecer al nivel 3 código 1402.",
)
semantic_record(
    "codigo_pp_140201",
    "CRITICAL",
    wrong_pp_code,
    "Todos los registros deben corresponder al producto principal 140201.",
)
semantic_record(
    "nivel3_etiqueta_camaron",
    "CRITICAL",
    wrong_level3_label,
    "La etiqueta Nivel 3 debe identificar camarones.",
)
semantic_record(
    "pp_etiqueta_camaron",
    "CRITICAL",
    wrong_pp_label,
    "La etiqueta PP debe identificar camarones.",
)
semantic_record(
    "descripcion_unica_por_subpartida",
    "WARNING",
    description_variations,
    "Se informa si un mismo código posee más de una descripción oficial en el snapshot.",
)
semantic_record(
    "anomalias_textuales_origen",
    "WARNING",
    source_text_anomaly_count,
    "Se preservan textos oficiales incluso cuando contienen variantes ortográficas/acentuales.",
)

critical_semantic_failures = [
    r for r in SEMANTIC_RESULTS
    if r["severity"] == "CRITICAL" and r["violations"] > 0
]

titulo("CONSISTENCIA SEMÁNTICA BCE")
for item in SEMANTIC_RESULTS:
    print(
        f"{item['status']:5s} | {item['check']:38s} | "
        f"violaciones={item['violations']}"
    )

if source_text_anomaly_count > 0:
    print("\nWARN | Anomalías textuales preservadas desde la fuente:")
    source_text_anomalies.orderBy("codigo_subpartida").show(100, truncate=False)

if critical_semantic_failures:
    raise RuntimeError(
        f"Fallaron controles semánticos críticos: {critical_semantic_failures}"
    )

print("\nCatálogo de subpartidas detectado:")
(
    DF_BCE_TYPED
    .groupBy("codigo_subpartida", "subpartida")
    .agg(
        F.min("fecha_periodo").alias("desde"),
        F.max("fecha_periodo").alias("hasta"),
        F.count("*").alias("registros"),
    )
    .orderBy("codigo_subpartida")
    .show(100, truncate=False)
)


# COMMAND ----------

# =============================================================================
# CELDA 8 — LLAVE DE NEGOCIO Y CONTINUIDAD TEMPORAL
# =============================================================================

BUSINESS_KEY = [
    "fecha_periodo",
    "codigo_producto_principal",
    "codigo_subpartida",
]

duplicate_key_groups = (
    DF_BCE_TYPED
    .groupBy(*BUSINESS_KEY)
    .count()
    .where(F.col("count") > 1)
)

duplicate_key_group_count = duplicate_key_groups.count()

period_stats = (
    DF_BCE_TYPED
    .agg(
        F.min("fecha_periodo").alias("min_period"),
        F.max("fecha_periodo").alias("max_period"),
        F.countDistinct("fecha_periodo").alias("distinct_periods"),
    )
    .first()
)

MIN_PERIOD = period_stats["min_period"]
MAX_PERIOD = period_stats["max_period"]
DISTINCT_PERIODS = int(period_stats["distinct_periods"])

if MIN_PERIOD is None or MAX_PERIOD is None:
    raise RuntimeError("No fue posible determinar el rango temporal Silver.")

observed_periods = {
    row["fecha_periodo"].strftime("%Y-%m")
    for row in DF_BCE_TYPED.select("fecha_periodo").distinct().collect()
}

expected_periods = []
year = MIN_PERIOD.year
month = MIN_PERIOD.month

while (year, month) <= (MAX_PERIOD.year, MAX_PERIOD.month):
    expected_periods.append(f"{year:04d}-{month:02d}")
    if month == 12:
        year += 1
        month = 1
    else:
        month += 1

missing_periods = sorted(set(expected_periods) - observed_periods)

PERIOD_AUDIT = {
    "min_period": MIN_PERIOD.isoformat(),
    "max_period": MAX_PERIOD.isoformat(),
    "distinct_periods": DISTINCT_PERIODS,
    "expected_periods_between_bounds": len(expected_periods),
    "missing_periods_between_bounds": missing_periods,
}

if duplicate_key_group_count > 0:
    print("ERROR | Duplicados en llave de negocio:")
    duplicate_key_groups.show(100, truncate=False)
    raise RuntimeError(
        f"Se detectaron {duplicate_key_group_count} grupos duplicados en la llave {BUSINESS_KEY}."
    )

if missing_periods:
    raise RuntimeError(
        "La serie temporal global tiene meses faltantes entre el primer y último período: "
        f"{missing_periods}"
    )

DF_SILVER_FINAL = DF_BCE_TYPED.orderBy(
    "fecha_periodo",
    "codigo_subpartida",
)

titulo("LLAVE Y COBERTURA TEMPORAL")
print(f"Llave de negocio        : {BUSINESS_KEY}")
print(f"Duplicados de llave     : {duplicate_key_group_count}")
print(f"Primer período          : {MIN_PERIOD}")
print(f"Último período          : {MAX_PERIOD}")
print(f"Meses distintos         : {DISTINCT_PERIODS}")
print(f"Meses faltantes         : {missing_periods}")


# COMMAND ----------

# =============================================================================
# CELDA 9 — REGLAS DE CALIDAD Y NEGOCIO EJECUTABLES
# =============================================================================

QUALITY_RESULTS: List[Dict[str, Any]] = []


def quality_record(
    rule: str,
    severity: str,
    violations: int,
    checked: int,
    detail: str,
) -> None:
    status = "PASS" if violations == 0 else ("FAIL" if severity == "CRITICAL" else "WARN")
    QUALITY_RESULTS.append({
        "rule": rule,
        "severity": severity,
        "status": status,
        "violations": int(violations),
        "checked": int(checked),
        "detail": detail,
    })


ROW_COUNT = DF_SILVER_FINAL.count()

critical_columns = [
    "periodo",
    "fecha_periodo",
    "anio",
    "mes_numero",
    "codigo_nivel3",
    "nivel3",
    "codigo_producto_principal",
    "producto_principal",
    "codigo_subpartida",
    "subpartida",
    "toneladas_metricas_peso_neto",
    "fob_miles_usd",
    "fob_usd",
]

null_condition = None
for c in critical_columns:
    condition = F.col(c).isNull()
    null_condition = condition if null_condition is None else (null_condition | condition)

critical_null_rows = DF_SILVER_FINAL.where(null_condition).count()

negative_tm = DF_SILVER_FINAL.where(F.col("toneladas_metricas_peso_neto") < 0).count()
negative_fob = DF_SILVER_FINAL.where(F.col("fob_miles_usd") < 0).count()
zero_tm = DF_SILVER_FINAL.where(F.col("toneladas_metricas_peso_neto") == 0).count()
zero_fob = DF_SILVER_FINAL.where(F.col("fob_miles_usd") == 0).count()

invalid_level3_format = (
    DF_SILVER_FINAL
    .where(~F.col("codigo_nivel3").rlike(r"^\d{4}$"))
    .count()
)
invalid_pp_format = (
    DF_SILVER_FINAL
    .where(~F.col("codigo_producto_principal").rlike(r"^\d{6}$"))
    .count()
)
invalid_subpartida_format = (
    DF_SILVER_FINAL
    .where(~F.col("codigo_subpartida").rlike(r"^\d{10}$"))
    .count()
)

invalid_usd_conversion = (
    DF_SILVER_FINAL
    .where(
        F.abs(
            F.col("fob_usd")
            - (F.col("fob_miles_usd") * F.lit(1000.0))
        ) > F.lit(ABS_TOL_USD_CONVERSION)
    )
    .count()
)

quality_record(
    "dataset_no_vacio",
    "CRITICAL",
    0 if ROW_COUNT > 0 else 1,
    1,
    "La tabla Silver debe contener al menos un registro.",
)
quality_record(
    "sin_nulos_criticos",
    "CRITICAL",
    critical_null_rows,
    ROW_COUNT,
    "No se admiten nulos en campos de negocio y medidas del contrato actual.",
)
quality_record(
    "tm_no_negativa",
    "CRITICAL",
    negative_tm,
    ROW_COUNT,
    "TM (Peso Neto) no puede ser negativa.",
)
quality_record(
    "fob_no_negativo",
    "CRITICAL",
    negative_fob,
    ROW_COUNT,
    "FOB no puede ser negativo.",
)
quality_record(
    "tm_cero",
    "WARNING",
    zero_tm,
    ROW_COUNT,
    "Los valores cero se informan, pero no se eliminan automáticamente.",
)
quality_record(
    "fob_cero",
    "WARNING",
    zero_fob,
    ROW_COUNT,
    "Los valores cero se informan, pero no se eliminan automáticamente.",
)
quality_record(
    "formato_codigo_nivel3",
    "CRITICAL",
    invalid_level3_format,
    ROW_COUNT,
    "Código Nivel 3 debe contener exactamente cuatro dígitos.",
)
quality_record(
    "formato_codigo_pp",
    "CRITICAL",
    invalid_pp_format,
    ROW_COUNT,
    "Código PP debe contener exactamente seis dígitos.",
)
quality_record(
    "formato_codigo_subpartida",
    "CRITICAL",
    invalid_subpartida_format,
    ROW_COUNT,
    "Código Subpartida debe contener exactamente diez dígitos.",
)
quality_record(
    "conversion_fob_miles_a_usd",
    "CRITICAL",
    invalid_usd_conversion,
    ROW_COUNT,
    "FOB USD debe ser exactamente FOB (miles USD) × 1000 dentro de tolerancia numérica.",
)
quality_record(
    "llave_negocio_unica",
    "CRITICAL",
    duplicate_key_group_count,
    ROW_COUNT,
    f"Llave evaluada: {BUSINESS_KEY}",
)
quality_record(
    "continuidad_temporal_global",
    "CRITICAL",
    len(missing_periods),
    len(expected_periods),
    "No deben existir huecos entre el primer y último mes observado en el snapshot.",
)

critical_quality_failures = [
    r for r in QUALITY_RESULTS
    if r["severity"] == "CRITICAL" and r["violations"] > 0
]

titulo("REGLAS DE CALIDAD SILVER BCE")
for item in QUALITY_RESULTS:
    print(
        f"{item['status']:5s} | {item['rule']:36s} | "
        f"violaciones={item['violations']:4d} | revisados={item['checked']}"
    )

if critical_quality_failures:
    raise RuntimeError(
        f"Silver BCE no supera las reglas críticas: {critical_quality_failures}"
    )


# COMMAND ----------

# =============================================================================
# CELDA 10 — MÉTRICAS DE CALIDAD, COBERTURA Y TRAZABILIDAD
# =============================================================================

aggregate_row = (
    DF_SILVER_FINAL
    .agg(
        F.count("*").alias("rows"),
        F.countDistinct("fecha_periodo").alias("distinct_periods"),
        F.countDistinct("codigo_subpartida").alias("distinct_subpartidas"),
        F.countDistinct("codigo_producto_principal").alias("distinct_pp"),
        F.sum("toneladas_metricas_peso_neto").alias("total_tm"),
        F.sum("fob_miles_usd").alias("total_fob_miles_usd"),
        F.sum("fob_usd").alias("total_fob_usd"),
        F.min("fecha_periodo").alias("min_period"),
        F.max("fecha_periodo").alias("max_period"),
    )
    .first()
)

null_exprs = [
    F.sum(F.when(F.col(c).isNull(), 1).otherwise(0)).alias(c)
    for c in DF_SILVER_FINAL.columns
]
null_row = DF_SILVER_FINAL.agg(*null_exprs).first().asDict()
TOTAL_NULLS = int(sum(int(v or 0) for v in null_row.values()))

SILVER_METRICS = {
    "rows": int(aggregate_row["rows"]),
    "columns": len(DF_SILVER_FINAL.columns),
    "nulls_total": TOTAL_NULLS,
    "distinct_periods": int(aggregate_row["distinct_periods"]),
    "distinct_subpartidas": int(aggregate_row["distinct_subpartidas"]),
    "distinct_productos_principales": int(aggregate_row["distinct_pp"]),
    "total_tm": float(aggregate_row["total_tm"] or 0.0),
    "total_fob_miles_usd": float(aggregate_row["total_fob_miles_usd"] or 0.0),
    "total_fob_usd": float(aggregate_row["total_fob_usd"] or 0.0),
    "min_period": aggregate_row["min_period"].isoformat(),
    "max_period": aggregate_row["max_period"].isoformat(),
    "business_key_duplicate_groups": int(duplicate_key_group_count),
}

YEAR_METRICS = [
    row.asDict()
    for row in (
        DF_SILVER_FINAL
        .groupBy("anio")
        .agg(
            F.count("*").alias("registros"),
            F.countDistinct("fecha_periodo").alias("meses"),
            F.sum("toneladas_metricas_peso_neto").alias("total_tm"),
            F.sum("fob_miles_usd").alias("total_fob_miles_usd"),
        )
        .orderBy("anio")
        .collect()
    )
]

titulo("MÉTRICAS SILVER BCE ANTES DE PERSISTENCIA")
for key, value in SILVER_METRICS.items():
    print(f"{key:32s}: {value}")

print("\nResumen por año:")
for item in YEAR_METRICS:
    print(item)


# COMMAND ----------

# =============================================================================
# CELDA 11 — PERSISTENCIA DELTA EN UNITY CATALOG
# =============================================================================

spark.sql(f"CREATE SCHEMA IF NOT EXISTS `{CATALOG}`.`{SILVER_SCHEMA}`")

LEGACY_TABLE_REF = f"{CATALOG}.{SILVER_SCHEMA}.exportaciones_camaron"
if spark.catalog.tableExists(LEGACY_TABLE_REF):
    print(
        "WARN | Existe la tabla heredada "
        f"{LEGACY_TABLE_REF}. No se elimina automáticamente; "
        "no debe utilizarse como sustituto de la nueva tabla BCE hasta revisar Gold."
    )

TABLE_REF_PLAIN = f"{CATALOG}.{SILVER_SCHEMA}.{OUTPUT_TABLE}"
TABLE_REF_SQL = f"`{CATALOG}`.`{SILVER_SCHEMA}`.`{OUTPUT_TABLE}`"

titulo("PERSISTENCIA SILVER BCE EN UNITY CATALOG")

(
    DF_SILVER_FINAL
    .write
    .format("delta")
    .mode("overwrite")
    .option("overwriteSchema", "true")
    .saveAsTable(TABLE_REF_PLAIN)
)

props = {
    "pipeline.layer": "silver",
    "pipeline.source": SOURCE_CODE,
    "pipeline.version": PIPELINE_VERSION,
    "pipeline.silver_run_id": RUN_ID,
    "lineage.selection_mode": str(BRONZE_LINEAGE.get("selection_mode") or ""),
    "lineage.bronze_batch_id": str(BRONZE_LINEAGE.get("bronze_batch_id") or ""),
    "lineage.bronze_sha256": BRONZE_LINEAGE["sha256"],
    "lineage.bronze_original_filename": str(BRONZE_LINEAGE.get("original_filename") or ""),
    "lineage.workbook_structure_fingerprint": WORKBOOK_STRUCTURE_FINGERPRINT,
    "source.report": REPORT_TITLE,
    "source.period_min": SILVER_METRICS["min_period"],
    "source.period_max": SILVER_METRICS["max_period"],
    "source.fob_unit": "thousand_USD",
    "canonical.fob_unit": "USD",
}

props_sql = ", ".join(
    f"'{k}'='{str(v).replace(chr(39), chr(39) * 2)}'"
    for k, v in props.items()
)

spark.sql(
    f"ALTER TABLE {TABLE_REF_SQL} SET TBLPROPERTIES ({props_sql})"
)

persisted_rows = spark.table(TABLE_REF_PLAIN).count()

PERSISTENCE_RESULT = {
    "table": OUTPUT_TABLE,
    "table_ref": TABLE_REF_PLAIN,
    "rows": persisted_rows,
    "status": "WRITTEN",
}

print(f"PASS | {TABLE_REF_PLAIN}")
print(f"Filas persistidas       : {persisted_rows}")


# COMMAND ----------

# =============================================================================
# CELDA 12 — VALIDACIÓN POST-ESCRITURA: ESQUEMA, FILAS Y EQUIVALENCIA
# =============================================================================

PERSISTED_DF = spark.table(TABLE_REF_PLAIN)

source_count = DF_SILVER_FINAL.count()
persisted_count = PERSISTED_DF.count()

same_columns = DF_SILVER_FINAL.columns == PERSISTED_DF.columns

source_types = [
    field.dataType.simpleString()
    for field in DF_SILVER_FINAL.schema.fields
]
persisted_types = [
    field.dataType.simpleString()
    for field in PERSISTED_DF.schema.fields
]
same_schema = source_types == persisted_types

missing_after_write = DF_SILVER_FINAL.exceptAll(PERSISTED_DF).count()
extra_after_write = PERSISTED_DF.exceptAll(DF_SILVER_FINAL).count()

POSTWRITE_RESULT = {
    "source_rows": source_count,
    "persisted_rows": persisted_count,
    "same_columns": same_columns,
    "same_schema": same_schema,
    "missing_after_write": missing_after_write,
    "extra_after_write": extra_after_write,
    "status": "PASS",
}

if source_count != persisted_count:
    POSTWRITE_RESULT["status"] = "FAIL"

if not same_columns or not same_schema:
    POSTWRITE_RESULT["status"] = "FAIL"

if missing_after_write != 0 or extra_after_write != 0:
    POSTWRITE_RESULT["status"] = "FAIL"

titulo("VALIDACIÓN POST-ESCRITURA")
for key, value in POSTWRITE_RESULT.items():
    print(f"{key:28s}: {value}")

if POSTWRITE_RESULT["status"] != "PASS":
    raise RuntimeError(
        f"La validación post-escritura falló: {POSTWRITE_RESULT}"
    )


# COMMAND ----------

# =============================================================================
# CELDA 13 — MANTENIMIENTO DELTA CON CRITERIO OBJETIVO
# =============================================================================

detail = spark.sql(f"DESCRIBE DETAIL {TABLE_REF_SQL}").first().asDict()

num_files = int(detail.get("numFiles") or 0)
size_bytes = int(detail.get("sizeInBytes") or 0)
size_mb = size_bytes / (1024 * 1024)

MAINTENANCE_RESULT = {
    "table": OUTPUT_TABLE,
    "num_files_before": num_files,
    "size_bytes_before": size_bytes,
    "size_mb_before": round(size_mb, 6),
    "optimize_min_files": OPTIMIZE_MIN_FILES,
    "optimize_min_size_mb": OPTIMIZE_MIN_SIZE_MB,
    "status": None,
    "zorder_column": None,
}

should_optimize = (
    ENABLE_DELTA_MAINTENANCE
    and (
        num_files >= OPTIMIZE_MIN_FILES
        or size_mb >= OPTIMIZE_MIN_SIZE_MB
    )
)

titulo("MANTENIMIENTO DELTA BCE")

if not should_optimize:
    MAINTENANCE_RESULT["status"] = "SKIPPED_SMALL_TABLE"
    print(
        "SKIP | La tabla no alcanza los umbrales de tamaño/fragmentación; "
        "OPTIMIZE no aportaría evidencia objetiva de mejora."
    )
else:
    zorder_column = "fecha_periodo"

    if zorder_column not in PERSISTED_DF.columns:
        raise RuntimeError(
            f"La columna ZORDER configurada no existe: {zorder_column}"
        )

    spark.sql(
        f"OPTIMIZE {TABLE_REF_SQL} ZORDER BY (`{zorder_column}`)"
    )

    MAINTENANCE_RESULT["status"] = "OPTIMIZED"
    MAINTENANCE_RESULT["zorder_column"] = zorder_column
    print(f"PASS | OPTIMIZE + ZORDER BY {zorder_column}")

print(MAINTENANCE_RESULT)


# COMMAND ----------

# =============================================================================
# CELDA 14 — SUITE TÉCNICA DE QA SOBRE LA TABLA PERSISTIDA
# =============================================================================

QA_RESULTS: List[Dict[str, Any]] = []


def qa_record(test: str, passed: bool, detail: str) -> None:
    QA_RESULTS.append({
        "table": OUTPUT_TABLE,
        "test": test,
        "status": "PASS" if passed else "FAIL",
        "detail": detail,
    })


titulo("QA TÉCNICO SILVER BCE")

try:
    qa_df = spark.table(TABLE_REF_PLAIN)
    qa_record("accesibilidad", True, "spark.table() ejecutó correctamente.")
except Exception as exc:
    qa_record("accesibilidad", False, str(exc))
    qa_df = None

if qa_df is not None:
    qa_rows = qa_df.count()
    qa_record(
        "row_count",
        qa_rows == SILVER_METRICS["rows"],
        f"esperado={SILVER_METRICS['rows']} observado={qa_rows}",
    )

    qa_record(
        "columnas",
        qa_df.columns == DF_SILVER_FINAL.columns,
        f"esperadas={DF_SILVER_FINAL.columns} observadas={qa_df.columns}",
    )

    qa_dup_groups = (
        qa_df
        .groupBy(*BUSINESS_KEY)
        .count()
        .where(F.col("count") > 1)
        .count()
    )
    qa_record(
        "llave_negocio_unica",
        qa_dup_groups == 0,
        f"grupos_duplicados={qa_dup_groups}",
    )

    qa_critical_nulls = qa_df.where(null_condition).count()
    qa_record(
        "nulos_criticos",
        qa_critical_nulls == 0,
        f"filas_con_nulos_criticos={qa_critical_nulls}",
    )

    props_df = spark.sql(f"SHOW TBLPROPERTIES {TABLE_REF_SQL}")
    props_map = {
        row["key"]: row["value"]
        for row in props_df.collect()
    }

    lineage_ok = (
        props_map.get("pipeline.layer") == "silver"
        and props_map.get("pipeline.source") == SOURCE_CODE
        and props_map.get("lineage.bronze_sha256") == BRONZE_LINEAGE["sha256"]
    )
    qa_record(
        "lineage_properties",
        lineage_ok,
        "Se verificaron pipeline.layer, pipeline.source y lineage.bronze_sha256.",
    )

QA_FAILURES = [item for item in QA_RESULTS if item["status"] == "FAIL"]

for item in QA_RESULTS:
    print(
        f"{item['status']:4s} | {item['test']:26s} | {item['detail']}"
    )

print(f"\nTests ejecutados         : {len(QA_RESULTS)}")
print(f"Tests exitosos          : {len(QA_RESULTS) - len(QA_FAILURES)}")
print(f"Tests fallidos          : {len(QA_FAILURES)}")

if QA_FAILURES:
    raise RuntimeError(f"QA técnico Silver BCE falló: {QA_FAILURES}")


# COMMAND ----------

# =============================================================================
# CELDA 15 — MANIFIESTO SILVER PERSISTENTE Y DURACIÓN REAL
# =============================================================================

PIPELINE_FINISHED_AT_UTC = datetime.now(timezone.utc)
PIPELINE_DURATION_SECONDS = round(
    time.perf_counter() - PIPELINE_STARTED_MONOTONIC,
    3,
)

spark.sql(
    f"CREATE VOLUME IF NOT EXISTS "
    f"`{CATALOG}`.`{SILVER_SCHEMA}`.`{SILVER_METADATA_VOLUME}`"
)

metadata_root = Path(
    f"/Volumes/{CATALOG}/{SILVER_SCHEMA}/{SILVER_METADATA_VOLUME}"
)
metadata_root.mkdir(parents=True, exist_ok=True)

manifest = {
    "manifest_schema_version": SILVER_MANIFEST_SCHEMA_VERSION,
    "pipeline_version": PIPELINE_VERSION,
    "run_id": RUN_ID,
    "source": SOURCE_CODE,
    "layer": "silver",
    "run_started_at_utc": PIPELINE_STARTED_AT_UTC.isoformat(),
    "run_finished_at_utc": PIPELINE_FINISHED_AT_UTC.isoformat(),
    "duration_seconds": PIPELINE_DURATION_SECONDS,
    "bronze_lineage": BRONZE_LINEAGE,
    "workbook_validation": WORKBOOK_VALIDATION,
    "cleaning_audit": CLEANING_AUDIT,
    "typing_audit": TYPING_AUDIT,
    "semantic_results": SEMANTIC_RESULTS,
    "period_audit": PERIOD_AUDIT,
    "quality_results": QUALITY_RESULTS,
    "metrics": SILVER_METRICS,
    "year_metrics": YEAR_METRICS,
    "persistence": PERSISTENCE_RESULT,
    "postwrite_validation": POSTWRITE_RESULT,
    "delta_maintenance": MAINTENANCE_RESULT,
    "qa_results": QA_RESULTS,
    "output": {
        "catalog": CATALOG,
        "schema": SILVER_SCHEMA,
        "table": OUTPUT_TABLE,
        "table_ref": TABLE_REF_PLAIN,
        "business_key": BUSINESS_KEY,
        "columns": DF_SILVER_FINAL.columns,
    },
    "software_versions": software_versions(),
}

manifest_name = f"silver_manifest_{RUN_ID}_BCE.json"
SILVER_MANIFEST_PATH = metadata_root / manifest_name

atomic_json_dump(manifest, SILVER_MANIFEST_PATH)

manifest_sha, manifest_size = sha256_file(SILVER_MANIFEST_PATH)

# Verificación de lectura del propio manifiesto.
manifest_reloaded = load_json(SILVER_MANIFEST_PATH)

if manifest_reloaded.get("run_id") != RUN_ID:
    raise RuntimeError("El manifiesto Silver persistido no superó la verificación de lectura.")

titulo("MANIFIESTO SILVER BCE")
print(f"Ruta                     : {SILVER_MANIFEST_PATH}")
print(f"Tamaño                   : {manifest_size} bytes")
print(f"SHA-256                  : {manifest_sha}")
print(f"Duración real            : {PIPELINE_DURATION_SECONDS} segundos")


# COMMAND ----------

# =============================================================================
# CELDA 16 — REPORTE FINAL DEL PIPELINE SILVER BCE
# =============================================================================

titulo("REPORTE FINAL — SILVER BCE")

print("Estado                  : SUCCESS")
print(f"Run ID                  : {RUN_ID}")
print(f"Pipeline version        : {PIPELINE_VERSION}")
print(f"Selección Bronze        : {BRONZE_LINEAGE['selection_mode']}")
print(f"Bronze batch            : {BRONZE_LINEAGE['bronze_batch_id'] or 'N/D'}")
print(f"Archivo Bronze          : {BRONZE_LINEAGE['original_filename']}")
print(f"SHA-256 fuente          : {BRONZE_LINEAGE['sha256']}")
print(f"Reporte BCE             : {REPORT_TITLE}")
print(f"Período mínimo          : {SILVER_METRICS['min_period']}")
print(f"Período máximo          : {SILVER_METRICS['max_period']}")
print(f"Meses distintos         : {SILVER_METRICS['distinct_periods']}")
print(f"Subpartidas             : {SILVER_METRICS['distinct_subpartidas']}")
print(f"Registros Silver        : {SILVER_METRICS['rows']}")
print(f"Total TM                : {SILVER_METRICS['total_tm']:.6f}")
print(f"FOB miles USD           : {SILVER_METRICS['total_fob_miles_usd']:.6f}")
print(f"FOB USD normalizado     : {SILVER_METRICS['total_fob_usd']:.2f}")
print(f"Tabla Delta             : {TABLE_REF_PLAIN}")
print(f"QA                      : {len(QA_RESULTS) - len(QA_FAILURES)}/{len(QA_RESULTS)} PASS")
print(f"Mantenimiento Delta     : {MAINTENANCE_RESULT['status']}")
print(f"Duración medida         : {PIPELINE_DURATION_SECONDS:.3f} segundos")
print(f"Manifiesto Silver       : {SILVER_MANIFEST_PATH}")

print("\nIMPORTANTE:")
print(
    "Este snapshot BCE no contiene 'País Destino'. "
    "La granularidad Silver es período × producto principal × subpartida."
)


# COMMAND ----------
