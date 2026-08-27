# Databricks notebook source

# =============================================================================
# CELDA 1 — IMPORTACIONES, CONFIGURACIÓN Y RELOJ DE EJECUCIÓN
# =============================================================================

import json
import re
import time
import hashlib
import unicodedata
import importlib.metadata as importlib_metadata
from dataclasses import dataclass, asdict
from pathlib import Path
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
%pip install openpyxl
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from pyspark.sql import DataFrame
from pyspark.sql import functions as F
from pyspark.sql import types as T
from pyspark.sql.window import Window

PIPELINE_VERSION = "2.1.0"
SILVER_MANIFEST_SCHEMA_VERSION = "1.0"
EXPECTED_BRONZE_STORAGE_LAYOUT_VERSION = "2.0"

CATALOG = "camaronera_2026"
BRONZE_SCHEMA = "bronce"
SILVER_SCHEMA = "plata"
BRONZE_VOLUME = "datoscna"
SILVER_METADATA_VOLUME = "metadatos_cna"
SOURCE_CODE = "CNA"

# Corte reproducible del estudio. El contrato tabular actual fue validado para mayo de 2026.
# Si en el futuro se procesa otro corte, primero deben actualizarse y revalidarse los rangos/contratos de la Celda 3.
TARGET_DOCUMENT_YEAR: Optional[int] = 2026
TARGET_DOCUMENT_MONTH: Optional[int] = 5

SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
SUCCESS_BRONZE_STATUSES = {"downloaded", "updated_version", "unchanged"}

# Tolerancias de validación matemática.
REL_TOL = 1e-6
ABS_TOL_TOTAL = 1.0
ABS_TOL_PRICE = 0.02
REL_TOL_PRICE = 0.01
ABS_TOL_VARIATION = 0.02
REL_TOL_VARIATION = 0.01

# Mantenimiento Delta: solo se ejecuta cuando la tabla lo justifica por tamaño/fragmentación.
ENABLE_DELTA_MAINTENANCE = True
OPTIMIZE_MIN_FILES = 10
OPTIMIZE_MIN_SIZE_MB = 128.0

RUN_ID = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
PIPELINE_STARTED_AT_UTC = datetime.now(timezone.utc)
PIPELINE_STARTED_MONOTONIC = time.perf_counter()


def titulo(texto: str) -> None:
    print("\n" + "=" * 100)
    print(texto)
    print("=" * 100)


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> Tuple[str, int]:
    digest = hashlib.sha256()
    size = 0
    with path.open("rb") as fh:
        while True:
            chunk = fh.read(chunk_size)
            if not chunk:
                break
            digest.update(chunk)
            size += len(chunk)
    return digest.hexdigest(), size


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def atomic_json_dump(payload: Dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + f".{RUN_ID}.tmp")
    with tmp.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2, default=str)
    tmp.replace(path)


def package_version(name: str) -> str:
    try:
        return importlib_metadata.version(name)
    except importlib_metadata.PackageNotFoundError:
        return "unknown"


def software_versions() -> Dict[str, str]:
    import os
    import platform
    versions = {
        "python": platform.python_version(),
        "spark": getattr(spark, "version", "unknown"),
        "databricks_runtime": os.getenv("DATABRICKS_RUNTIME_VERSION", "unknown"),
    }
    for pkg in ["pyspark", "pandas", "openpyxl", "delta-spark", "playwright"]:
        try:
            versions[pkg] = importlib_metadata.version(pkg)
        except Exception:
            versions[pkg] = "not_installed"
    try:
        versions["spark_databricks_version"] = spark.conf.get(
            "spark.databricks.clusterUsageTags.sparkVersion"
        )
    except Exception:
        versions["spark_databricks_version"] = "unavailable"
    return versions

titulo("SILVER CNA — INICIO")
print(f"Run ID                 : {RUN_ID}")
print(f"Pipeline version       : {PIPELINE_VERSION}")
print(f"Catálogo/Schema Silver : {CATALOG}.{SILVER_SCHEMA}")
print(f"Corte configurado      : {TARGET_DOCUMENT_YEAR}-{TARGET_DOCUMENT_MONTH:02d}" if TARGET_DOCUMENT_YEAR and TARGET_DOCUMENT_MONTH else "Corte configurado      : último período disponible")


# COMMAND ----------

# =============================================================================
# CELDA 2 — SELECCIÓN DESDE MANIFIESTO BRONZE Y VERIFICACIÓN SHA-256
# =============================================================================

MONTHS_ES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}


def normalize_for_match(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip().upper()


def infer_month_from_record(record: Dict[str, Any]) -> Optional[int]:
    combined = " ".join([
        str(record.get("original_filename") or ""),
        str(record.get("anchor_text") or ""),
        str(record.get("source_url") or ""),
    ])
    normalized = normalize_for_match(combined)
    for month_name, month_number in MONTHS_ES.items():
        if re.search(rf"\b{month_name}\b", normalized):
            return month_number
    return None


def load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


bronze_root = Path(f"/Volumes/{CATALOG}/{BRONZE_SCHEMA}/{BRONZE_VOLUME}")
metadata_dir = bronze_root / "_metadata"

if not metadata_dir.is_dir():
    raise FileNotFoundError(
        f"No existe el directorio de metadatos Bronze esperado: {metadata_dir}. "
        "Silver v2 requiere el contrato Bronze v2 (raw/ + _metadata/)."
    )

manifest_paths = sorted(metadata_dir.glob("ingestion_manifest_*_CNA.json"), reverse=True)
if not manifest_paths:
    raise FileNotFoundError(f"No se encontraron manifiestos CNA en {metadata_dir}")

compatible_manifest = None
compatible_manifest_path = None
for manifest_path in manifest_paths:
    candidate_manifest = load_json(manifest_path)
    if (
        candidate_manifest.get("source") == SOURCE_CODE
        and candidate_manifest.get("storage_layout_version") == EXPECTED_BRONZE_STORAGE_LAYOUT_VERSION
    ):
        compatible_manifest = candidate_manifest
        compatible_manifest_path = manifest_path
        break

if compatible_manifest is None:
    raise RuntimeError(
        "No existe un manifiesto CNA compatible con storage_layout_version=2.0. "
        "Ejecute primero el Bronze CNA corregido."
    )

bronze_batch_id = compatible_manifest.get("batch_id")
records = []
for record in compatible_manifest.get("files", []):
    if record.get("status") not in SUCCESS_BRONZE_STATUSES:
        continue
    original_filename = record.get("original_filename") or record.get("stored_filename") or ""
    extension = Path(original_filename).suffix.lower()
    if extension not in SUPPORTED_EXCEL_EXTENSIONS:
        continue
    if not record.get("stored_path") or not record.get("sha256") or not record.get("document_year"):
        continue
    enriched = dict(record)
    enriched["document_month"] = infer_month_from_record(record)
    records.append(enriched)

if not records:
    raise RuntimeError("El último manifiesto Bronze CNA compatible no contiene archivos Excel procesables.")

if TARGET_DOCUMENT_YEAR is not None and TARGET_DOCUMENT_MONTH is not None:
    period_records = [
        r for r in records
        if int(r["document_year"]) == int(TARGET_DOCUMENT_YEAR)
        and r.get("document_month") == int(TARGET_DOCUMENT_MONTH)
    ]
    if not period_records:
        available = sorted({(int(r["document_year"]), r.get("document_month")) for r in records})
        raise RuntimeError(
            f"No se encontró el período objetivo {TARGET_DOCUMENT_YEAR}-{TARGET_DOCUMENT_MONTH:02d}. "
            f"Períodos identificados: {available}"
        )
else:
    period_records = [r for r in records if r.get("document_month") is not None]
    if not period_records:
        raise RuntimeError("No fue posible inferir el mes documental de ningún archivo CNA.")
    latest_period = max((int(r["document_year"]), int(r["document_month"])) for r in period_records)
    period_records = [r for r in period_records if (int(r["document_year"]), int(r["document_month"])) == latest_period]

# Evita elegir silenciosamente entre contenidos distintos para el mismo período.
unique_hashes = {r["sha256"] for r in period_records}
if len(unique_hashes) > 1:
    detail = [(r.get("original_filename"), r.get("source_url"), r.get("sha256")) for r in period_records]
    raise RuntimeError(f"Existen múltiples contenidos distintos para el mismo período CNA: {detail}")

# Si las URLs son distintas pero el contenido es idéntico, la elección es determinista.
selected_record = sorted(period_records, key=lambda r: (r.get("source_url") or "", r.get("stored_path") or ""))[0]
source_path = Path(selected_record["stored_path"])

if not source_path.is_file():
    raise FileNotFoundError(f"El manifiesto Bronze referencia un archivo inexistente: {source_path}")

computed_sha256, source_size_bytes = sha256_file(source_path)
if computed_sha256 != selected_record["sha256"]:
    raise RuntimeError(
        "Fallo de integridad Bronze→Silver: SHA-256 del archivo físico no coincide con el manifiesto. "
        f"Esperado={selected_record['sha256']} | Obtenido={computed_sha256}"
    )

BRONZE_LINEAGE = {
    "manifest_path": str(compatible_manifest_path),
    "bronze_batch_id": bronze_batch_id,
    "source_url": selected_record.get("source_url"),
    "original_filename": selected_record.get("original_filename"),
    "stored_filename": selected_record.get("stored_filename"),
    "stored_path": str(source_path),
    "document_year": int(selected_record["document_year"]),
    "document_month": int(selected_record["document_month"]),
    "sha256": computed_sha256,
    "size_bytes": source_size_bytes,
}

RUTA_ARCHIVO_BRONZE = str(source_path)
REPORT_YEAR = BRONZE_LINEAGE["document_year"]
REPORT_MONTH = BRONZE_LINEAGE["document_month"]
REPORT_CUTOFF_DATE = f"{REPORT_YEAR:04d}-{REPORT_MONTH:02d}-01"

# La salida esperada del estudio está parametrizada respecto al mes seleccionado.
if REPORT_MONTH != 5:
    raise RuntimeError(
        f"El contrato tabular validado de este notebook corresponde al corte mayo; se seleccionó mes={REPORT_MONTH}. "
        "Actualice rangos/contratos antes de procesar un corte distinto."
    )

titulo("BRONZE CNA SELECCIONADO Y VERIFICADO")
print(f"Manifest       : {compatible_manifest_path.name}")
print(f"Bronze batch   : {bronze_batch_id}")
print(f"Archivo        : {BRONZE_LINEAGE['original_filename']}")
print(f"Ruta física    : {RUTA_ARCHIVO_BRONZE}")
print(f"Período        : {REPORT_YEAR}-{REPORT_MONTH:02d}")
print(f"Tamaño         : {source_size_bytes:,} bytes")
print(f"SHA-256        : {computed_sha256}")


# COMMAND ----------

# =============================================================================
# CELDA 3 — CONTRATOS TABULARES Y CONFIGURACIÓN VERSIONADA
# =============================================================================

@dataclass(frozen=True)
class TableContract:
    id_table: int
    name: str
    sheet: str
    header_range: str
    data_range: str
    semantic_columns: Tuple[str, ...]
    required_header_tokens: Tuple[str, ...]
    description: str


CONTRACTS: Dict[str, TableContract] = {
    "exportaciones_mensuales": TableContract(
        1, "exportaciones_mensuales", "RESUMEN", "AB9:AE9", "AB10:AE122",
        ("mes", "libras", "dolares", "precio_promedio"),
        ("MES", "LIBRAS", "DOLARES"),
        "Serie histórica mensual de libras, dólares y precio promedio."
    ),
    "exportaciones_mensuales_dolares": TableContract(
        2, "exportaciones_mensuales_dolares", "RESUMEN", "AH45:AU45", "AH46:AU78",
        ("anio", "enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre", "total"),
        ("ANO", "ENERO", "DICIEMBRE", "TOTAL"),
        "Serie anual de exportaciones en dólares con meses como columnas."
    ),
    "exportaciones_mayo_historico": TableContract(
        3, "exportaciones_mayo_historico", "RESUMEN", "AB127:AF128", "AB129:AF135",
        ("anio", "libras", "dolares", "libras_variacion", "dolares_variacion"),
        ("MAYO", "LIBRAS", "DOLARES"),
        "Comparativo histórico del mes de mayo."
    ),
    "exportaciones_acumuladas_historico": TableContract(
        4, "exportaciones_acumuladas_historico", "RESUMEN", "AB138:AF139", "AB140:AF146",
        ("periodo", "libras", "dolares", "libras_variacion", "dolares_variacion"),
        ("PERIODO", "LIBRAS", "DOLARES"),
        "Comparativo histórico acumulado enero-mayo."
    ),
    "exportaciones_mensuales_libras": TableContract(
        5, "exportaciones_mensuales_libras", "RESUMEN", "AH9:AW9", "AH10:AW42",
        ("anio", "enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre", "total", "precio_promedio_libra", "crecimiento_anual"),
        ("ANO", "ENERO", "DICIEMBRE", "TOTAL"),
        "Serie anual de exportaciones en libras con meses como columnas."
    ),
    "mercado_pais_mayo": TableContract(
        6, "mercado_pais_mayo", "MERCADO PAÍS", "A10:H11", "A12:H91",
        ("pais", "dolares_2025_05", "libras_2025_05", "dolares_2026_05", "libras_2026_05", "variacion_dolares", "variacion_libras", "participacion_libras"),
        ("PAIS", "DOLARES", "LIBRAS", "VARIACION"),
        "Distribución por mercado de destino para mayo de 2025 y 2026."
    ),
    "mercado_pais_acumulado": TableContract(
        7, "mercado_pais_acumulado", "MERCADO PAÍS ACUM", "A10:H11", "A12:H100",
        ("pais", "dolares_ene_may_2025", "libras_ene_may_2025", "dolares_ene_may_2026", "libras_ene_may_2026", "variacion_dolares", "variacion_libras", "participacion_libras"),
        ("PAIS", "DOLARES", "LIBRAS", "VARIACION"),
        "Distribución acumulada por mercado de destino enero-mayo."
    ),
}

if len(CONTRACTS) != 7:
    raise RuntimeError(f"Contrato Silver inválido: se esperaban 7 tablas y existen {len(CONTRACTS)}")
if sorted(c.id_table for c in CONTRACTS.values()) != list(range(1, 8)):
    raise RuntimeError("Los identificadores de contratos deben ser únicos y consecutivos entre 1 y 7.")

contract_payload = {k: asdict(v) for k, v in CONTRACTS.items()}
CONFIG_HASH = sha256_text(json.dumps(contract_payload, sort_keys=True, ensure_ascii=False))

titulo("CONTRATOS SILVER CNA")
for contract in sorted(CONTRACTS.values(), key=lambda x: x.id_table):
    print(f"{contract.id_table:02d}. {contract.name:42s} | {contract.sheet:18s} | {contract.data_range}")
print(f"Hash configuración: {CONFIG_HASH}")


# COMMAND ----------

# =============================================================================
# CELDA 4 — VALIDACIÓN FÍSICA DEL LIBRO EXCEL
# =============================================================================


def cells_from_range(ws, excel_range: str) -> List[Any]:
    min_col, min_row, max_col, max_row = range_boundaries(excel_range)
    return [ws.cell(row=r, column=c).value for r in range(min_row, max_row + 1) for c in range(min_col, max_col + 1)]


def dimensions_of_range(excel_range: str) -> Tuple[int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(excel_range)
    return max_row - min_row + 1, max_col - min_col + 1


def normalized_tokens(values: List[Any]) -> str:
    return " ".join(normalize_for_match(v) for v in values if v is not None)


validation_results = []
wb = load_workbook(RUTA_ARCHIVO_BRONZE, read_only=False, data_only=True)
try:
    titulo("VALIDACIÓN FÍSICA DE ESTRUCTURAS CNA")
    for contract in sorted(CONTRACTS.values(), key=lambda x: x.id_table):
        if contract.sheet not in wb.sheetnames:
            validation_results.append({"table": contract.name, "status": "FAIL", "detail": f"Hoja ausente: {contract.sheet}"})
            continue

        ws = wb[contract.sheet]
        header_rows, header_cols = dimensions_of_range(contract.header_range)
        data_rows, data_cols = dimensions_of_range(contract.data_range)

        if data_cols != len(contract.semantic_columns):
            validation_results.append({"table": contract.name, "status": "FAIL", "detail": f"Columnas rango={data_cols}, contrato={len(contract.semantic_columns)}"})
            continue

        header_values = cells_from_range(ws, contract.header_range)
        header_text = normalized_tokens(header_values)
        missing_tokens = [token for token in contract.required_header_tokens if normalize_for_match(token) not in header_text]

        data_values = cells_from_range(ws, contract.data_range)
        non_empty = sum(v is not None and str(v).strip() != "" for v in data_values)

        if missing_tokens:
            status, detail = "FAIL", f"Tokens de encabezado ausentes: {missing_tokens}"
        elif non_empty == 0:
            status, detail = "FAIL", "El rango de datos está completamente vacío."
        else:
            status, detail = "PASS", f"{data_rows} filas x {data_cols} columnas; {non_empty} celdas con contenido."

        validation_results.append({
            "table": contract.name,
            "sheet": contract.sheet,
            "header_range": contract.header_range,
            "data_range": contract.data_range,
            "status": status,
            "detail": detail,
        })
        print(f"{status:4s} | {contract.name:42s} | {detail}")
finally:
    wb.close()

PHYSICAL_VALIDATION_DF = pd.DataFrame(validation_results)
physical_failures = PHYSICAL_VALIDATION_DF[PHYSICAL_VALIDATION_DF["status"] == "FAIL"]
if not physical_failures.empty:
    raise RuntimeError("Validación física Silver fallida:\n" + physical_failures.to_string(index=False))

print("\n✓ Las 7 estructuras físicas cumplen el contrato configurado.")


# COMMAND ----------

# =============================================================================
# CELDA 5 — EXTRACCIÓN DE RANGOS Y MATERIALIZACIÓN INICIAL EN SPARK
# =============================================================================


def extract_data_range(ws, excel_range: str, columns: Tuple[str, ...]) -> pd.DataFrame:
    min_col, min_row, max_col, max_row = range_boundaries(excel_range)
    rows = [
        [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
        for r in range(min_row, max_row + 1)
    ]
    df = pd.DataFrame(rows, columns=list(columns))
    return df


def pandas_to_string_spark(df: pd.DataFrame) -> DataFrame:
    rows = []
    for row in df.itertuples(index=False, name=None):
        rows.append(tuple(None if pd.isna(v) else str(v) for v in row))
    schema = T.StructType([T.StructField(str(c), T.StringType(), True) for c in df.columns])
    return spark.createDataFrame(rows, schema=schema)


TABLAS_RAW_SPARK: Dict[str, DataFrame] = {}
EXTRACTION_METRICS: List[Dict[str, Any]] = []

wb = load_workbook(RUTA_ARCHIVO_BRONZE, read_only=False, data_only=True)
try:
    titulo("EXTRACCIÓN CNA → SPARK")
    for contract in sorted(CONTRACTS.values(), key=lambda x: x.id_table):
        pdf = extract_data_range(wb[contract.sheet], contract.data_range, contract.semantic_columns)
        sdf = pandas_to_string_spark(pdf)
        TABLAS_RAW_SPARK[contract.name] = sdf
        rows = sdf.count()
        EXTRACTION_METRICS.append({"table": contract.name, "rows_extracted": rows, "columns": len(sdf.columns)})
        print(f"PASS | {contract.name:42s} | {rows:4d} filas | {len(sdf.columns):2d} columnas")
finally:
    wb.close()

if set(TABLAS_RAW_SPARK) != set(CONTRACTS):
    raise RuntimeError("La extracción no produjo exactamente las siete tablas configuradas.")

EXTRACTION_METRICS_DF = pd.DataFrame(EXTRACTION_METRICS)
print("\n✓ Extracción finalizada. Desde este punto el procesamiento se realiza con Spark.")


# COMMAND ----------

# =============================================================================
# CELDA 6 — LIMPIEZA, NORMALIZACIÓN TEXTUAL Y DEDUPLICACIÓN EN SPARK
# =============================================================================

NULL_TOKENS = {"", "NULL", "NONE", "NAN", "N/A", "NA", "-"}
COUNTRY_ALIASES = {"EEUU": "ESTADOS UNIDOS", "USA": "ESTADOS UNIDOS", "US": "ESTADOS UNIDOS"}


def clean_text_expr(column_name: str):
    c = F.col(column_name)
    c = F.regexp_replace(c, "[\\u200b\\u200c\\u200d\\ufeff\\u00a0]", "")
    c = F.regexp_replace(c, r"\s+", " ")
    c = F.trim(c)
    return F.when(F.upper(c).isin(list(NULL_TOKENS)), F.lit(None)).otherwise(c)


def canonical_country_expr(column_name: str):
    cleaned = clean_text_expr(column_name)
    upper = F.upper(cleaned)
    normalized = F.translate(upper, "ÁÉÍÓÚÜÑ", "AEIOUUN")
    expr = normalized
    for alias, canonical in COUNTRY_ALIASES.items():
        expr = F.when(normalized == F.lit(alias), F.lit(canonical)).otherwise(expr)
    return expr


def remove_all_null_rows(df: DataFrame) -> DataFrame:
    condition = None
    for c in df.columns:
        term = F.col(c).isNotNull()
        condition = term if condition is None else (condition | term)
    return df.filter(condition) if condition is not None else df


TABLAS_LIMPIAS: Dict[str, DataFrame] = {}
CLEANING_AUDIT: List[Dict[str, Any]] = []

titulo("LIMPIEZA SILVER CNA")
for name, df in TABLAS_RAW_SPARK.items():
    initial_rows = df.count()

    cleaned = df.select(*[clean_text_expr(c).alias(c) for c in df.columns])
    if "pais" in cleaned.columns:
        cleaned = cleaned.withColumn("pais", canonical_country_expr("pais"))

    without_empty = remove_all_null_rows(cleaned)
    rows_after_empty = without_empty.count()

    dedup = without_empty.dropDuplicates()
    final_rows = dedup.count()

    TABLAS_LIMPIAS[name] = dedup
    audit = {
        "table": name,
        "rows_initial": initial_rows,
        "rows_removed_all_null": initial_rows - rows_after_empty,
        "duplicates_removed": rows_after_empty - final_rows,
        "rows_final": final_rows,
        "country_homologation": "pais" in dedup.columns,
    }
    CLEANING_AUDIT.append(audit)
    print(
        f"PASS | {name:42s} | inicial={initial_rows:4d} | "
        f"vacías={audit['rows_removed_all_null']:2d} | duplicados={audit['duplicates_removed']:2d} | final={final_rows:4d}"
    )

CLEANING_AUDIT_DF = pd.DataFrame(CLEANING_AUDIT)
print(f"\nDuplicados exactos eliminados: {int(CLEANING_AUDIT_DF['duplicates_removed'].sum())}")


# COMMAND ----------

# =============================================================================
# CELDA 7 — TIPADO CONTROLADO Y AUDITORÍA DE CONVERSIÓN
# =============================================================================

MONTH_COLUMNS = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

TYPE_CONTRACTS: Dict[str, Dict[str, str]] = {
    "exportaciones_mensuales": {"mes": "date", "libras": "double", "dolares": "double", "precio_promedio": "double"},
    "exportaciones_mensuales_dolares": {"anio": "int", **{m: "double" for m in MONTH_COLUMNS}, "total": "double"},
    "exportaciones_mayo_historico": {"anio": "int", "libras": "double", "dolares": "double", "libras_variacion": "double", "dolares_variacion": "double"},
    "exportaciones_acumuladas_historico": {"periodo": "string", "libras": "double", "dolares": "double", "libras_variacion": "double", "dolares_variacion": "double"},
    "exportaciones_mensuales_libras": {"anio": "int", **{m: "double" for m in MONTH_COLUMNS}, "total": "double", "precio_promedio_libra": "double", "crecimiento_anual": "double"},
    "mercado_pais_mayo": {"pais": "string", "dolares_2025_05": "double", "libras_2025_05": "double", "dolares_2026_05": "double", "libras_2026_05": "double", "variacion_dolares": "double", "variacion_libras": "double", "participacion_libras": "double"},
    "mercado_pais_acumulado": {"pais": "string", "dolares_ene_may_2025": "double", "libras_ene_may_2025": "double", "dolares_ene_may_2026": "double", "libras_ene_may_2026": "double", "variacion_dolares": "double", "variacion_libras": "double", "participacion_libras": "double"},
}


def numeric_expr(column_name: str):
    s = F.trim(F.col(column_name))
    # Admite números almacenados como texto y elimina símbolo de porcentaje/moneda si apareciera.
    s = F.regexp_replace(s, r"[%$]", "")
    # Caso miles con coma (1,234.56) vs decimal con coma (1234,56).
    s = F.when(s.rlike(r"^-?\d{1,3}(,\d{3})+(\.\d+)?$"), F.regexp_replace(s, ",", "")) \
         .otherwise(F.regexp_replace(s, ",", "."))
    return s.cast(T.DoubleType())


def date_expr(column_name: str):
    c = F.col(column_name)
    return F.coalesce(
        F.to_date(c, "yyyy-MM-dd HH:mm:ss"),
        F.to_date(c, "yyyy-MM-dd"),
        F.to_date(c, "dd/MM/yyyy"),
    )


TABLAS_TIPADAS: Dict[str, DataFrame] = {}
TYPE_AUDIT: List[Dict[str, Any]] = []
conversion_failures: List[Dict[str, Any]] = []

titulo("TIPADO CONTROLADO SILVER")
for table_name, df in TABLAS_LIMPIAS.items():
    contract = TYPE_CONTRACTS[table_name]
    if list(df.columns) != list(contract.keys()):
        raise RuntimeError(f"Contrato de columnas alterado en {table_name}: {df.columns}")

    out = df
    for column_name, target_type in contract.items():
        original_non_null = out.filter(F.col(column_name).isNotNull()).count()

        if target_type == "double":
            converted = numeric_expr(column_name)
        elif target_type == "int":
            converted = numeric_expr(column_name).cast(T.IntegerType())
        elif target_type == "date":
            converted = date_expr(column_name)
        elif target_type == "string":
            converted = F.col(column_name).cast(T.StringType())
        else:
            raise ValueError(f"Tipo no soportado: {target_type}")

        temp_name = f"__typed_{column_name}"
        out = out.withColumn(temp_name, converted)
        invalid_count = out.filter(F.col(column_name).isNotNull() & F.col(temp_name).isNull()).count()
        TYPE_AUDIT.append({
            "table": table_name,
            "column": column_name,
            "target_type": target_type,
            "source_non_null": original_non_null,
            "conversion_failures": invalid_count,
        })
        if invalid_count:
            conversion_failures.append({"table": table_name, "column": column_name, "count": invalid_count})
        out = out.drop(column_name).withColumnRenamed(temp_name, column_name)

    # Restaura orden del contrato.
    out = out.select(*contract.keys())
    TABLAS_TIPADAS[table_name] = out
    print(f"{'FAIL' if any(f['table']==table_name for f in conversion_failures) else 'PASS'} | {table_name:42s} | {out.count():4d} filas")

TYPE_AUDIT_DF = pd.DataFrame(TYPE_AUDIT)
if conversion_failures:
    raise RuntimeError(f"Se detectaron errores de conversión de tipos: {conversion_failures}")

print("\n✓ Todas las conversiones de tipos fueron determinísticas y sin pérdida por coerción.")


# COMMAND ----------

# =============================================================================
# CELDA 8 — NORMALIZACIÓN ANALÍTICA EN SPARK
# =============================================================================

MONTH_TO_NUMBER = {m: i for i, m in enumerate(MONTH_COLUMNS, start=1)}
NORMALIZATION_AUDIT: List[Dict[str, Any]] = []


def unpivot_months(df: DataFrame, value_name: str) -> DataFrame:
    structs = [
        F.struct(F.lit(num).alias("mes_numero"), F.col(month).alias("valor"))
        for month, num in MONTH_TO_NUMBER.items()
    ]
    out = (
        df.select("anio", F.explode(F.array(*structs)).alias("m"))
          .select(
              F.col("anio").cast("int").alias("anio"),
              F.col("m.mes_numero").cast("int").alias("mes_numero"),
              F.col("m.valor").cast("double").alias("valor"),
          )
          .withColumn("fecha_periodo", F.make_date("anio", "mes_numero", F.lit(1)))
          .select("fecha_periodo", "anio", "mes_numero", "valor")
    )
    # Excluye meses posteriores al corte documental: no son observaciones disponibles.
    cutoff = F.to_date(F.lit(REPORT_CUTOFF_DATE))
    out = out.filter(F.col("fecha_periodo") <= cutoff)
    before_null_filter = out.count()
    out = out.filter(F.col("valor").isNotNull())
    NORMALIZATION_AUDIT.append({
        "table": value_name,
        "transformation": "wide_to_long_months",
        "rows_before_null_filter": before_null_filter,
        "null_observations_excluded": before_null_filter - out.count(),
        "rows_final": out.count(),
    })
    return out


def unpivot_market(df: DataFrame, output_name: str) -> DataFrame:
    value_cols = [c for c in df.columns if c != "pais"]
    structs = [F.struct(F.lit(c).alias("indicador"), F.col(c).alias("valor")) for c in value_cols]
    out = (
        df.select("pais", F.explode(F.array(*structs)).alias("m"))
          .select("pais", F.col("m.indicador").alias("indicador"), F.col("m.valor").cast("double").alias("valor"))
          .filter(F.col("pais").isNotNull() & F.col("valor").isNotNull())
    )
    NORMALIZATION_AUDIT.append({
        "table": output_name,
        "transformation": "wide_to_long_indicator",
        "rows_final": out.count(),
    })
    return out


TABLAS_SILVER_FINAL: Dict[str, DataFrame] = {}

titulo("NORMALIZACIÓN ANALÍTICA")
TABLAS_SILVER_FINAL["exportaciones_mensuales_dolares_normalizada"] = unpivot_months(
    TABLAS_TIPADAS["exportaciones_mensuales_dolares"], "exportaciones_mensuales_dolares_normalizada"
)
TABLAS_SILVER_FINAL["exportaciones_mensuales_libras_normalizada"] = unpivot_months(
    TABLAS_TIPADAS["exportaciones_mensuales_libras"], "exportaciones_mensuales_libras_normalizada"
)
TABLAS_SILVER_FINAL["mercado_pais_mayo_normalizado"] = unpivot_market(
    TABLAS_TIPADAS["mercado_pais_mayo"], "mercado_pais_mayo_normalizado"
)
TABLAS_SILVER_FINAL["mercado_pais_acumulado_normalizado"] = unpivot_market(
    TABLAS_TIPADAS["mercado_pais_acumulado"], "mercado_pais_acumulado_normalizado"
)

monthly = (
    TABLAS_TIPADAS["exportaciones_mensuales"]
    .withColumnRenamed("mes", "fecha_periodo")
    .withColumn("anio", F.year("fecha_periodo"))
    .withColumn("mes_numero", F.month("fecha_periodo"))
    .filter(F.col("fecha_periodo") <= F.to_date(F.lit(REPORT_CUTOFF_DATE)))
    .select("fecha_periodo", "anio", "mes_numero", "libras", "dolares", "precio_promedio")
)
TABLAS_SILVER_FINAL["exportaciones_mensuales"] = monthly
TABLAS_SILVER_FINAL["exportaciones_mayo_historico"] = TABLAS_TIPADAS["exportaciones_mayo_historico"]
TABLAS_SILVER_FINAL["exportaciones_acumuladas_historico"] = TABLAS_TIPADAS["exportaciones_acumuladas_historico"]

EXPECTED_FINAL_TABLES = {
    "exportaciones_mensuales_dolares_normalizada",
    "exportaciones_mensuales_libras_normalizada",
    "mercado_pais_mayo_normalizado",
    "mercado_pais_acumulado_normalizado",
    "exportaciones_mensuales",
    "exportaciones_mayo_historico",
    "exportaciones_acumuladas_historico",
}
if set(TABLAS_SILVER_FINAL) != EXPECTED_FINAL_TABLES:
    raise RuntimeError("La normalización no produjo exactamente las siete tablas Silver esperadas.")

for name, df in TABLAS_SILVER_FINAL.items():
    print(f"PASS | {name:48s} | {df.count():4d} filas | {len(df.columns):2d} columnas")

NORMALIZATION_AUDIT_DF = pd.DataFrame(NORMALIZATION_AUDIT)


# COMMAND ----------

# =============================================================================
# CELDA 9 — REGLAS DE NEGOCIO Y CONTROLES MATEMÁTICOS EJECUTABLES
# =============================================================================

QUALITY_RESULTS: List[Dict[str, Any]] = []


def record_quality(rule: str, table: str, severity: str, violations: int, checked: int, detail: str) -> None:
    status = "PASS" if violations == 0 else ("FAIL" if severity == "CRITICAL" else "WARN")
    QUALITY_RESULTS.append({
        "rule": rule,
        "table": table,
        "severity": severity,
        "status": status,
        "checked": int(checked),
        "violations": int(violations),
        "detail": detail,
    })
    print(f"{status:4s} | {rule:28s} | {table:42s} | revisados={checked:4d} | violaciones={violations}")


def count_negative(df: DataFrame, columns: List[str]) -> Tuple[int, int]:
    checked = 0
    violations = 0
    for c in columns:
        checked += df.filter(F.col(c).isNotNull()).count()
        violations += df.filter(F.col(c) < 0).count()
    return checked, violations


def variation_violation_expr(current_col: str, previous_col: str, stored_col: str):
    prev = F.col(previous_col)
    curr = F.col(current_col)
    stored = F.col(stored_col)
    ratio = (curr - prev) / prev
    # Acepta que Excel almacene porcentajes como fracción (0.12) o como porcentaje (12.0).
    err_fraction = F.abs(stored - ratio)
    err_percent = F.abs(stored - ratio * F.lit(100.0))
    tolerance_fraction = F.greatest(F.lit(ABS_TOL_VARIATION / 100.0), F.abs(ratio) * F.lit(REL_TOL_VARIATION))
    tolerance_percent = F.greatest(F.lit(ABS_TOL_VARIATION), F.abs(ratio * 100.0) * F.lit(REL_TOL_VARIATION))
    return (err_fraction > tolerance_fraction) & (err_percent > tolerance_percent)


titulo("REGLAS DE NEGOCIO SILVER CNA")

# R1 — No negatividad para magnitudes que no admiten valores negativos.
for table, cols in {
    "exportaciones_mensuales": ["libras", "dolares", "precio_promedio"],
    "exportaciones_mensuales_dolares_normalizada": ["valor"],
    "exportaciones_mensuales_libras_normalizada": ["valor"],
}.items():
    checked, violations = count_negative(TABLAS_SILVER_FINAL[table], cols)
    record_quality("R1_NO_NEGATIVIDAD", table, "CRITICAL", violations, checked, "Volumen, valor monetario y precio deben ser >= 0.")

# R2 — Total anual = suma de meses disponibles.
for table in ["exportaciones_mensuales_dolares", "exportaciones_mensuales_libras"]:
    df = TABLAS_TIPADAS[table]
    month_sum = sum((F.coalesce(F.col(m), F.lit(0.0)) for m in MONTH_COLUMNS), F.lit(0.0))
    eligible = df.filter(F.col("total").isNotNull())
    tolerance = F.greatest(F.lit(ABS_TOL_TOTAL), F.abs(F.col("total")) * F.lit(REL_TOL))
    violations = eligible.filter(F.abs(F.col("total") - month_sum) > tolerance).count()
    record_quality("R2_TOTAL_MENSUAL", table, "CRITICAL", violations, eligible.count(), "El total debe coincidir con la suma de los meses disponibles dentro de tolerancia numérica.")

# R3 — Precio promedio reportado ≈ dólares / libras.
price_df = TABLAS_SILVER_FINAL["exportaciones_mensuales"].filter(
    F.col("libras").isNotNull() & (F.col("libras") > 0) & F.col("dolares").isNotNull() & F.col("precio_promedio").isNotNull()
)
calc_price = F.col("dolares") / F.col("libras")
price_tol = F.greatest(F.lit(ABS_TOL_PRICE), F.abs(F.col("precio_promedio")) * F.lit(REL_TOL_PRICE))
price_violations = price_df.filter(F.abs(F.col("precio_promedio") - calc_price) > price_tol).count()
record_quality("R3_PRECIO_PROMEDIO", "exportaciones_mensuales", "CRITICAL", price_violations, price_df.count(), "Precio promedio contrastado con dólares/libras; tolerancia 0.02 o 1%.")

# R4 — Variaciones interanuales históricas.
hist = TABLAS_TIPADAS["exportaciones_mayo_historico"]
w = Window.orderBy("anio")
hist = hist.withColumn("prev_libras", F.lag("libras").over(w)).withColumn("prev_dolares", F.lag("dolares").over(w))
for metric, prev, stored in [("libras", "prev_libras", "libras_variacion"), ("dolares", "prev_dolares", "dolares_variacion")]:
    eligible = hist.filter(F.col(prev).isNotNull() & (F.col(prev) != 0) & F.col(metric).isNotNull() & F.col(stored).isNotNull())
    violations = eligible.filter(variation_violation_expr(metric, prev, stored)).count()
    record_quality("R4_VARIACION_INTERANUAL", "exportaciones_mayo_historico", "CRITICAL", violations, eligible.count(), f"Variación {metric} contrastada contra el año anterior.")

# R4b — Variaciones en tablas de mercado (2025 vs 2026).
for table, pairs in {
    "mercado_pais_mayo": [
        ("dolares_2026_05", "dolares_2025_05", "variacion_dolares"),
        ("libras_2026_05", "libras_2025_05", "variacion_libras"),
    ],
    "mercado_pais_acumulado": [
        ("dolares_ene_may_2026", "dolares_ene_may_2025", "variacion_dolares"),
        ("libras_ene_may_2026", "libras_ene_may_2025", "variacion_libras"),
    ],
}.items():
    df = TABLAS_TIPADAS[table]
    for curr, prev, stored in pairs:
        eligible = df.filter(F.col(prev).isNotNull() & (F.col(prev) != 0) & F.col(curr).isNotNull() & F.col(stored).isNotNull())
        violations = eligible.filter(variation_violation_expr(curr, prev, stored)).count()
        record_quality("R4_VARIACION_MERCADO", table, "CRITICAL", violations, eligible.count(), f"{stored} contrastada con {curr} y {prev}.")

# R5 — Participación dentro de rango. Se detecta automáticamente si el origen usa fracción o porcentaje.
for table in ["mercado_pais_mayo", "mercado_pais_acumulado"]:
    df = TABLAS_TIPADAS[table].filter(F.col("participacion_libras").isNotNull())
    stats = df.agg(F.min("participacion_libras").alias("min_v"), F.max("participacion_libras").alias("max_v")).first()
    checked = df.count()
    if checked == 0:
        record_quality("R5_PARTICIPACION", table, "WARNING", 0, 0, "No hay valores de participación para validar.")
    else:
        max_v = float(stats["max_v"])
        upper = 1.000001 if max_v <= 1.000001 else 100.0001
        violations = df.filter((F.col("participacion_libras") < 0) | (F.col("participacion_libras") > upper)).count()
        scale = "fracción [0,1]" if upper <= 1.1 else "porcentaje [0,100]"
        record_quality("R5_PARTICIPACION", table, "CRITICAL", violations, checked, f"Escala detectada: {scale}.")

# R6 — Integridad temporal y unicidad de claves analíticas.
for table in ["exportaciones_mensuales", "exportaciones_mensuales_dolares_normalizada", "exportaciones_mensuales_libras_normalizada"]:
    df = TABLAS_SILVER_FINAL[table]
    duplicate_keys = df.groupBy("anio", "mes_numero").count().filter(F.col("count") > 1).count()
    invalid_periods = df.filter(
        F.col("anio").isNull() | F.col("mes_numero").isNull() |
        (F.col("mes_numero") < 1) | (F.col("mes_numero") > 12) |
        (F.col("fecha_periodo") > F.to_date(F.lit(REPORT_CUTOFF_DATE)))
    ).count()
    record_quality("R6_CLAVE_TEMPORAL", table, "CRITICAL", duplicate_keys + invalid_periods, df.count(), "Clave (anio, mes_numero) única y período dentro del corte documental.")

for table in ["mercado_pais_mayo_normalizado", "mercado_pais_acumulado_normalizado"]:
    df = TABLAS_SILVER_FINAL[table]
    duplicate_keys = df.groupBy("pais", "indicador").count().filter(F.col("count") > 1).count()
    record_quality("R6_CLAVE_MERCADO", table, "CRITICAL", duplicate_keys, df.count(), "Clave (pais, indicador) debe ser única.")

QUALITY_RESULTS_DF = pd.DataFrame(QUALITY_RESULTS)
critical_failures = QUALITY_RESULTS_DF[(QUALITY_RESULTS_DF["severity"] == "CRITICAL") & (QUALITY_RESULTS_DF["status"] == "FAIL")]

print("\nResumen de reglas:")
print(QUALITY_RESULTS_DF[["rule", "table", "status", "checked", "violations"]].to_string(index=False))

if not critical_failures.empty:
    raise RuntimeError("Reglas críticas de negocio incumplidas. Silver no será persistido:\n" + critical_failures.to_string(index=False))

print("\n✓ Todas las reglas críticas fueron superadas.")


# COMMAND ----------

# =============================================================================
# CELDA 10 — MÉTRICAS DE CALIDAD Y TRAZABILIDAD DE TRANSFORMACIÓN
# =============================================================================

SILVER_TABLE_METRICS: List[Dict[str, Any]] = []

titulo("MÉTRICAS SILVER ANTES DE PERSISTENCIA")
for name, df in TABLAS_SILVER_FINAL.items():
    rows = df.count()
    null_exprs = [F.sum(F.when(F.col(c).isNull(), 1).otherwise(0)).alias(c) for c in df.columns]
    null_row = df.agg(*null_exprs).first().asDict() if rows > 0 else {c: 0 for c in df.columns}
    nulls = int(sum(int(v or 0) for v in null_row.values()))
    metric = {
        "table": name,
        "rows": rows,
        "columns": len(df.columns),
        "null_cells": nulls,
        "schema": {f.name: f.dataType.simpleString() for f in df.schema.fields},
    }
    SILVER_TABLE_METRICS.append(metric)
    print(f"{name:48s} | filas={rows:4d} | columnas={len(df.columns):2d} | nulos={nulls:3d}")

SILVER_TABLE_METRICS_DF = pd.DataFrame([{k: v for k, v in m.items() if k != "schema"} for m in SILVER_TABLE_METRICS])

TRANSFORMATION_SUMMARY = {
    "run_id": RUN_ID,
    "pipeline_version": PIPELINE_VERSION,
    "config_hash": CONFIG_HASH,
    "bronze_lineage": BRONZE_LINEAGE,
    "extraction": EXTRACTION_METRICS,
    "cleaning": CLEANING_AUDIT,
    "normalization": NORMALIZATION_AUDIT,
    "quality_rules": QUALITY_RESULTS,
    "silver_tables": SILVER_TABLE_METRICS,
}

print("\n✓ Resumen de transformación generado sin estimaciones artificiales de tamaño o duración.")


# COMMAND ----------

# =============================================================================
# CELDA 11 — PERSISTENCIA DELTA EN UNITY CATALOG
# =============================================================================

spark.sql(f"CREATE SCHEMA IF NOT EXISTS `{CATALOG}`.`{SILVER_SCHEMA}`")

PERSISTENCE_RESULTS: List[Dict[str, Any]] = []
TABLE_REFS: Dict[str, str] = {}

titulo("PERSISTENCIA SILVER EN UNITY CATALOG")
for name, df in TABLAS_SILVER_FINAL.items():
    table_ref = f"`{CATALOG}`.`{SILVER_SCHEMA}`.`{name}`"
    table_ref_plain = f"{CATALOG}.{SILVER_SCHEMA}.{name}"

    (
        df.write.format("delta")
          .mode("overwrite")
          .option("overwriteSchema", "true")
          .saveAsTable(table_ref_plain)
    )

    # Lineage sin alterar el esquema analítico de la tabla.
    props = {
        "pipeline.layer": "silver",
        "pipeline.source": SOURCE_CODE,
        "pipeline.version": PIPELINE_VERSION,
        "pipeline.silver_run_id": RUN_ID,
        "lineage.bronze_batch_id": str(BRONZE_LINEAGE.get("bronze_batch_id") or ""),
        "lineage.bronze_sha256": BRONZE_LINEAGE["sha256"],
        "lineage.bronze_original_filename": str(BRONZE_LINEAGE.get("original_filename") or ""),
        "lineage.report_period": f"{REPORT_YEAR:04d}-{REPORT_MONTH:02d}",
    }
    props_sql = ", ".join([f"'{k}'='{str(v).replace(chr(39), chr(39)*2)}'" for k, v in props.items()])
    spark.sql(f"ALTER TABLE {table_ref} SET TBLPROPERTIES ({props_sql})")

    TABLE_REFS[name] = table_ref_plain
    persisted_rows = spark.table(table_ref_plain).count()
    PERSISTENCE_RESULTS.append({"table": name, "table_ref": table_ref_plain, "rows": persisted_rows, "status": "WRITTEN"})
    print(f"PASS | {table_ref_plain:80s} | {persisted_rows:4d} filas")

if len(TABLE_REFS) != 7:
    raise RuntimeError("No se persistieron exactamente siete tablas Silver.")

PERSISTENCE_RESULTS_DF = pd.DataFrame(PERSISTENCE_RESULTS)


# COMMAND ----------

# =============================================================================
# CELDA 12 — VALIDACIÓN POST-ESCRITURA: ESQUEMA, FILAS Y EQUIVALENCIA DE DATOS
# =============================================================================

POSTWRITE_RESULTS: List[Dict[str, Any]] = []

titulo("VALIDACIÓN POST-ESCRITURA")
for name, source_df in TABLAS_SILVER_FINAL.items():
    table_ref = TABLE_REFS[name]
    persisted_df = spark.table(table_ref)

    source_count = source_df.count()
    persisted_count = persisted_df.count()
    same_columns = source_df.columns == persisted_df.columns
    same_schema = [f.dataType.simpleString() for f in source_df.schema.fields] == [f.dataType.simpleString() for f in persisted_df.schema.fields]

    source_minus_persisted = source_df.exceptAll(persisted_df).limit(1).count()
    persisted_minus_source = persisted_df.exceptAll(source_df).limit(1).count()
    same_data = source_minus_persisted == 0 and persisted_minus_source == 0

    detail = spark.sql(f"DESCRIBE DETAIL `{CATALOG}`.`{SILVER_SCHEMA}`.`{name}`").first().asDict()
    is_delta = str(detail.get("format", "")).lower() == "delta"

    passed = source_count == persisted_count and same_columns and same_schema and same_data and is_delta
    POSTWRITE_RESULTS.append({
        "table": name,
        "source_rows": source_count,
        "persisted_rows": persisted_count,
        "same_columns": same_columns,
        "same_schema": same_schema,
        "same_data": same_data,
        "is_delta": is_delta,
        "status": "PASS" if passed else "FAIL",
    })
    print(f"{'PASS' if passed else 'FAIL'} | {name:48s} | filas={persisted_count:4d} | data_equal={same_data} | delta={is_delta}")

POSTWRITE_RESULTS_DF = pd.DataFrame(POSTWRITE_RESULTS)
postwrite_failures = POSTWRITE_RESULTS_DF[POSTWRITE_RESULTS_DF["status"] == "FAIL"]
if not postwrite_failures.empty:
    raise RuntimeError("Validación post-escritura fallida:\n" + postwrite_failures.to_string(index=False))

print("\n✓ Persistencia verificada por equivalencia bidireccional con exceptAll().")


# COMMAND ----------

# =============================================================================
# CELDA 13 — MANTENIMIENTO DELTA CON CRITERIO OBJETIVO
# =============================================================================

MAINTENANCE_RESULTS: List[Dict[str, Any]] = []

# Z-Order solo se define para una futura escala donde sea justificable.
ZORDER_COLUMNS = {
    "exportaciones_mensuales_dolares_normalizada": "anio",
    "exportaciones_mensuales_libras_normalizada": "anio",
    "mercado_pais_mayo_normalizado": "pais",
    "mercado_pais_acumulado_normalizado": "pais",
    "exportaciones_mensuales": "anio",
    "exportaciones_mayo_historico": "anio",
    "exportaciones_acumuladas_historico": "periodo",
}

titulo("EVALUACIÓN DE MANTENIMIENTO DELTA")
for name, table_ref in TABLE_REFS.items():
    detail = spark.sql(f"DESCRIBE DETAIL `{CATALOG}`.`{SILVER_SCHEMA}`.`{name}`").first().asDict()
    num_files = int(detail.get("numFiles") or 0)
    size_bytes = int(detail.get("sizeInBytes") or 0)
    size_mb = size_bytes / (1024 * 1024)

    should_optimize = ENABLE_DELTA_MAINTENANCE and (num_files >= OPTIMIZE_MIN_FILES or size_mb >= OPTIMIZE_MIN_SIZE_MB)
    if should_optimize:
        zcol = ZORDER_COLUMNS[name]
        if zcol not in spark.table(table_ref).columns:
            raise RuntimeError(f"Columna ZORDER configurada inexistente: {name}.{zcol}")
        spark.sql(f"OPTIMIZE `{CATALOG}`.`{SILVER_SCHEMA}`.`{name}` ZORDER BY (`{zcol}`)")
        action = f"OPTIMIZE_ZORDER({zcol})"
    else:
        action = "SKIPPED_SMALL_TABLE"

    MAINTENANCE_RESULTS.append({
        "table": name,
        "num_files_before": num_files,
        "size_mb_before": size_mb,
        "action": action,
    })
    print(f"{name:48s} | archivos={num_files:2d} | {size_mb:8.4f} MB | {action}")

MAINTENANCE_RESULTS_DF = pd.DataFrame(MAINTENANCE_RESULTS)
print("\n✓ No se atribuye mejora de rendimiento a OPTIMIZE si la operación fue omitida o no fue medida experimentalmente.")


# COMMAND ----------

# =============================================================================
# CELDA 14 — SUITE TÉCNICA DE QA SOBRE LAS TABLAS PERSISTIDAS
# =============================================================================

QA_RESULTS: List[Dict[str, Any]] = []

EXPECTED_COLUMNS = {name: TABLAS_SILVER_FINAL[name].columns for name in TABLAS_SILVER_FINAL}
EXPECTED_ROWS = {name: TABLAS_SILVER_FINAL[name].count() for name in TABLAS_SILVER_FINAL}


def qa_record(table: str, test: str, passed: bool, detail: str) -> None:
    QA_RESULTS.append({"table": table, "test": test, "status": "PASS" if passed else "FAIL", "detail": detail})


titulo("QA TÉCNICO SILVER CNA")
for name, table_ref in TABLE_REFS.items():
    try:
        df = spark.table(table_ref)
        qa_record(name, "accessibility", True, "spark.table() exitoso")
    except Exception as exc:
        qa_record(name, "accessibility", False, str(exc))
        continue

    row_count = df.count()
    qa_record(name, "row_count", row_count == EXPECTED_ROWS[name], f"esperado={EXPECTED_ROWS[name]}, obtenido={row_count}")
    qa_record(name, "columns", df.columns == EXPECTED_COLUMNS[name], f"esperado={EXPECTED_COLUMNS[name]}, obtenido={df.columns}")

    duplicate_rows = df.groupBy(*df.columns).count().filter(F.col("count") > 1).count() if df.columns else 0
    qa_record(name, "exact_duplicates", duplicate_rows == 0, f"grupos duplicados={duplicate_rows}")

    props = spark.sql(f"SHOW TBLPROPERTIES `{CATALOG}`.`{SILVER_SCHEMA}`.`{name}`").collect()
    props_dict = {r[0]: r[1] for r in props}
    lineage_ok = (
        props_dict.get("pipeline.silver_run_id") == RUN_ID
        and props_dict.get("lineage.bronze_sha256") == BRONZE_LINEAGE["sha256"]
    )
    qa_record(name, "lineage_properties", lineage_ok, "propiedades de lineage verificadas")

QA_RESULTS_DF = pd.DataFrame(QA_RESULTS)
qa_failures = QA_RESULTS_DF[QA_RESULTS_DF["status"] == "FAIL"]

print(QA_RESULTS_DF.groupby(["status"]).size().to_string())
print(f"Tests ejecutados: {len(QA_RESULTS_DF)}")
print(f"Tests fallidos  : {len(qa_failures)}")

if not qa_failures.empty:
    raise RuntimeError("La suite técnica QA detectó fallos:\n" + qa_failures.to_string(index=False))

print("\n✓ Suite QA técnica superada. Esto verifica integridad operacional; no se interpreta como '100% de calidad' del dominio.")


# COMMAND ----------

# =============================================================================
# CELDA 15 — MANIFIESTO SILVER PERSISTENTE Y DURACIÓN REAL
# =============================================================================

PIPELINE_FINISHED_AT_UTC = datetime.now(timezone.utc)
PIPELINE_DURATION_SECONDS = time.perf_counter() - PIPELINE_STARTED_MONOTONIC

spark.sql(f"CREATE VOLUME IF NOT EXISTS `{CATALOG}`.`{SILVER_SCHEMA}`.`{SILVER_METADATA_VOLUME}`")
metadata_root = Path(f"/Volumes/{CATALOG}/{SILVER_SCHEMA}/{SILVER_METADATA_VOLUME}")
metadata_root.mkdir(parents=True, exist_ok=True)

manifest = {
    "manifest_schema_version": SILVER_MANIFEST_SCHEMA_VERSION,
    "pipeline_version": PIPELINE_VERSION,
    "run_id": RUN_ID,
    "source": SOURCE_CODE,
    "started_at_utc": PIPELINE_STARTED_AT_UTC.isoformat(),
    "finished_at_utc": PIPELINE_FINISHED_AT_UTC.isoformat(),
    "duration_seconds": round(PIPELINE_DURATION_SECONDS, 6),
    "report_period": f"{REPORT_YEAR:04d}-{REPORT_MONTH:02d}",
    "config_hash": CONFIG_HASH,
    "bronze_lineage": BRONZE_LINEAGE,
    "physical_validation": validation_results,
    "extraction": EXTRACTION_METRICS,
    "cleaning": CLEANING_AUDIT,
    "type_audit": TYPE_AUDIT,
    "normalization": NORMALIZATION_AUDIT,
    "business_quality": QUALITY_RESULTS,
    "silver_tables": SILVER_TABLE_METRICS,
    "persistence": PERSISTENCE_RESULTS,
    "postwrite_validation": POSTWRITE_RESULTS,
    "delta_maintenance": MAINTENANCE_RESULTS,
    "technical_qa": QA_RESULTS,
    "software_versions": software_versions(),
    "status": "SUCCESS",
}

SILVER_MANIFEST_PATH = metadata_root / f"silver_manifest_{RUN_ID}_CNA.json"
atomic_json_dump(manifest, SILVER_MANIFEST_PATH)

# Verificación inmediata del manifiesto persistido.
reloaded_manifest = load_json(SILVER_MANIFEST_PATH)
if reloaded_manifest.get("run_id") != RUN_ID or reloaded_manifest.get("status") != "SUCCESS":
    raise RuntimeError("El manifiesto Silver no pudo verificarse después de su escritura.")

manifest_sha256, manifest_size = sha256_file(SILVER_MANIFEST_PATH)
print(f"Manifiesto Silver : {SILVER_MANIFEST_PATH}")
print(f"SHA-256 manifiesto: {manifest_sha256}")
print(f"Tamaño manifiesto : {manifest_size:,} bytes")
print(f"Duración real      : {PIPELINE_DURATION_SECONDS:.3f} s")


# COMMAND ----------

# =============================================================================
# CELDA 16 — REPORTE FINAL DEL PIPELINE SILVER CNA
# =============================================================================

titulo("REPORTE FINAL — SILVER CNA")

print(f"Estado                  : SUCCESS")
print(f"Run ID                  : {RUN_ID}")
print(f"Período fuente          : {REPORT_YEAR}-{REPORT_MONTH:02d}")
print(f"Bronze batch            : {BRONZE_LINEAGE['bronze_batch_id']}")
print(f"Archivo Bronze          : {BRONZE_LINEAGE['original_filename']}")
print(f"SHA-256 fuente          : {BRONZE_LINEAGE['sha256']}")
print(f"Duración medida         : {PIPELINE_DURATION_SECONDS:.3f} segundos")
print(f"Tablas Silver           : {len(TABLE_REFS)}")
print(f"Filas materializadas*   : {sum(m['rows'] for m in SILVER_TABLE_METRICS):,}")
print(f"Reglas ejecutadas       : {len(QUALITY_RESULTS)}")
print(f"Reglas críticas fallidas: {len(critical_failures)}")
print(f"Tests técnicos QA       : {len(QA_RESULTS_DF)}")
print(f"Tests técnicos fallidos : {len(qa_failures)}")
print(f"Manifiesto              : {SILVER_MANIFEST_PATH}")

print("\nTablas persistidas:")
for name, ref in TABLE_REFS.items():
    rows = next(m["rows"] for m in SILVER_TABLE_METRICS if m["table"] == name)
    print(f"  - {ref} | {rows:,} filas")

print("\n* 'Filas materializadas' es la suma de filas de siete tablas diferentes; no representa observaciones independientes.")
print("\n✓ PIPELINE SILVER CNA COMPLETADO Y VALIDADO")


# COMMAND ----------
